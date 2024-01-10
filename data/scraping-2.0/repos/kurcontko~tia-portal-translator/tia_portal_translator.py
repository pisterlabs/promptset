import os
import multiprocessing as mp
import argparse
from openpyxl import load_workbook
from googletrans import Translator  # use 3.1.0a0 or later
import openai
from deepl import Translator as DeepLTranslator
import textwrap

my_excel = 'TIAProjectTexts.xlsx'
my_excel_sheet_name = 'User Texts'
n_processes = min(os.cpu_count(), 64) #64 is maximum number in Windows, you can try to push the no of processes to the limits, but it can hit your system's stability
result_excel = f'{my_excel[:-5]}_translated.xlsx'

def parse_arguments():
    parser = argparse.ArgumentParser(description='Translate an Excel file using Google Translate, GPT, or DeepL.')
    parser.add_argument('--service', choices=['google', 'gpt', 'deepl'], required=True, help='Choose the translation service (google, gpt, or deepl).')
    parser.add_argument('--source', required=True, help='Source language and region (e.g., en-US, fr-FR, de-DE).')
    parser.add_argument('--dest', required=True, help='Destination language and region (e.g., en-US, fr-FR, de-DE).')
    args = parser.parse_args()
    return args

class TranslationService:
    def __init__(self, api_key=None, destination_language=None):
        self.api_key = api_key
        self.destination_language = destination_language

    def translate(self, text):
        pass

class GoogleTranslationService(TranslationService):
    def translate(self, text):
        translator = Translator()
        return translator.translate(text, dest=self.destination_language).text

class GPTTranslationService(TranslationService):
    def translate(self, text):
        openai.api_key = self.api_key
        prompt = f'Translate the following text to "{self.destination_language}" language:\n{text}'
        response = openai.Completion.create(
            engine='text-davinci-002',
            prompt=prompt,
            max_tokens=100,
            n=1,
            stop=None,
            temperature=0.5,
        )
        return response.choices[0].text.strip()

class DeepLTranslationService(TranslationService):
    def translate(self, text):
        translator = DeepLTranslator(self.api_key)
        return translator.translate_text(text, target_lang=self.destination_language)

def translation_service_factory(service, api_key=None, destination_language=None):
    if service == 'google':
        return GoogleTranslationService(api_key, destination_language)
    elif service == 'gpt':
        return GPTTranslationService(api_key, destination_language)
    elif service == 'deepl':
        return DeepLTranslationService(api_key, destination_language)
    else:
        raise ValueError(f'Invalid service: {service}')
    
def process_frame(chunk_tuple, translator_instance, ws, destination_to_translation_col):
    index, chunk = chunk_tuple
    print(f'Translating chunk {index+1}...')
    translated_chunk = []
    for cell in chunk:
        if cell.value:
            # Split the source text into lines
            source_lines = cell.value.split('\n')

            # Translate each line separately
            translated_lines = [translator_instance.translate(line) for line in source_lines]

            # Try to re-wrap the translated lines to match the source lines
            wrapped_translated_lines = []
            for source_line, translated_line in zip(source_lines, translated_lines):
                if len(source_line) > 0:  # Only apply textwrap if the line is not empty
                    wrapped_translated_line = textwrap.wrap(translated_line, width=len(source_line))
                    wrapped_translated_lines.extend(wrapped_translated_line)
                else:
                    wrapped_translated_lines.append('')  # Append an empty line if the source line was empty

            # Join the lines back together
            cell_translation = '\n'.join(wrapped_translated_lines)
            translated_chunk.append(cell_translation)
        else:
            translated_chunk.append(ws[destination_to_translation_col][index].value)
    return index, translated_chunk

def find_column_letter(column_name, ws):
    for cell in ws[1]:
        if cell.value == column_name:
            return cell.column_letter
    return None

if __name__ == '__main__':
    try:

        # Parse command-line arguments
        args = parse_arguments()
        translator_service = args.service
        source_to_translation = args.source
        destination_to_translation = args.dest

        # Extract the destination language from the column name
        destination_language = destination_to_translation.split('-')[0]

        # Check if the API key is required and available
        api_key = None
        if translator_service == 'gpt' or translator_service == 'deepl':
            api_key_env_var = 'OPENAI_API_KEY' if translator_service == 'gpt' else 'DEEPL_API_KEY'
            try:
                api_key = os.environ[api_key_env_var]
            except KeyError:
                print(f'Error: {translator_service.upper()} translation requires the {api_key_env_var} environment variable.')
                exit(1)

        # Read Excel file
        wb = load_workbook(my_excel)
        ws = wb[my_excel_sheet_name]

        # Find column letters for source and destination
        source_to_translation_col = find_column_letter(source_to_translation, ws)
        destination_to_translation_col = find_column_letter(destination_to_translation, ws)

        if not source_to_translation_col or not destination_to_translation_col:
            print('Could not find column names.')
            exit(1)

        # Split data into chunks
        row_count = ws.max_row - 1
        chunk_size = row_count // n_processes
        data_chunks = [(i, ws[source_to_translation_col][i*chunk_size+1:(i+1)*chunk_size+1]) for i in range(n_processes)]

        # Instantiate the translator
        translator_instance = translation_service_factory(translator_service, api_key, destination_language)

        # Use multiprocessing to translate chunks
        pool = mp.Pool(n_processes)
        result_list = pool.starmap(process_frame, [(chunk_tuple, translator_instance, ws, destination_to_translation_col) for chunk_tuple in data_chunks])
        pool.close()
        pool.join()

        # Sort results by index
        result_list.sort(key=lambda x: x[0])

        # Write translations back to the worksheet
        for index, chunk in result_list:
            for idx, cell in enumerate(chunk):
                ws[f'{destination_to_translation_col}{index * chunk_size + idx + 2}'].value = cell

        # Save the workbook
        if os.path.exists(result_excel):
            print(f'Removed file {result_excel}')
        wb.save(result_excel)
        print(f'Created new file {result_excel}')
        print('Translating finished!')

    except Exception as e:
        print(f'An error occurred: {e}')
