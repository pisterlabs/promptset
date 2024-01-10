import sys
import openai
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import os
import keyfile
openai.api_key = keyfile.OpenAikey
import re
import random
import requests
from time import sleep
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

failed_generations = 0

# Set the step-by-step mode based on the command-line argument
if len(sys.argv) > 1 and sys.argv[1] == "step-by-step":
    step_by_step = True
else:
    step_by_step = False



def update_story_log(story_name, num_chapters, current_chapter, status):
    file_path = "story_log.xlsx"

    # Load the workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    # Select the first sheet
    sheet = workbook.active

    # Write the headers if they don't exist
    if sheet.cell(row=1, column=1).value is None:
        headers = ["Story Name", "Open Time", "Num Chapters", "Current Chapter", "Time Completed", "Status"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

    # Find the next available row to write the data
    row_idx = sheet.max_row + 1

    # Write the data to the cells
    sheet.cell(row=row_idx, column=1).value = story_name
    # Write other information as needed

    # Save the workbook
    workbook.save(file_path)

def promptGPT(prompt_text):
    while True:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt_text},
            ],
            temperature=0.9,
            max_tokens=2300,
            top_p=1,
            frequency_penalty=0,
            presence_penalty=0,
            n=1,
            stop=["\nUser:"],
        )
        content = response["choices"][0]["message"]["content"]
        print("Response Received.")
        if "That model is currently overloaded with other requests. " in content or "error" in content:
            # Phrase found in the response, retry after 60 seconds
            print("Model overloaded. Retrying after 60 seconds...")
            sleep(60)
        else:
            # Valid response, return the content
            return content

def promptGPT_bak(prompt_text):
    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": prompt_text},
        ],
        temperature=0.9,
        max_tokens=2300,
        top_p=1,
        frequency_penalty=0,
        presence_penalty=0,
        n=1,
        stop=["\nUser:"],
    )
    content = response["choices"][0]["message"]["content"]
    if "That model is currently overloaded with other requests." in content or "error" in content:
        # Phrase found in the response, retry after 60 seconds
        print("Model overloaded. Retrying after 60 seconds...")
        sleep(60)
    else:
        # Valid response, return the content
        return content

    #print(response["choices"][0]["message"]["content"])
    #return response["choices"][0]["message"]["content"]

def BookTitle():
    keywords = promptGPT('please generate 3 keywords to use as inspiration for a fun, interesting, an not in anyway average story. Think outside the box, please do not provide generic keywors. Please ONLY PROVIDE 3 KEYWORDS, your response shoul be only 3 words, nothing else')
    print(keywords)
    prompt_text = "provide an  interesting, an not in anyway average - think outside the box - story book title for an illustrated children's adventure book. please make sure the response ends with nothing else but 'Title:' and the title of the book, an example: Title: Dinosaurs vs Aliens: One Brontosaur's battle for the planet. Use the following keywords as inspiration: " + keywords
    title = None
    if not step_by_step:
        string = promptGPT(prompt_text)
        print("Riding Solo: please wait for your book.")

        title = string[len('Title: '):]
        print(title)
        return title
    else:
        print("Manual Mode: Generating Title.")
        approved = False
        while not approved:
            string = promptGPT(prompt_text)

            title = string[len('Title: '):]
            print("Title: " + title)
            approved = input("Do you approve? (yes/no) ")
            approved = approved.lower() == "yes"

        return title

def DefineChapters(title):
    file_path = 'chapterprompt.txt'  # Replace with the actual path to your file
    chapters = {}

    # Open the file in read mode
    with open(file_path, 'r') as file:
        file_content = file.read()

    prompt_text = file_content + title

    if not step_by_step:
        string = promptGPT(prompt_text)
        print("Riding Solo: please wait for your chapters.")
        chapters_list = string.split('Chapter ')[1:]  # Split the string into chapters
        print(f"chapter list: {chapters_list}")
        for i, chapter_text in enumerate(chapters_list):
            chapter_number = i + 1
            chapter_key = f"{chapter_number}"
            _, chapter_text = chapter_text.split(':', 1)  # Split at the first colon
            chapters[chapter_key] = chapter_text.strip()  # Remove leading/trailing spaces
        print(chapters)
        return chapters
    else:
        print("Manual Mode: Generating Chapters.")
        approved = False
        while not approved:
            string = promptGPT(prompt_text)

            chapters_list = string.split('Chapter ')[1:]  # Split the string into chapters
            for i, chapter_text in enumerate(chapters_list):
                chapter_number = i + 1
                chapter_key = f"Chapter {chapter_number}"
                _, chapter_text = chapter_text.split(':', 1)  # Split at the first colon
                chapters[chapter_key] = chapter_text.strip()  # Remove leading/trailing spaces

            print(chapters)
            approved = input("Do you approve? (yes/no) ")
            approved = approved.lower() == "yes"
        print(chapters)
        return chapters

def WriteChapters(chapters, title):
    prompt_text = "use the following 'Chapter Descriptions' to write content of the current chapter. I will give you a description of the previous chapter, as 'previous chapter: ' then the current chapter as 'current chapter' and a description of the next chapter as 'next chapter:' use these descriptions to write a coherent current chapter that incorporates the previous chapter and leads into the next chapter. Do Not Include any information about the previous or next chapters, that information is strickly for your use, you should only present the current chapter in your response, but that chapter should be written taking into account the previous and next chapters. Here are the descriptions: "
    title = re.sub(r'[^\w\s-]', '', title).replace(" ", "_")
    chapter_items = list(chapters.items())

    folder_path = f"chapters\\{title}"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    for i, (chapter_number, chapter_text) in enumerate(chapter_items):
        file_name = f"{chapter_number}.txt"
        previous_chapter_text = chapter_items[i-1][1] if i > 0 else ""
        current_chapter_text = chapter_text
        next_chapter_text = chapter_items[i+1][1] if i < len(chapter_items)-1 else ""
        approved = False
        if step_by_step:
            while not approved:
                chapter_content = promptGPT(f"{prompt_text} previous chapter: {previous_chapter_text}, current chapter: {current_chapter_text}, next chapter: {next_chapter_text}")
                print(f"Chapter {chapter_number}:\n{chapter_content}")
                approved = input(f"Do you approve of Chapter {chapter_number}? (yes/no) ").lower() == "yes"

        else:
            chapter_content = promptGPT(f"{prompt_text} previous chapter: {previous_chapter_text}, current chapter: {current_chapter_text}, next chapter: {next_chapter_text}")
            print(f"Chapter {chapter_number}:\n{chapter_content}")


        # Write chapter text to a file


        file_path = f"{folder_path}/{file_name}"
        with open(file_path, 'w') as file:
            file.write(chapter_content)



        make_images(chapter_content)
        sleep(30)
        download_images(title)
    return True

def make_images(chapter_content):
    global failed_generations
    file_path = "LeoGen.txt"
    with open(file_path, 'r') as file:
        prompt_text = file.read()

    chapter_moments = promptGPT(f"Please provide 3 or 4 moments from this text that would make good images for a kids book, label them as 'moments' {chapter_content}")
    print(chapter_moments)

    prompt_text = promptGPT(prompt_text + chapter_moments)
    print("")
    print(prompt_text)
    # Regular expressions to extract prompts and negatives
    prompt_pattern = r"Prompt \d+: \"(.+?)\""
    negative_pattern = r"Negative \d+: \"(.+?)\""

    # Extract prompts and negatives using regular expressions
    prompts = re.findall(r"Prompt \d+:(.+)", prompt_text)
    negatives = re.findall(r"Negative \d+:(.+)", prompt_text)

    # Print the prompts and negatives
    print("Prompts:")
    for prompt in prompts:
        print(prompt)

    print("Negatives:")
    for negative in negatives:
        print(negative)



    guidance_scale = random.randint(5, 20)
    ids = [
        "6bef9f1b-29cb-40c7-b9df-32b51c1f67d3",
        "cd2b2a15-9760-4174-a5ff-4d2925057376",
        "291be633-cb24-434f-898f-e662799936ad"
    ]

    random_id = random.choice(ids)
    random_prompt = random.randint(0, len(prompts) - 1)

    print(random_id)
    url = "https://cloud.leonardo.ai/api/rest/v1/generations"
    print("Generating Image Prompt")

    print()
    print("::Generating Image::")
    payload = {
        "prompt": f"{prompts[random_prompt]}",
        "negative_prompt": f"((grainy)) ((pixelated)) {negatives[random_prompt]}",
        "modelId": random_id,
        "width": 1024,
        "height": 1024,
        "public": True,
        "sd_version": "v2",
        "num_images": 4,
        # need to make these 3 variable
        "guidance_scale": guidance_scale,
        "presetStyle": "LEONARDO",
        "promptMagic": True,
        "num_inference_steps": guidance_scale,
        "scheduler": "LEONARDO",

    }
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "authorization": "Bearer " + keyfile.LeoKey
    }
    attempts = 0
    httpresponse = False
    while not httpresponse and attempts < 10:
        response = requests.post(url, json=payload, headers=headers)
        print(response)
        print(f"HTTP Response: {response.status_code}")
        if response.status_code == 200:
            httpresponse = True
            sleep(30)
            #print(download_images(title))

            return prompts[random_prompt]
        else:
            attempts +=1
            sleep(30)
    if attempts == 10:
        print("Sorry, the image was not generated, Leonardo is giving us the following status_code: " + str(response.status_code))
        print(response)
        failed_generations +=1

    return prompts[random_prompt]


def generate_text(prompt_text):

    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content":
                "You are a helpful assistant."},
            {"role": "user", "content": prompt_text},
        ],
        temperature=0.5,
        max_tokens=1500,
        top_p=1,
        frequency_penalty=0,
        presence_penalty=0,
        n=1,
        stop=["\nUser:"],
    )
    print("Image Prompts Generated: ")
    print(response["choices"][0]["message"]["content"])
    print("Extracting and Organizing Prompts")
    # Parse the response to extract the prompts
    prompts = re.findall(r"Prompt \d+:(.+)", response["choices"][0]["message"]["content"])
    negPrompts = re.findall(r"Negative \d+:(.+)", response["choices"][0]["message"]["content"])

    print("Choosing a Random Prompt.")
    prompts = [p.strip() for p in prompts]
    print("prompts")
    print(prompts)
    print("negPrompts")
    print(negPrompts)
    if len(prompts) > 0 and len(negPrompts) > 0:
         random_index = random.randint(0, len(prompts) - 1)
    else:
        # Handle the case when prompts is empty
        print("no prompts generated: Likely GPT server error.!")
        print(prompts)
        print(negPrompts)
        random_index = random.randint(0, len(prompts) - 1)
        return"", "", ""
    print("Chosen Prompts")
    print("prompt: " + prompts[random_index])
    print("NegPrompt: " + negPrompts[random_index])
    # Choose a random prompt and return it
    GenKeyWords = None#PKeyword1 + "_" + PKeyword2 + "_" + PKeyword3
    return prompts[random_index], negPrompts[random_index], GenKeyWords


def download_images(title):
    url = f"https://cloud.leonardo.ai/api/rest/v1/generations/user/73c05e39-d703-4464-ac1a-aa493227f9fe?offset=0&limit=1"
    headers = {
        "accept": "application/json",
        "authorization": "Bearer " + keyfile.LeoKey
    }

    response = requests.get(url, headers=headers)
    data = response.json()

    generated_images = data['generations'][0]['generated_images']
    prompt_id = data['generations'][0]['id']
    image_name = data['generations'][0]['prompt']
    print(image_name)

    folder_name = f"chapters/{title}/images/{prompt_id}"

    # Create the folder if it doesn't exist
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)

    downloaded_images = []
    for image in generated_images:
        url = image['url']
        if url.endswith('.jpg'):
            filename = url.split("/")[-1]
            print("Downloading Image:", filename)
            filepath = os.path.join(folder_name, filename)

            # Download and save the image to the folder
            response = requests.get(url)
            with open(filepath, 'wb') as f:
                f.write(response.content)

            downloaded_images.append(filepath)
            print("Images downloaded to:", folder_name)

            # Return a random image file path
            if downloaded_images:
                random_image = random.choice(downloaded_images)
                return random_image
            else:
                return None






title = BookTitle()
chapters = DefineChapters(title)
sleep(30)
approved = WriteChapters(chapters, title)
print(chapters)
print(f"number of failed generations: {failed_generations}")

#print(approved)