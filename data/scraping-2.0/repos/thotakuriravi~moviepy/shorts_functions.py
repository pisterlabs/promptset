import uuid
import openpyxl as xl
import random
import shorts_path as path
import openai
from moviepy.editor import *
import os
import textwrap
from gtts import gTTS


openai.api_key = path.openai_api_key

def uniq_id():
    id = str(uuid.uuid4()).replace('-', '')
    # print(id)
    return id


def topic_pick():
    # read data to the excel file using python openpyxl
    
    workbook = xl.load_workbook(path.data_file)
    workbook_active = workbook.active
    # print(workbook_active.max_column)
    # print(workbook_active.max_row) # data in between 1, 3

    index_of_row = random.randint(1, workbook_active.max_row)
    # print(index_of_row)

    topic = workbook_active.cell(row= index_of_row, column= 1 ).value
    
    return index_of_row, topic


def chatGPT(topic):
    
    chat_completion = openai.ChatCompletion.create(model="gpt-3.5-turbo", 
                                               messages=[{"role": "user", 
                                                          "content": f"explain {topic}"}])


    result = chat_completion['choices'][0]['message']['content']
    
    result = result.replace('\n\n', '\n')

    final_lines = []
    for line in result.splitlines():
        if '. ' in line:
            new_line = line.split('. ', 1)[1]
            final_lines.append(new_line)
        else:
            final_lines.append(line)

    # final_lines
    
    return final_lines



def gtts_text_to_audio(text, id):
    tts = gTTS(text)
    file_name = 'output/' + id + '.mp3'
    tts.save(file_name)
    
    return file_name






def title_intro(topic, id):
    
    # Background Video
    bg_video = VideoFileClip( path.bg_videos_path + random.choice(os.listdir(path=path.bg_videos_path)) )
    
    # Background Image
    
    # bg_video = ImageClip(path.bg_images_path + random.choice(os.listdir(path.bg_images_path)) )
    
    title_txt = '\n'.join(textwrap.wrap(topic, width=15))   
    text_audio_file = gtts_text_to_audio(text=topic, id=id+'_title')
    
    title_audio = AudioFileClip(text_audio_file)
    
    title = TextClip(txt=title_txt,
                     fontsize= 150,
                     font='melton',
                     color= random.choice(path.color_codes),
                     stroke_color=random.choice(path.stroke_colors),
                     stroke_width=2).set_audio(title_audio).set_position(('center', 'center'))
    
    title1 = TextClip(txt=title_txt,
                     fontsize= 50,
                     font='melton',
                     color= random.choice(path.color_codes),
                     stroke_color=random.choice(path.stroke_colors),
                     stroke_width=2).set_duration(1).set_position(('center', 'center'))
    
    title2 = TextClip(txt=title_txt,
                     fontsize= 100,
                     font='melton',
                     color= random.choice(path.color_codes),
                     stroke_color=random.choice(path.stroke_colors),
                     stroke_width=2).set_duration(1).set_position(('center', 'center'))
    
    title3 = TextClip(txt=title_txt,
                     fontsize= 150,
                     font='melton',
                     color= random.choice(path.color_codes),
                     stroke_color=random.choice(path.stroke_colors),
                     stroke_width=2).set_duration(1).set_position(('center', 'center'))
    
    bg_audio = AudioFileClip(path.title_musiq + random.choice(os.listdir(path.title_musiq)) ).set_duration(3)
    
    bg_video = bg_video.set_audio(bg_audio)
    
    def get_txt_color(t):
        if t < 0.5:
            return random.choice(path.color_codes)
        elif t < 1:
            return '#a32372'
        elif t < 1.5:
            return random.choice(path.color_codes)
        elif t < 2:
            return '#008080'
        elif t < 2.5:
            return random.choice(path.color_codes)
        elif t < 3:
            return 'red'
        else:
            return random.choice(path.color_codes)
    
    def modified_title_clip(get_frame, t):
        modified_clip = TextClip(txt=title_txt,
                        fontsize= 150,
                        font='melton',
                        color= get_txt_color(t),
                        stroke_color=random.choice(path.stroke_colors),
                        stroke_width=2).set_position(('center', 'center'))
        return modified_clip.get_frame(t)
    
    modified_title = title.fl(modified_title_clip)
    
    if modified_title.w > 1080:
        ratio = 1070 / modified_title.w
        new_size = ( int(modified_title.w * ratio), int(modified_title.h * ratio))
        
        modified_title = modified_title.resize(new_size)
        title3 = title3.resize(new_size)
      
    
    duration = 3 + title_audio.duration
    
    transition_video = VideoFileClip('shorts_gallery\\transition.gif', has_mask=True).subclip(0,1)
    
    
    final_video = CompositeVideoClip(clips=[bg_video, 
                                            title1, 
                                            title2.set_start(1), 
                                            title3.set_start(2), 
                                            modified_title.set_start(3),
                                            transition_video.set_start(duration)
                                            ]).set_duration(duration + 1)
    
    # file_name = 'final_videos/' + topic.replace(' ', '_').replace('\'', '') + '.mp4'

    # final_video.write_videofile(file_name, fps= 30, codec='libx264', logger = None)
    
    return final_video
    




def topic_clips(text, id, i):
    
    # Background Image
    
    bg_image = ImageClip(path.bg_images_path + random.choice(os.listdir(path.bg_images_path)) ).resize((1080,1920))
    
    title_txt = '\n'.join(textwrap.wrap(text, width=30))   
    text_audio_file = gtts_text_to_audio(text=text, id=id + f'_clip_{i}')
    
    title_audio = AudioFileClip(text_audio_file)
    
    title = TextClip(txt=title_txt,
                     fontsize= 70,
                     font='subtitle',
                     color= random.choice(path.color_codes),
                     stroke_color=random.choice(path.stroke_colors),
                     stroke_width=2).set_audio(title_audio).set_position(lambda t: ('center', 1900-100*t))
    
    ending_transition_video = VideoFileClip('shorts_gallery\\transition.gif', has_mask=True).subclip(0,1).resize((1080,1920))
    starting_transition_video = VideoFileClip('shorts_gallery\\transition.gif', has_mask=True).subclip(1,2).resize((1080,1920))
    
    duration = 1 + title_audio.duration
    
    final_video = CompositeVideoClip([bg_image, 
                                      starting_transition_video, 
                                      title.set_start(1), 
                                      ending_transition_video.set_start(duration)]).set_duration(duration+1)
    
    

    # final_video.write_videofile('output/shorts_clip.mp4', fps= 30, codec='libx264', logger = None)
    
    return final_video
    
    
    
