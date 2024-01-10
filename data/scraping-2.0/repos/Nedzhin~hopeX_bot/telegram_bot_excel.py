import logging
from telegram import ReplyKeyboardMarkup, ReplyKeyboardRemove, Update
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

import constants
# Enable logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)


from langchain.chat_models import ChatOpenAI
from langchain.document_loaders import DirectoryLoader
from langchain.indexes import VectorstoreIndexCreator
from langchain.chains import ConversationalRetrievalChain
from langchain.memory import ConversationSummaryMemory
import openai
import os

os.environ['OPENAI_API_KEY']= constants.API_KEY
from openai import OpenAI

client = OpenAI(
  api_key=os.environ.get("CUSTOM_ENV_NAME"),
)

#excell materials
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl

excel_file_path = 'question_answer.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
sheet_name = 'Sheet1'  # Replace with the actual sheet name
save_sheet = workbook.active

# loading for the pdf retireval task
loader = DirectoryLoader("documents", glob="*.txt")
index = VectorstoreIndexCreator().from_loaders([loader])



def get_spreadsheet():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name(
        'hopeX.json', scope
    )
    gc = gspread.authorize(credentials)
    spreadsheet = gc.open("HopeX")  # Replace with your spreadsheet name
    return spreadsheet


def read_last_row(spreadsheet):
    # Get the last row in the sheet
    sheet = spreadsheet.sheet1

    all_values = sheet.get_all_values()

    if not all_values:
        return None

    # Get the last column
    last_column = [row[-1] for row in all_values]

    return last_column

#Function to generate the OpenAi response to the Doctors entered information
def generate_openai_response(question: str) -> str:
    # Use the OpenAI API to generate a response
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",  # or choose another model
        messages=[
            {"role": "system", "content": "You are a assistant at fatigue topics, skilled in explaining complex fatigue concepts with treating recommendations for the fatigue and other categories."},
            {"role": "user", "content": question},
        ]
    )

    return str(response.choices[0].message.content)

sheet = get_spreadsheet()
current_day = read_last_row(sheet)[2]

# making the current_day as the global variable
# async def modify_global_variable(update, context):
#     # Use the global keyword to modify the global variable
#     sheet = get_spreadsheet()
#     global current_day
#     current_day = read_last_row(sheet)[2]
#     print(current_day)
#     update.message.reply_text(f"Global Variable Modified: {current_day}")

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Starts the conversation and introduces itself."""

    await context.bot.send_message(chat_id=update.effective_chat.id,
        text="""Hi! I am HopeX bot. I am professional at fatigue. Ask any question in this field. I will try to answer!
      Send /cancel to stop talking to me.\n\n 
      Your user ID: {}""".format(update.effective_chat.id))


async def bot_reply(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Returns the reply to user after getting reply from server."""
    user = update.message.from_user
    logger.info("Question from User: %s", update.message.text)
    if update.message.text != '':
        
        llm_reply = index.query(update.message.text, llm=ChatOpenAI(model="gpt-3.5-turbo"))
      

        # Add a new question and answer
        new_question = update.message.text
        new_answer = llm_reply

        # Assuming your questions and answers start from the second row (row index 2)
        next_row = save_sheet.max_row + 1

        # Write the question and answer to the sheet
        save_sheet.cell(row=next_row, column=1, value=new_question)
        save_sheet.cell(row=next_row, column=2, value=new_answer)

        # Save the changes to the Excel file
        workbook.save(excel_file_path)
    else:
        return 

    await update.message.reply_text(llm_reply)


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancels and ends the conversation."""
    user = update.message.from_user
    logger.info("User %s canceled the conversation.", user.first_name)
    await update.message.reply_text(
        "Bye! I hope we can talk again some day."
    )

def get_chat_id_from_spreadsheet():
    spreadsheet = get_spreadsheet()
    chat_id = spreadsheet.sheet1.acell('A1').value
    return chat_id

async def check_spreadsheet_changes(context):
    """Check for changes in the spreadsheet and send updates to the chat."""
    global current_day
    try:
        chat_id = get_chat_id_from_spreadsheet()
        if chat_id:
            sheet = get_spreadsheet()
            new_row_data = read_last_row(sheet)
            if new_row_data[2] != current_day:
              current_day = new_row_data[2]
              orig_message = f"You scores from the Doctor: \n Fatigue: {new_row_data[3]}\n Stress: {new_row_data[4]}\n Sleep qlty: {new_row_data[5]}\n Pain: {new_row_data[6]}\n Stiffness: {new_row_data[7]}\n Swelling: {new_row_data[8]}\n"
              
              await context.bot.send_message(chat_id=chat_id, text=
                                             orig_message)
              message = generate_openai_response(orig_message+"Give the recommendations to improve this scores")
              await context.bot.send_message(chat_id=chat_id, text=
                                        message)
              
    except Exception as e:
        logger.error(f"An error occurred while checking spreadsheet changes: {e}")



def main() -> None:
    """Run the bot."""
    # Create the Application and pass it your bot's token.
    application = Application.builder().token(constants.Telegram_BOT_TOKEN).build()

    # Add conversation handler with the states GENDER, PHOTO, LOCATION and BIO
    application.add_handler(CommandHandler("start", start))
    #application.add_handler(CommandHandler("modify", modify_global_variable))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND,bot_reply))
    application.add_handler(CommandHandler("cancel", cancel))

    
    application.job_queue.run_repeating(
        check_spreadsheet_changes, interval=15, first=0
    )
    # Run the bot until the user presses Ctrl-C
    application.run_polling()

import asyncio
if __name__ == "__main__":
    asyncio.run(main())