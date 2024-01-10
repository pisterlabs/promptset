# script revision for processing prime factorization calls to the ChatGPT API

import os
import math
from time import sleep
import openai
from openpyxl import load_workbook
    
openai.api_key="API_KEY"
# openai.api_key = os.getenv("OPENAI_API_KEY")


def check(num, response):
    if str(math.factorial(num)) in response:
        return True
    else:
        return False

# replaces '!' in the question with num
def process_question(msg, num):
    msg = msg.replace('!', str(10**num))
    return msg
    
    
# perform a request
def gpt_request(message):
    while True:
        try:
            sleep(0.1)
            response = openai.ChatCompletion.create(model="gpt-3.5-turbo", messages=[{"role": "system", "content": "please output just the numerical answer"}, {"role": "user", "content": str(message)}])
            return str(response['choices'][0]['message']['content'])
        except:
            print("Exception thrown, sleeping for 2 seconds...")
            sleep(2)
            print("Trying API call again...")

def main(file):

    messages = [] # messages grabbed from file to send to API
    new = []
    correct = []
    gpt_responses = []

    excel = load_workbook("questions.xlsx") # name of file with responses
    sheet = excel.active
    
    # read messages from file
    columns = sheet["A"]
    for data in columns:
        messages.append(data.value)

    # read recursion level
    columns = sheet["B"]
    for data in columns:
        recursion = int(data.value)

    # read correct responses

    #columns = sheet["C"]
    #for data in columns:
    #   correct = int(data.value)
        
    # get responses from chatGPT
    for message in messages:
        for num in range(recursion):
            num += 1
            new.append(process_question(message, num))
            print("posing question " + str(num) + " out of " + str(recursion) + ": " + new[num-1])
            gpt_responses.append(gpt_request(new[num-1]))
            
    # write results to file
    sheet.cell(row=1, column=1).value="Inputted Value"
    sheet.cell(row=1, column=2).value="ChatGPT Response"
    sheet.cell(row=1, column=2).value="Correct Answer"


    index = 0
    for x in gpt_responses:
        sheet.cell(row=index+2, column=1).value=str(new[index])
        sheet.cell(row=index+2, column=2).value=str(gpt_responses[index])

        index += 1    

    # cleanup
    excel.save("output" + str(file) + ".xlsx")

if __name__ == '__main__':
    for x in range(10):
        print ("Beginning run number " + str(x+1))
        main(x)

    print("!!!SUCCESS!!!")
