import asyncio
import csv
import openai
import openai_async
import os


def processCSV(filename, query):
    import csv

    # Read the CSV file.
    csv_data = ""
    with open(filename, "r") as input_file:
        csvFile = csv.reader(input_file)
        for row in csvFile:
            csv_data += '\n' + ','.join(row)

    prompt = f"""
Given the following CSV file, answer this question: "{query}"

Data:
{csv_data}
"""

    return prompt
    

def processExcel(filename, query):

    import openpyxl

    # Load the workbook
    wb = openpyxl.load_workbook(filename, data_only=False)

    # Load value workbook
    wb_value = openpyxl.load_workbook(filename, data_only=True)

    result_str = ""

    # Iterate over all sheets
    for sheet in wb:
        sheet_name = sheet.title
        result_str += (f'Sheet "{sheet_name}":\n')
        sheet_values = wb_value[sheet_name]
        for row, row_values in zip(sheet.iter_rows(), sheet_values.iter_rows()):
            for cell, cell_value in zip(row, row_values):
                if cell.data_type == 'f':  # 'f' stands for formula content
                    result_str += (f'\t({cell.coordinate}, {cell.value}, {cell_value.value})\n')
                else:
                    if cell_value.value != None:
                        result_str += (f'\t({cell.coordinate}, {cell_value.value})\n')


    prompt = f"""
    Given the following information about an Excel spreadsheet, where each
    sheet starts with "Sheet: (sheet name)" and then is followed by the
    contents of the sheet in the form (cell, formula, cell value) or (cell, cell value),
    answer the following query: "{query}".

    Excel information:
    {result_str}
    """

    return prompt


def precheck():
    import os
    if not 'OPENAI_API_KEY' in os.environ:
        print('You need a valid OpenAI key to use ChatDBG. You can get a key here: https://openai.com/api/')
        print('Set the environment variable OPENAI_API_KEY to your key value.')
        return False
    return True


async def chat(user_prompt, model="gpt-3.5-turbo"):
    import asyncio
    import httpx
    import openai
    import openai_async
    import os
    try:
        completion = await openai_async.chat_complete(openai.api_key, timeout=30, payload={'model': model, 'messages': [{'role': 'user', 'content': user_prompt}]})
        json_payload = completion.json()
        text = json_payload['choices'][0]['message']['content']
    except (openai.error.AuthenticationError, httpx.LocalProtocolError, KeyError):
        # Something went wrong.
        print()
        print('You need a valid OpenAI key to use ChatDBG. You can get a key here: https://openai.com/api/')
        print('Set the environment variable OPENAI_API_KEY to your key value.')
        import sys
        sys.exit(1)
    except Exception as e:
        print(f'EXCEPTION {e}, {type(e)}')
        pass
    return text

def main():

    if not precheck():
        return
    
    import argparse
    import os

    #create parser
    parser = argparse.ArgumentParser()

    #add arguments
    parser.add_argument("--file", help="Input file to be processed.", required=True)
    parser.add_argument("--query", help="Query to be executed", required=True)

    #parse arguments
    args = parser.parse_args()

    # check if the file exists
    if not os.path.isfile(args.file):
        print(f"The file {args.file} does not exist.")
        return
            
    #identify file type and call appropriate function
    extension = os.path.splitext(args.file)[1]
    if extension == '.csv':
        user_prompt = processCSV(args.file, args.query)
    elif extension == '.xlsx':
        user_prompt = processExcel(args.file, args.query)
    elif extension == '.txt':
        pass
        # user_prompt = processText(args.file, args.query)
    else:
        print("Unrecognized file type. The file must be a .csv, .xlsx or .txt file")

    text = asyncio.run(chat(user_prompt))
    print(text)

if __name__ == "__main__":
    main()
