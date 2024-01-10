from PyPDF2 import PdfReader
from langchain.text_splitter import CharacterTextSplitter
from langchain.embeddings import OpenAIEmbeddings
from langchain.vectorstores import FAISS
import os
import openai
from dotenv import load_dotenv, find_dotenv
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from langchain.chat_models import ChatOpenAI
from langchain.document_loaders import WebBaseLoader
import re
import openpyxl
import pandas as pd
_ = load_dotenv(find_dotenv())
openai.api_key = os.getenv("OPENAI_API_KEY")

def get_pdf_text(pdf_docs):
    """
    PDF documents are converted to text format
    input: PDF documents (list of file paths)
    output: Concatenated text from all PDF pages
    """
    text = ""
    for pdf in pdf_docs:
        pdf_reader = PdfReader(pdf)
        for page in pdf_reader.pages:
            text += page.extract_text()
    return text

def get_excel_text(excel_files):
    """
    Excel files are read, and text from all cells is extracted and concatenated.
    input: Excel files (list of file paths)
    output: Concatenated text from all Excel files and all cells
    """
    text = ""
    for excel_file in excel_files:
        wb = openpyxl.load_workbook(excel_file)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        text += str(cell.value) + " "
    return text

def detect_file_type(file_path):
    if str(file_path).lower().endswith('.pdf'):
        file_type = 'pdf'
    elif str(file_path).lower().endswith('.csv'):
        file_type = 'csv'
    elif str(file_path).lower().endswith(('.xlsx', '.xls')):
        file_type = 'excel'
    else:
        file_type = 'other'
    return file_type
    
def get_text_from_file(files):
    """
    Function to handle PDF, CSV, and Excel files based on their file type.
    input: File path of the uploaded file
    output: Text extracted from the file
    """
    text = ""
    for file in files:
        type_file = detect_file_type(file.name)
        if type_file == 'pdf':
            text += get_pdf_text([file])
        elif type_file == 'csv': 
              df = pd.read_csv(file)
              all_text = df.to_string(index=False,header=False)
              text+=re.sub(r'\s+', ' ', all_text)
        elif type_file == 'excel':
            text += get_excel_text([file])
        else:
            print(f"File type not supported,{file.name}") 

        
    return text 

def get_text_chunks(raw_text,chunk_size=1000,chunk_overlap=200):
    """
    divides the given text into chunks
    """
    splitter = CharacterTextSplitter(
        separator="\n",
        chunk_size =chunk_size,
        chunk_overlap = chunk_overlap,
        length_function = len
        )
    chunks = splitter.split_text(raw_text)
    return chunks

def get_vectorstore(chunks):
    embeddings = OpenAIEmbeddings()
    vectorstore = FAISS.from_texts(texts=chunks,embedding=embeddings)

    return vectorstore 

def get_conversation_chain(vectorstore):
    llm = ChatOpenAI()
    memory = ConversationBufferMemory(
        memory_key='chat_history',
        return_messages=True
    )
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm,
        retriever=vectorstore.as_retriever(),
        memory=memory
    )
    return conversation_chain 


def load_url(links):
    text = ""
    for link in links:
        link = str(link)
        loader = WebBaseLoader(link)
        data = loader.load()
        for i in range(len(data)):
            text+=data[i].page_content
    return str(text)