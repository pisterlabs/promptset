# coding=utf-8
__author__ = 'stefano'
import logging
from pprint import pprint
from datetime import datetime
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from openpyxl import load_workbook, Workbook
from django.core.management.base import BaseCommand
from openaid.projects.models import Project, Activity, Initiative, NewProject


class Command(BaseCommand):

    help = 'export file with new projects associated with initiatives '
    logger = logging.getLogger('openaid')
    special_cases = {
        "9377 (Fondo Esperti e Fondo in Loco) - F.ROT/AID 99/009/01":"009377",
        "0010521":"010521",
    }


    def get_code(self,new_project):
        # gets correct 6 chars code from np.number
        code = new_project.number.strip()

        # deals with hand picked special cases
        if code in self.special_cases.keys():
            return self.special_cases[code]

        code = code.replace("AID ","")
        if code is None or code == '':
            return code
        code_split_slash = code.split("/")
        code_split_dot = code.split(".")

        if len(code_split_slash)>1:
            if len(code_split_slash[0]) <=6:
                code = code_split_slash[0]
            else:
                return -1
        elif len(code_split_dot)>1:
            if len(code_split_dot[0]) <=6:
                code = code_split_dot[0]
            else:
                return -1
        elif len(code) <= 6:
            code = code
        else:
            # malformed code
            return -1
        # check that the string contains only numbers
        if code.isdigit():
            return code.zfill(6)
        else:
            return -1


    def handle(self, *args, **options):
        verbosity = options['verbosity']
        if verbosity == '0':
            self.logger.setLevel(logging.ERROR)
        elif verbosity == '1':
            self.logger.setLevel(logging.WARNING)
        elif verbosity == '2':
            self.logger.setLevel(logging.INFO)
        elif verbosity == '3':
            self.logger.setLevel(logging.DEBUG)

        output_filename= 'newproj_init.xlsx'
        counters={'null':0, 'malformed':0,'existing':0, 'not_existing':0}
        newproject_fields = ['number', 'title_it', 'title_en','description_it','description_en', 'recipient']
        initiative_fields = ['code', 'title_it', 'title_en','description_temp_it','description_temp_en', 'recipient_temp']
        np_to_fix = []
        workbook = Workbook()
        ws_output = workbook.create_sheet(index=0, title='sheet')
        self.logger.info(u"start")

        # append headers to output file
        headers = ["new_project_"+x for x in newproject_fields]
        headers.append("new_project_photos")
        headers.extend(["init_"+x for x in initiative_fields])
        headers.append("init_photos")

        ws_output.append(headers)

        for new_project in NewProject.objects.all().order_by('number'):
            initiative = None
            code = self.get_code(new_project)
            if code is not None and code != '' and code != -1:
                try:
                    initiative = Initiative.objects.get(code=code)
                except ObjectDoesNotExist:
                    self.logger.info(u"Initiative with code:'{}' does NOT exist".format(code))
                    counters['not_existing'] += 1
                    np_to_fix.append(new_project.pk)
                    continue
                else:
                    self.logger.debug(u"Initiative with code:'{}' exist".format(code))
                    counters['existing'] += 1

            else:
                if code == None or code == '':
                    self.logger.error(u"NewProj:{} has code None or ''. Skip".format(new_project.pk))
                    counters['null'] += 1
                    np_to_fix.append(new_project.pk)
                elif code == -1:
                    self.logger.error(u"NewProj:{} - Code '{}' is malformed, cannot process. Skip".format(new_project.pk, new_project.number))
                    counters['malformed'] += 1
                    np_to_fix.append(new_project.pk)
                continue

            row = []
            # gets data from NP
            for f in newproject_fields:
                value = getattr(new_project, f)
                if f == 'recipient' and value is not None:
                    row.append(value.name)
                else:
                    row.append(value)

            # has photos: true/false
            if new_project.photo_set.all().count() == 0:
                row.append("FALSE")
            else:
                row.append("TRUE")

            # gets data from initiative
            for f in initiative_fields:
                value = getattr(initiative, f)
                if f == 'recipient_temp' and value is not None:
                    row.append(value.name)
                else:
                    row.append(value)

            # has photos: true/false
            if initiative.photo_set.all().count() == 0:
                row.append("FALSE")
            else:
                row.append("TRUE")

            ws_output.append(row)

        pprint(counters)
        # save output file
        workbook.save(output_filename)
        if len(np_to_fix) > 0:
            self.logger.error("The following NewProject PK have code malformed or null")
            pprint(np_to_fix)
        self.logger.info(u"finish")