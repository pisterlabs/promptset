# coding=utf-8
__author__ = 'stefano'
import logging
from pprint import pprint
from optparse import make_option
from datetime import datetime
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from openpyxl import load_workbook, Workbook
from django.core.management.base import BaseCommand
from openaid.projects.models import Project, Activity, Initiative

#import projects start/end dates from xlsx file
#at the moment just outputs xls file with error dates

class Command(BaseCommand):

    option_list = BaseCommand.option_list + (
        make_option('--file',
                    dest='file',
                    default='',
                    help='path to input file'),
    )

    help = 'import projects start/end dates from xlsx file'
    logger = logging.getLogger('openaid')
    default_date_format = '%d/%m/%y'

    def date_to_string(self, date, date_format=None):
        if date_format is None:
            date_format = self.default_date_format
        else:
            date_format = date_format
        if date is not None:
            try:
                return datetime.strftime(date, date_format)
            except (ValueError, TypeError):
                self.logger.error("Wrong date value:{}".format(date))
                return ''
        return ''

    def get_reduced_row(self, row, skip_dates):
        rr = []
        for i in range(9):
            value = row[i].value
            if skip_dates is False and (4 <= i <= 7):
                value = self.date_to_string(value)
            rr.append(value)
        return rr

    def handle(self, *args, **options):
        verbosity = options['verbosity']
        input_filename = options['file']
        error_filename= '{}_dates_errorlog.xlsx'.format(self.date_to_string(datetime.now(),"%Y%m%d%H%M"))
        if verbosity == '0':
            self.logger.setLevel(logging.ERROR)
        elif verbosity == '1':
            self.logger.setLevel(logging.WARNING)
        elif verbosity == '2':
            self.logger.setLevel(logging.INFO)
        elif verbosity == '3':
            self.logger.setLevel(logging.DEBUG)

        error_workbook = Workbook()
        ws_pr_notfound = error_workbook.create_sheet(index=0, title='Project number not found')
        ws_pr_notunique = error_workbook.create_sheet(index=1, title='Project number not unique')
        ws_in_notfound = error_workbook.create_sheet(index=2, title='Initiative code not found')

        self.logger.info(u"Opening input file: {}".format(input_filename))
        input_file = open(input_filename, 'rb')
        input_workbook = load_workbook(input_file, data_only=True)
        input_ws = input_workbook['Foglio1']
        not_found=multiple=initiative_counter=0
        row_counter = 0
        for row in input_ws.rows:
            if row_counter == 0 :
                header = self.get_reduced_row(row=row, skip_dates=True)
                row_counter+=1
                # write header in error log file
                ws_pr_notfound.append(header)
                ws_pr_notunique.append(header)
                ws_in_notfound.append(header)
                continue

            project_number = row[0].value
            initiative_code = project_number[:6]
            reduced_row = self.get_reduced_row(row, False)
            # datainizio = row[5].value
            # datafine = row[6].value
            try:
                Project.objects.get(number=project_number)
            except ObjectDoesNotExist:

                # try to look for an activity with that code

                activities = Activity.objects.filter(number=project_number)
                if activities.count() == 0:
                    self.logger.error("Proj not found:'{}'".format(project_number))
                    not_found+=1
                    ws_pr_notfound.append(reduced_row)

            except MultipleObjectsReturned:
                self.logger.error("Multiple proj found:'{}'".format(project_number))
                multiple+=1
                ws_pr_notunique.append(reduced_row)

            try:
                Initiative.objects.get(code=initiative_code)
            except ObjectDoesNotExist:
                self.logger.error("Initiative not found found:'{}'".format(initiative_code))
                initiative_counter +=1
                ws_in_notfound.append(reduced_row)


        self.logger.error("{} proj.number not found, {} proj.number not unique, {} initiative not found".format(not_found, multiple, initiative_counter))

        if not_found > 0 or multiple > 0 or initiative_counter > 0:
            self.logger.info(u"Error log file:{}".format(error_filename))
            error_workbook.save(error_filename)
