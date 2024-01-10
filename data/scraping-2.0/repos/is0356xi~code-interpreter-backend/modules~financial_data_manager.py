import pypdf
import re
import os
from dotenv import load_dotenv
import openai
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Font


load_dotenv()

# FunctionCalling用のクラス
from modules.function_calling import(
    create_financial_data_dict,
)

CreateFinancialDataDict = create_financial_data_dict.CreateFinancialDataDict()

# テキスト抽出&表作成用のプレフィックス
PREFIX = """
# ゴール
- テキストから「会社名と決算期」に該当しそうなデータを抽出する
- テキストから今期の$(バランスシート)に該当しそうなデータを抽出する
- テキストから今期の$(損益計算書)に該当しそうなデータを抽出する
- テキストから今期の$(キャッシュフロー計算書)に該当しそうなデータを抽出する
- テキストから「直近に公表されている予想から修正があるか抽出する
- テキストから「株式分割があるか」抽出する

# 変数の定義
- $(バランスシート)：流動資産合計・固定(非流動)資産合計・流動負債合計・固定(非流動)負債合計・純資産(資本)合計
- $(損益計算書)：売上高・売上原価・販管費・経常利益
- $(キャッシュフロー計算書)：営業活動・投資活動・財務活動

# 実行のプロセス
- テキストファイルを読み込む (今期のデータがどこにあるか入念に確認する)
- 変数の定義を確認する
- 変数の定義に示されたデータを抽出し、表形式で結果を出力する

それでは、実行のプロセスに従って表形式の成果物を作成してください。
必ず全ての表を埋めて成果物を作成すること。
以下はテキストファイルです。必ず必要なデータが含まれているので入念に探してください。
   
###

"""

# テキスト抽出し、表を出力するシステムプロンプト
EXTRACT_TEXT_SYSTEM_PROMPT = """
あなたは与えられたテキストから財務データを抽出し、表を出力するAIです。
回答は表形式で、必ず表のみを出力することを忘れないでください。
"""

# function callingのシステムプロンプト
FUNCTION_CALLING_SYSTEM_PROMPT = """
あなたはユーザを助けるアシスタントです。必要に応じてfunctionを使用することができます。
"""

# Function calling用のサフィックス
SUFFIX = """
###

上記のテキストから財務データの辞書を作成して。
"""


class FinancialDataManager:
    def __init__(self) -> None:
        # Azure OpenAIの設定
        openai.api_type = "azure"
        openai.api_key = os.getenv("AZURE_OPENAI_API_KEY")
        openai.api_base = os.getenv("AZURE_OPENAI_ENDPOINT")
        openai.api_version = os.getenv("AZURE_OPENAI_API_VERSION")


    # テキスト抽出
    def extract_text(self, target_file:str) -> str:
        reader1 = pypdf.PdfReader(target_file)

        # 抽出するキーワード
        keywords = [
            r"上場会社名", 
            r"営業外.益", 
            r"営業活動によるキャッシュ・フロー",
            # r"発行.*株数"
            r"(固定負債合計|固定負債計|非流動資産合計|非流動資産計)",
            r"(流動負債合計|流動負債計)",
            r"(流動資産合計|流動資産計)",
            r"(資本合計|純資産合計)",
        ]

        # マッチしたページとキーワードを格納するリスト
        match_pages = []
        match_keywords = []

        # テキストを抽出
        pdf_text = ""
        for i in range(len(reader1.pages)):
            page_text = reader1.pages[i].extract_text()
            
            for keyword in keywords:
                # 上場会社名が抽出できない場合は、スペースを削除
                if keyword == r"上場会社名" and re.search(keyword, page_text) == None:
                    page_text = self._remove_space(page_text)

                # テキスト整形処理
                page_text = self._replace_number(page_text)
                page_text = self._process_data(page_text)

                # キーワードが含まれ、かつ、既に抽出済みのページでなければ、抽出
                if re.search(keyword, page_text) and (keyword not in match_keywords) and (i not in match_pages):
                    match_pages.append(i)
                    match_keywords.append(keyword)
                    print(keyword)
                    print("page:", i, "\n")
                    # 1ページ目だけ文字列をカット
                    if i == 0:
                        page_text = page_text[:200]
                    pdf_text += page_text + "\n\n"
                else:
                    pass

        return pdf_text
    
    # 数字データを抽出するための整形処理
    def _process_data(self, text):
        # 正規表現パターンを使って、小数点の後の数字の後ろにスペースを追加
        result = re.sub(r"\.(\d)", r".\1 ", text)

        # 正規表現パターンを使って、コンマの後の3つの数字にスペースを追加
        result = re.sub(r",(\d{3})", r",\1 ", result)

        # 正規表現パターンを使って、コンマの後に数字が4つ以上続く場合は、最後の3つの数字の後ろにスペースを追加
        result = re.sub(r",(\d{3})(\d)", r",\1 \2", result)

        # 売上が兆の場合は、無駄なスペースが入るので削除
        result = result.replace(" ,", ",")

        # "yyyy年m月期第q四半期" を空文字列に置換して切り落とす
        result = re.sub(r"\d{4}年\d{1,2}月期第\d{1}四半期", "", result)

        # "通期"を空文字列に置換して切り落とす
        result = result.replace("通期", "")

        return result


    # 数字部分が全角になっている場合は半角数字に変換
    def _replace_number(self, text:str) -> str:
        text = text.replace("０", "0")
        text = text.replace("１", "1")
        text = text.replace("２", "2")
        text = text.replace("３", "3")
        text = text.replace("４", "4")
        text = text.replace("５", "5")
        text = text.replace("６", "6")
        text = text.replace("７", "7")
        text = text.replace("８", "8")
        text = text.replace("９", "9")

        return text


    # 無駄なスペースを削除
    def _remove_space(self, text):
        result = text.replace(" ", "")

        return result
    
    # 財務分析用のデータに関連するテキストを抽出し、表形式で出力させる
    def extract_text_and_create_table(self, text):
        # メッセージを初期化
        messages = [
            {"role":"system", "content": FUNCTION_CALLING_SYSTEM_PROMPT},
            {"role": "user", "content": PREFIX + text},
        ]

        # OpenAI APIを使って、推論実行
        response = openai.ChatCompletion.create(
            messages=messages,
            engine=os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME_16k"),
            temperature=0.0,
        )

        # 結果を抽出
        result = response["choices"][0]["message"]["content"]

        print(result)

        return result


    # 財務分析用のデータを抽出して辞書型にする
    def extract_financial_data(self, text):
        # function callingを実行
        result = self._exec_function_calling(text)

        # データをチェック
        if result == None:
            return None
        else:
            result = self._check_data(result)
            return result


    def _exec_function_calling(self, text):
        # メッセージを初期化
        messages = [
            {"role":"system", "content": FUNCTION_CALLING_SYSTEM_PROMPT},
            {"role": "user", "content": text + SUFFIX},
        ]

        # 使用する関数のメタデータ
        functions_metadata = [CreateFinancialDataDict.metadata]

        # 推論実行
        response = openai.ChatCompletion.create(
            engine = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME"),
            messages = messages,
            functions=functions_metadata,
            function_call="auto",
            temperature=0,
        )

        # 関数呼び出し有無を判定
        if response["choices"][0]["message"].get("function_call"):
            msg = response["choices"][0]["message"]
            return json.loads(msg["function_call"]["arguments"])
        else:
            return None
        
    def _check_data(self, data:dict) -> dict:
        # キーを満たしているか確認
        keys = [
            "営業活動キャッシュフロー",
            "投資活動キャッシュフロー",
            "財務活動キャッシュフロー",
        ]

        # キーがない場合は値をNoneで追加
        for key in keys:
            if key not in data.keys():
                data[key] = None

        return data
    

    def create_excel(self, data:dict, file_path:str):
        # データフレームを作成
        df = pd.DataFrame([data])

        # Excelが存在する場合は読み込み、存在しない場合は新規作成
        try:
            wb = load_workbook(file_path)
        except FileNotFoundError:
            wb = Workbook()

        # 会社名のシートが存在する場合は、そのシートに追記
        if data["会社名"] in wb.sheetnames:
            ws = wb[data["会社名"]]

            # データを追記
            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append(r)
        else:
            # 会社名のシートが存在しない場合は、新規作成
            ws = wb.create_sheet(data["会社名"])
            
            # データを追記
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # ヘッダーのフォントを太字にする
            for cell in ws[1]:
                cell.font = Font(bold=True)

        # 空のシートが存在する場合は削除
        for sheet in wb.worksheets:
            if sheet.title == "Sheet":
                wb.remove(sheet)

        # Excelファイルを保存
        wb.save(file_path)

