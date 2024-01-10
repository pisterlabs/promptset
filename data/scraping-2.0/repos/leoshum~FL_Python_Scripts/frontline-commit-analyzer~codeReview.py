import openai
import tiktoken
import os
from openpyxl import load_workbook

class CodeReviewProvider:
    def __init__(self, chat_completion=False):
        self.chat_completion = chat_completion
        openai.api_key = os.environ.get("OPENAI_API_TOKEN")
        with open(f"{os.path.dirname(__file__)}\\preprompt.txt", "r") as file:
            self.prepromt = file.read()
        
        with open(f"{os.path.dirname(__file__)}\\c-sharp-issues.txt", "r") as file:
            self.csharp_preprompt = file.read()

        with open(f"{os.path.dirname(__file__)}\\sql-issues.txt", "r") as file:
            self.sql_preprompt = file.read()
        
        with open(f"{os.path.dirname(__file__)}\\angular-issues.txt", "r") as file:
            self.angular_preprompt = file.read()
        
        with open(f"{os.path.dirname(__file__)}\\js-issues.txt", "r") as file:
            self.js_preprompt = file.read()
        
        with open(f"{os.path.dirname(__file__)}\\binary-answer-preprompt.txt", "r") as file:
            self.binary_prepromt = file.read()

    def get_chat_completion_answer(self, prompt):
        response = openai.ChatCompletion.create(
        model="gpt-4-0613",
        messages=[
                {"role": "user", "content": prompt}
            ]
        )
        return response['choices'][0]['message']['content']

    def get_completion_answer(self, prompt):
        model_engine = "text-davinci-003"
        encoding = tiktoken.get_encoding("p50k_base")
        max_tokens = 4097 - len(encoding.encode(prompt))
        if max_tokens < 0:
            max_tokens = 4097
        completion = openai.Completion.create(
            engine=model_engine,
            prompt=prompt,
            max_tokens=max_tokens,
            temperature=0.5,
            top_p=1,
            frequency_penalty=0,
            presence_penalty=0
        )
        return completion.choices[0].text.strip()

    def get_bot_answer(self, prepromt, code, file_path, binary_answer=False):
        file_ext = os.path.splitext(file_path)[1]
        
        ext_to_excluede = [".xml", ".rdlc", ".resx", ".json", ".md", ".csproj", ".sln"]
        if file_ext in ext_to_excluede:
            return "Skipped"
        
        code_issues = ""
        if file_ext == ".cs":
            code_issues = self.csharp_preprompt
        elif file_ext == ".sql":
            code_issues = self.sql_preprompt
        elif file_ext == ".ts":
            code_issues = self.angular_preprompt
        elif file_ext == ".js":
            code_issues = self.js_preprompt
        else:
            code_issues = prepromt

        prompt = ""
        if binary_answer:
            prompt = f"{self.binary_prepromt}\n{code}"
        else:
            prompt = f"{code_issues}\n{code}"

        result = ""
        if self.chat_completion:
            result = self.get_chat_completion_answer(prompt)
        else:
            result = self.get_completion_answer(prompt)
        return result
    
    def get_code_review(self, code, file_path):
        return self.get_bot_answer(self.prepromt, code, file_path)
    
    def get_binary_answer(self, review, file_path):
        return self.get_bot_answer(self.binary_prepromt, review, file_path, binary_answer=True)
    
def main():
    code_review = CodeReviewProvider(chat_completion=True)
    file_name = f"{os.path.dirname(__file__)}\\code_issues.xlsx"
    wb = load_workbook(file_name, data_only=True)
    wb_sheet = wb.active
    file_ext = ".md"
    for row in wb_sheet.iter_rows():
        if row[0].value in ["C#", "JS/TS/Angular", "SQL"]:
            file_ext = row[1].value
        else:
            row[4].value = code_review.get_code_review(row[0].value, f"https://api.github.com/repos/octokit/octokit.rb/contents/README{file_ext}")
            row[5].value = code_review.get_binary_answer(row[0].value, f"https://api.github.com/repos/octokit/octokit.rb/contents/README{file_ext}")
    wb.save(file_name)
if __name__ == "__main__":
    main()