# Interpreters MSNScript 2.0
#
# See documentation for more information,
# documentation could lack functions or
# capabilities in this interpreter, as
# this is a work in progress.
#
# docs: masonmarker.com/#/msn2docs
#
# Author : Mason Marker
# Start date : 09/15/2022


# TODO
# speed up function interpretation
# by determining obj and func by argument count first
# as opposed to iterating through all functions.
# this also entails testing and reordering function priority to
# maximize speed.
#
# TODO
# implement string parsing of the same character at which
# it was defined. ex: "hello \"world\"" -> hello "world"
# currently, this is not possible
#
# TODO
# Determine absolutely necessary dependencies
#
# TODO (less important)
# split interpreter up into multiple files
# for better readability
#
# TODO
# implement warnings and warning handling, as this
# language was designed to be safe yet flexible
#
# TODO
# implement linear interpretation in areas of heavy logic, this applies
# easily non linear approaches in several blocks
# such as <<>> or system calls such as script()
#
# TODO
# implement an interpretation for block syntax
# that permits the existence of whitespace / tabs / carriage returns
# in the multilined block to interpret
#
# TODO
# ensure no code repetition
# (it definitely exists)


# the current logical implementation is conceptual,
# deoptimized, and exists to prove functionality as speed can 
# be enhanced later

import os
import math
import shutil
import psutil
import pywinauto
import win32api
import warnings
import pyperclip

# pywinauto automation
from pywinauto.application import Application
from pywinauto import mouse, \
                timings, \
                controls, \
                findwindows, \
                ElementAmbiguousError, \
                ElementNotFoundError

# automating Excel
import openpyxl
import concurrent.futures

# ChatGPT API (just in case)
import openai

# multiprogramming
import subprocess
import threading

# APIs
import requests
from flask import Flask, request, jsonify
from flask_restful import Resource, Api
from flask_cors import CORS


# general
import random
import time
import logging
import socket
import sys
import re

# web scraping
from bs4 import BeautifulSoup

# remove warnings for calling of integers: "10()"
warnings.filterwarnings("ignore", category=SyntaxWarning)

# pywinauto defaults
timings.Timings.after_clickinput_wait = 0.001
timings.Timings.after_setcursorpos_wait = 0.001
timings.Timings.after_sendkeys_key_wait = 0.001
timings.Timings.after_menu_wait = 0.001


# error reporting
class Err:
    def __init__(self, errorcode):
        self.errorcode = errorcode

# variable
class Var:

    # constructs a new Var
    def __init__(self, _msn2_reserved_varname, _msn2_reserved_varvalue):
        self.name = _msn2_reserved_varname
        self.value = _msn2_reserved_varvalue

    # determines equality of another Var
    def __eq__(self, other):
        if isinstance(other, Var):
            return other.name == self.name


# creates an ai model
# creates and returns a custom ai model
def ai_response(model, prompt, creativity):
    return openai.Completion.create(
        model=model,
        prompt=prompt,
        temperature=creativity
    ).choices[0].text

# alias to run Python in the terminal
python_alias = 'python'

# msn2 implementation of None
msn2_none = '___msn2_None_'

# thread serial
thread_serial = 0

# global vars
lock = threading.Lock()
auxlock = threading.Lock()

# lock for pointer controls
pointer_lock = threading.Lock()

# pywinauto automation lock
auto_lock = threading.Lock()

# user defined syntax
syntax = {}

# user defined enclosing syntax
enclosed = {}

# user defined macros
macros = {}

# user defined post macros
# aka macros that are defined that the end of a line
postmacros = {}

# user defined inline syntax
inlines = {}

# OpenAI model presets
# these are the latest models at the time of creation
models = {

    # standard model, answers questions will little to no creativity
    'basic': {'model': 'text-davinci-003', 'creativity': 0},

    # creative model, answers questions with creativity
    'creative': {'model': 'text-davinci-003', 'creativity': 0.5},

    # advanced model, answers questions with high creativity
    'advanced': {'model': 'text-davinci-003', 'creativity': 1}

}


# accounting information
lines_ran = []
total_ints = 0

# automation
apps = {}

# interprets MSNScript2, should create a new interpreter for each execution iteration
class Interpreter:

    # initializer
    def __init__(self):
        self.version = 2.0
        self.lines = []
        self.out = ''
        self.log = ''
        self.errors = []

        self.vars = {}
        self.methods = {}
        self.loggedmethod = []
        self.objects = {}
        self.calledmethod = None

        self.current_line = 0
        self.breaking = False
        self.redirect = None
        self.redirecting = False
        self.redirect_inside = []
        self.imports = set()

        self.thread = None
        self.threads = {}
        self.parent = None

        self.openaikey = None
        self.tokens = 100
        self.browser_path = None
        self.serial_1 = 0

        self.endpoints = {}
        self.endpoint_datas = {}
        self.endpoint_path = 'demos/practical/apidata/apitestdata.csv'

        self.processes = {}
        self.breaking_return = []
    
    # executes stored script
    def execute(self, script):

        # # convert script to lines
        self.lines = []

        # for aggregate syntax support !{}
        inmultiline = False
        multiline = ''

        # for block syntax
        inblock = False
        p = 0
        
        # whether or not to keep 
        keep_space = False
        for line in filter(None, script.split("\n")):
            self.lines.append(line)
            if line.startswith('\\\\'):
                keep_space = True
                line = line[2:]
            elif keep_space:
                ...
            else:
                line = line.strip()
            if line.startswith("::") or line.startswith("#"):
                self.current_line += 1
                continue

            else:

                # aggregate syntax !{} (not recommended for most cases)
                if line.startswith('!{') and line.endswith('}'):
                    multiline = line[2:-1]
                    self.interpret(multiline)
                    multiline = ''
                elif not inmultiline and line.startswith("!{"):
                    inmultiline = True
                    multiline += line[2:]
                elif inmultiline and line.endswith("}"):
                    inmultiline = False
                    multiline += line[0:len(line) - 1]
                    self.interpret(multiline)
                    multiline = ''
                elif inmultiline:
                    multiline += line

                # block syntax (recommended for most cases)
                elif not inblock and line.endswith('('):
                    for c in line:
                        if c == '(':
                            p += 1
                        if c == ')':
                            p -= 1
                        multiline += c
                    inblock = True
                elif inblock:
                    for i in range(len(line)):
                        c = line[i]
                        if c == '(':
                            p += 1
                        if c == ')':
                            p -= 1

                        # end of syntax met
                        if p == 0:
                            multiline += line[i:]
                            inter = multiline
                            multiline = ''
                            inblock = False
                            self.interpret(inter, keep_space=keep_space)
                            break
                        multiline += c
                else:
                    self.interpret(line, keep_space=keep_space)

            self.current_line += 1
        return self.out

    def replace_vars(self, line):
        boo = line
        for varname in sorted(self.vars.keys(), key=len, reverse=True):

            try:

                boo = boo.replace(varname, str(
                    self.get_var(eval(f'"{varname}"', {}, {}))))
            except:
                None
        return boo

    # interprets a line

    def interpret(self, line, block={}, keep_space=False):

        # acquiring globals
        global total_ints
        global lock
        global auxlock
        global auto_lock
        global pointer_lock
        global python_alias

        # accounting
        total_ints += 1
        lines_ran.append(line)

        # interpreter is breaking
        if self.breaking:
            return self.breaking_return

        # strip line for interpretation
        try:
            if not keep_space:
                line = line.strip()
        except:
            return
        l = len(line)

        # whether the line should immediately continue or not
        cont = False
        if line == '':
            return

        # the below conditions interpret a line based on initial appearances
        # beneath these conditions will the Interpreter then parse the arguments from the line as a method call

        # method-specific line reached
        if line.startswith('--'):
            line = line[2:]
            try:
                if not self.methods[self.loggedmethod[-1]].ended:
                    self.methods[self.loggedmethod[-1]].add_body(line)
            except:
                None
            return

        # new variable setting and modification syntax as of 12/20/2022
        # iterates to the first '=' sign, capturing the variable name in the
        # process (as it should)
        # msn1 fallback
        if line[0] == '@':
            return self.interpret_msnscript_1(line[1:])

        # python fallback mode specification, 
        # both <<>> and
        if line.startswith('<<'):
            
            # parse all text in the line for text surrounded by |
            funccalls = []
            infunc = False
            func = ''
            for i in range(line.rindex('>>')):
                if line[i] == '|' and not infunc:
                    infunc = True
                elif line[i] == '|' and infunc:
                    infunc = False
                    funccalls.append(func)
                    func = ''
                elif infunc:
                    func += line[i]
            for function in funccalls:
                ret = self.interpret(function)
                if isinstance(ret, str):
                    line = line.replace(
                        f'|{function}|', f'"{str(ret)}"')
                else:
                    line = line.replace(f'|{function}|', str(ret))
            line = line[2:-2]
            try:
                return eval(line)
            except:
                return line

        # embedded MSN2 interpretation macro
        if line.startswith('<2>'):
            # parse all text in the line for text surrounded by %
            funccalls = []
            infunc = False
            func = ''
            for i in range(3, line.rindex('<2>')):
                if line[i] == '%' and not infunc:
                    infunc = True
                elif line[i] == '%' and infunc:
                    infunc = False
                    funccalls.append(func)
                    func = ''
                elif infunc:
                    func += line[i]

            # for each msn2 evaluation
            for function in funccalls:
                ret = self.interpret(function)
                if isinstance(ret, str):
                    line = line.replace(
                        f'%{function}%', f'"{str(ret)}"')
                else:
                    line = line.replace(f'%{function}%', str(ret))
            line = line[3:-3]
            try:
                return self.interpret(line)
            except:
                try:
                    return eval(line, {}, {})
                except:
                    return line

        # # offers a way to repetively call methods on a variable.
        # # generally unsafe, yet retained for specific cases
        # if line.startswith('()'):
        #     line = line[2:]

        #     # split line by --> macro
        #     line = line.split('-->')
        #     ret = None
        #     varname = ''

        #     # iterate to the first '.', this will be the literal or variable name
        #     for i in range(len(line[0])):
        #         if line[0][i] != ' ':
        #             if line[0][i] == '.':
        #                 break
        #             varname += line[0][i]

        #     if varname in self.vars:
        #         # interpret the first instruction
        #         ret = self.interpret(line[0])

        #         # interpret the rest of the instructions prefixed with the variable name and '.'
        #         for i in range(1, len(line)):
        #             ret = self.interpret(f"{varname}.{line[i]}")
        #     else:
        #         self.vars[varname]

        #     return ret

        # user defined syntax
        for key in syntax:
            if line.startswith(key):
                return self.run_syntax(key, line)

        # user defined macro
        for token in macros:
            if line.startswith(token):

                # if the macro returns a value instead of executing a function
                if len(macros[token]) == 4:
                    return macros[token][3]

                # variable name
                varname = macros[token][1]

                # function to execute
                function = macros[token][2]

                val = line[len(token):]

                # store extended for user defined syntax
                self.vars[varname] = Var(varname, val)

                # execute function
                return self.interpret(function)

        # user defined postmacro
        for token in postmacros:
            if line.endswith(token):
                # if the macro returns a value instead of executing a function
                if len(postmacros[token]) == 4:
                    return postmacros[token][3]
                varname = postmacros[token][1]
                function = postmacros[token][2]
                val = line[0:len(line) - len(token)]
                self.vars[varname] = Var(varname, val)
                return self.interpret(function)

        # variable replacement, generally unsafe
        if line[0] == '*':
            line = self.replace_vars(line[1:])
            return self.interpret(line)

        # invoking user defined enclosing syntax
        for key in enclosed:
            start = enclosed[key][0]
            end = enclosed[key][1]
            if line.startswith(start) and line.endswith(end):
                if len(enclosed[key]) == 5:
                    return enclosed[key][4]
                varname = enclosed[key][2]
                block = enclosed[key][3]
                val = line[len(start):len(line) - len(end)]
                self.vars[varname] = Var(varname, val)
                return self.interpret(block)

        # checks for active Interpreter redirect request
        if self.redirecting and not 'stopredirect()' in line.replace(' ', ''):
            _block = self.redirect[1]
            self.redirect_inside.append([_block, line])
            return self.redirect

        # try base literal
        try:
            if not line.startswith('--'):
                
                # try evaluating the line
                _ret = eval(line, {}, {})

                # eval cannot be a python class, because names of variables
                # could result in python classes
                # should also not be a built in function
                if not isinstance(_ret, type) and not isinstance(_ret, type(eval)):
                    return _ret
        except:
            None

        func = ''
        objfunc = ''
        obj = ''
        s = 0
        sp = 0
        for i in range(l):
            if cont:
                continue
            try:
                c = line[i]
            except:
                break
            if c == ' ' and s == 0:
                sp += 1
                continue

            if c == '.':
                obj = func
                func = ''
                objfunc = ''
                continue

                                    
            # function to attempt to pull an integer out of the function
            # if it is an integer, then it is a loop that runs func times
            def get_int(func):
                try:
                    return int(func)
                except:
                    return None

            # basic method creation
            if c == '~':
                returnvariable = ''
                self.loggedmethod.append('')
                for j in range(i + 1, len(line)):
                    if line[j] != ' ':
                        if line[j] == '(':
                            args = self.method_args(line, j)
                            for k in range(args[1], len(line)):
                                if line[k] != ' ':
                                    if line[k] == '-' and line[k + 1] == '>':
                                        for l in range(k + 2, len(line)):
                                            if line[l] != ' ':
                                                returnvariable += line[l]
                            break
                        self.loggedmethod[-1] += line[j]
                if self.loggedmethod[-1] not in self.methods.keys():
                    self.methods[self.loggedmethod[-1]
                                 ] = self.Method(self.loggedmethod[-1], self)
                else:
                    break
                for arg in args[0]:
                    if arg != '':
                        self.vars[arg] = None
                        self.methods[self.loggedmethod[-1]].add_arg(arg)
                self.methods[self.loggedmethod[-1]].add_return(returnvariable)
                return self.loggedmethod[-1]

            # interpreting a function
            elif c == '(':

                mergedargs = ''
                p = 1
                l = len(line)
                for j in range(i + 1, l - 1):
                    c2 = line[j]
                    if p == 0:
                        break
                    if c2 == '(':
                        p += 1
                    if c2 == ')':
                        p -= 1
                    mergedargs += c2
                args = self.get_args(mergedargs)
                f = len(func)

                # clean function for handling
                func = func.strip()
                objfunc = objfunc.strip()

                # class attribute / method access
                if obj in self.vars:
                    vname = obj
                    try:
                        var = self.get_var(vname)
                    except:
                        var = self.vars[vname]
                    try:
                        object = self.vars[obj].value
                    except:
                        object = self.vars[obj]

                    try:
                        # if the object is a class
                        if objfunc in object:

                            # if the object is a self.Method
                            if type(object[objfunc]) == self.Method:

                                # get the Method object
                                method = object[objfunc]

                                # get the number of arguments to the method
                                num_args = len(method.args)

                                # args to pass to the function
                                to_pass = [vname]

                                # if there is no argument
                                if args[0][0] != '':

                                    # for each parsed argument
                                    for k in range(num_args):
                                        try:
                                            to_pass.append(self.parse(
                                                k, line, f, sp, args)[2])
                                        except:
                                            None
                                # create return variable
                                ret_name = method.returns[0]

                                # if the return variable doesn't exist
                                if ret_name not in self.vars:
                                    self.vars[ret_name] = Var(ret_name, None)
                                    
                                # # insert vname into args[0]
                                args.insert(0, [vname])

                                    
                                # execute method
                                # if objfunc == 'wait_for_field':
                                #     print(to_pass, args)
                                method.run(to_pass, self, args)

                                try:
                                    return eval(str(self.vars[method.returns[0]].value), {}, {})
                                except:
                                    try:
                                        return str(self.vars[method.returns[0]].value)
                                    except:
                                        return str(self.vars[method.returns[0]])

                            # otherwise if we're accessing an attribute

                            # no arguments given
                            if args[0][0] == '':
                                return object[objfunc]

                            # parameter provided, wants to set attribute
                            param = self.parse(0, line, f, sp, args)[2]

                            self.vars[obj].value[objfunc] = param
                            return param
                    except:
                        None

                    # continuously accesses a key
                    if '->' in objfunc:
                        keys = objfunc.split('->')

                        # no arguments given
                        if args[0][0] == '':
                            # get the value on this object at the keys
                            for key in keys:
                                try:
                                    object = object[eval(key, {}, {})]
                                except:
                                    object = object[key]

                        # argument given to set
                        else:
                            to_set = self.parse(0, line, f, sp, args)[2]
                            for key in keys[:-1]:
                                object = object[key]
                            object[keys[-1]] = to_set
                            return to_set
                        return object
                    
                    # working with Excel sheets
                    elif isinstance(object, self.Sheet):
                        
                        # active elements
                        title = object.title
                        workbook = object.workbook
                        path = object.path
                        sheet = object.sheet
                        
                        # gets the value of a cell
                        if objfunc == 'get':
                            
                            # column of the cell
                            column = self.parse(0, line, f, sp, args)[2]
                            
                            # row of the cell
                            row = self.parse(1, line, f, sp, args)[2]

                            # returns the value of the cell
                            return sheet.cell(row + 1, column + 1).value
                    
                        # sets the value of a cell
                        if objfunc == 'set':

                            # column of the cell
                            column = self.parse(0, line, f, sp, args)[2]
                            
                            # row of the cell
                            row = self.parse(1, line, f, sp, args)[2]
                            
                            # value to set the cell to
                            value = self.parse(2, line, f, sp, args)[2]
                            
                            # sets the value of the cell
                            sheet.cell(row + 1, column + 1, value)
                            
                            # returns the sheet
                            return value
                        
                        
                        # clears the sheet
                        if objfunc == 'clear':
                                
                            # clears the sheet
                            for row in sheet.iter_rows():
                                for cell in row:
                                    cell.value = None
                                                        
                            # returns the sheet
                            return object
                                                
                        # gets the populated cell values of a column
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        if objfunc == 'column':
                            
                            # column, either an integer or string
                            col = self.parse(0, line, f, sp, args)[2]
                            
                            column_values = []
                            
                            # if number
                            if isinstance(col, int):
                                col += 1
                                for cell in sheet.iter_cols(min_col=col, max_col=col):
                                    for row in cell:
                                        if row.value != None:
                                            column_values.append(row.value)
                                            
                            # otherwise, get column by title
                            elif isinstance(col, str):
                                
                                # for each column
                                for cell in sheet.iter_cols():
                                        # if the title matches
                                        if cell[0].value == col:
                                            
                                            # get the column values
                                            for row in cell:
                                                if row.value != None:
                                                    column_values.append(row.value)
                            return column_values
                            
                        # gets the populated cell values of a row
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        if objfunc == 'row':
                            
                            # row, either an integer or string
                            r = self.parse(0, line, f, sp, args)[2]
                            
                            row_values = []
                            
                            # if number
                            if isinstance(r, int):
                                r += 1
                                for cell in sheet.iter_rows(min_row=r, max_row=r):
                                    for row in cell:
                                        if row.value != None:
                                            row_values.append(row.value)     
                            # otherwise, get row by title
                            elif isinstance(r, str):
                                
                                # for each row
                                for cell in sheet.iter_rows():
                                        # if the title matches
                                        if cell[0].value == r:
                                            
                                            # get the row values
                                            for row in cell:
                                                if row.value != None:
                                                    row_values.append(row.value)
                            return row_values
                    
                        # gets the index of a column with a given title
                        def get_column_index(title):
                            for cell in sheet.iter_cols():
                                if cell[0].value == title:
                                    return cell[0].column
                            return None
                        def get_row_index(title):
                            for cell in sheet.iter_rows():
                                if cell[0].value == title:
                                    return cell[0].row
                            return None
                    
                        # rewrite the above method, but with
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        if objfunc == 'set_column':
                                
                            # column, either an integer or string
                            col = self.parse(0, line, f, sp, args)[2]
                            
                            # array of values
                            values = self.parse(1, line, f, sp, args)[2]
                            
                            # if number
                            if isinstance(col, int):
                                col += 1
                                for i in range(len(values)):
                                    sheet.cell(i + 1, col, values[i])
                                    
                            # otherwise, get column by title
                            elif isinstance(col, str):
                                
                                # for each column
                                for cell in sheet.iter_cols():
                                        # if the title matches
                                        if cell[0].value == col:
                                            
                                            # get the column values
                                            for i in range(len(values)):
                                                sheet.cell(i + 1, get_column_index(col), values[i])
                                                
                            return values
                                
                        # sets a row to an array of values
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        elif objfunc == 'set_row':
                            
                            # row, either an integer or string
                            r = self.parse(0, line, f, sp, args)[2]
                            
                            # array of values
                            values = self.parse(1, line, f, sp, args)[2]
                            
                            # if number
                            if isinstance(r, int):
                                r += 1
                                for i in range(len(values)):
                                    sheet.cell(r, i + 1, values[i])
                                    
                            # otherwise, get row by title
                            elif isinstance(r, str):
                                
                                # for each row
                                for cell in sheet.iter_rows():
                                        # if the title matches
                                        if cell[0].value == r:
                                            
                                            # get the row values
                                            for i in range(len(values)):
                                                sheet.cell(get_row_index(r), i + 1, values[i])          
                            return values
                        
                    
                        # reqrite the above method, but with
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        elif objfunc == 'add_to_column':
                            
                            # column
                            column = self.parse(0, line, f, sp, args)[2]
                            
                            # value to add
                            value = self.parse(1, line, f, sp, args)[2]
                            
                            # if number
                            if isinstance(column, int):
                                column += 1
                                # find the first empty cell in the column
                                for i in range(sheet.max_row + 1):
                                    if sheet.cell(i + 1, column).value == None:
                                        sheet.cell(i + 1, column, value)
                                        return value
                                return value
                                
                            # otherwise, get column by title
                            elif isinstance(column, str):
                                column_index = get_column_index(column)
                                # find the first empty cell in the column
                                for i in range(sheet.max_row + 1):
                                    if sheet.cell(i + 1, column_index).value == None:
                                        sheet.cell(i + 1, column_index, value)
                                        return value
                            return value
                            
                        # adds a value to a row
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        elif objfunc == 'add_to_row':
                                
                            # row
                            row = self.parse(0, line, f, sp, args)[2]
                            
                            # value to add
                            value = self.parse(1, line, f, sp, args)[2]
                            
                            # if number
                            if isinstance(row, int):
                                row += 1
                                # find the first empty cell in the row
                                for i in range(sheet.max_column):
                                    if sheet.cell(row, i + 1).value == None:
                                        sheet.cell(row, i + 1, value)
                                        return value
                                return value
                                
                            # otherwise, get row by title
                            elif isinstance(row, str):
                                row_index = get_row_index(row)
                                # find the first empty cell in the row
                                for i in range(sheet.max_column):
                                    if sheet.cell(row_index, i + 1).value == None:
                                        sheet.cell(row_index, i + 1, value)
                                        return value
                            return value
                        
                        # writes a matrix (2D array) to this Excel Worksheet
                        # with the offsets given by the second and third arguments,
                        # if no second or third arguments, then the matrix is written
                        # starting at the first cell
                        elif objfunc == 'import_matrix':
                            
                            # get the 2D list
                            matrix = self.parse(0, line, f, sp, args)[2]
                            
                            # default offset
                            offsetx = 0
                            offsety = 0
                            
                            # if there is a second argument
                            if len(args) == 2:
                                offsetx = self.parse(1, line, f, sp, args)[2]
                            if len(args) == 3:
                                offsetx = self.parse(1, line, f, sp, args)[2]
                                offsety = self.parse(2, line, f, sp, args)[2]
                            
                            # for each row
                            for i in range(len(matrix)):
                                # for each column
                                for j in range(len(matrix[i])):
                                    w = matrix[i][j]
                                    # if w is an AppElement, write its name
                                    if 'name' in dir(w):
                                        w = w.name
                                    sheet.cell(i + offsety + 1, j + offsetx + 1, w)
                            return matrix    
                            
                        
                        # if nothing else, return the object
                        return object
                    
                    # methods available to all types
                    if objfunc == 'copy':
                        return object.copy()

                    if objfunc == 'print':

                        # if no arguments
                        if args[0][0] == '':
                            print(object)
                            return object

                        # if one argument
                        elif len(args) == 1:

                            # what to print
                            to_print = f"{self.parse(0, line, f, sp, args)[2]}{object}"

                            # print the object
                            print(to_print)
                            return to_print

                        # if two arguments
                        elif len(args) == 2:

                            # what to print
                            to_print = f"{self.parse(0, line, f, sp, args)[2]}{object}{self.parse(1, line, f, sp, args)[2]}"

                            # print the object
                            print(to_print)

                            # return the printed object
                            return to_print
                        return object

                    if objfunc == 'val':
                        return object

                    if objfunc == 'type':
                        return type(object)

                    if objfunc == 'len':
                        return len(object)

                    if objfunc == 'str':
                        return str(object)

                    if objfunc == 'int':
                        return int(object)

                    if objfunc == 'float':
                        return float(object)

                    if objfunc == 'complex':
                        return complex(object)

                    if objfunc == 'bool':
                        return bool(object)

                    if objfunc == 'dict':
                        return dict(object)
                    

                    # gets values from the object if the statement is true for each object
                    # runs the function on each element / kv pair
                    #
                    # method is not destructive
                    if objfunc == 'if':

                        # variable name
                        varname = self.parse(0, line, f, sp, args)[2]

                        # function to execute
                        function = args[1][0]

                        new_list = []

                        # perform logic
                        for el in self.vars[vname].value:
                            self.vars[varname] = Var(varname, el)
                            if self.interpret(function):
                                new_list.append(el)

                        return new_list

                    # comparing object types
                    if objfunc == 'is':
                        return object is self.parse(0, line, f, sp, args)[2]

                    # test if the object is equal to all the parameters
                    if objfunc == 'equals':
                        for i in range(len(args)):
                            if object != self.parse(i, line, f, sp, args)[2]:
                                return False
                        return True

                    # obtains a slice of the iterable
                    if objfunc == 'slice':
                        return self.vars[vname].value[self.parse(0, line, f, sp, args)[2]:self.parse(1, line, f, sp, args)[2]]

                    # gets the index of the object
                    if objfunc == 'index':
                        return self.vars[vname].value.index(self.parse(0, line, f, sp, args)[2])

                    # exports a variable to the parent context
                    if objfunc == 'export':
                        name = vname
                        # if an argument is provided
                        # export as name
                        if args[0][0] != '':
                            name = self.parse(0, line, f, sp, args)[2]
                        self.parent.vars[name] = Var(name, object)
                        return object
                    
                    # if no objfunc, there has been a request
                    # for repeated method calls/access
                    if objfunc == '':
                        ret = self.vars[vname].value
                        # for each block
                        for arg in args:
                            block = arg[0]
                            ret = self.interpret(f"{vname}.{block}")
                        return ret
                    
                    # string_name() returns the string name of the object
                    if objfunc == 'string_name':
                        return vname

                    # performs a function for each element in the iterable
                    if objfunc == 'each':
                        # get the variable name
                        varname = self.parse(0, line, f, sp, args)[2]
                        # get the function
                        func = args[1][0]

                        # try an indexable
                        try:
                            for i in range(len(object)):
                                self.vars[varname] = Var(varname, object[i])
                                self.interpret(func)

                        except:
                            # try a set
                            for i in object:
                                self.vars[varname] = Var(varname, i)
                                self.interpret(func)

                        return object
                    
                    # rfind and lfind
                    # find the index of the first element that satisfies the condition
                    if objfunc == 'rfind':
                        return self.vars[vname].value.rfind(self.parse(0, line, f, sp, args)[2])
                    if objfunc == 'lfind':
                        return self.vars[vname].value.find(self.parse(0, line, f, sp, args)[2])
                    # find
                    # find the index of the first element that satisfies the condition
                    if objfunc == 'find':
                        return self.vars[vname].value.find(self.parse(0, line, f, sp, args)[2])

                    # filters the iterable
                    if objfunc == 'filter':
                        
                        # get the variable name
                        varname = self.parse(0, line, f, sp, args)[2]

                        # get the function
                        block = args[1][0]
                        
                        # filtered
                        filtered = []
                        
                        # filter the iterable
                        for el in object:
                            self.vars[varname] = Var(varname, el)
                            if self.interpret(block):
                                filtered.append(el)
                                
                        # set the variable to the filtered list
                        self.vars[vname].value = filtered
                        
                        # return the filtered list
                        return self.vars[vname].value
                    
                    # basic arithmetic, non-destructive
                    # takes any amount of arguments
                    if objfunc == '+':
                        ret = object
                        for i in range(len(args)):
                            ret += self.parse(i, line, f, sp, args)[2]
                        return ret
                    if objfunc == '-':
                        ret = object
                        for i in range(len(args)):
                            ret -= self.parse(i, line, f, sp, args)[2]
                        return ret
                    if objfunc == '*' or objfunc == 'x':
                        ret = object
                        for i in range(len(args)):
                            ret *= self.parse(i, line, f, sp, args)[2]
                        return ret
                    if objfunc == '/':
                        ret = object
                        for i in range(len(args)):
                            ret /= self.parse(i, line, f, sp, args)[2]
                        return ret
                    if objfunc == '%':
                        ret = object
                        for i in range(len(args)):
                            ret %= self.parse(i, line, f, sp, args)[2]
                        return ret
                    if objfunc == '**':
                        ret = object
                        for i in range(len(args)):
                            ret **= self.parse(i, line, f, sp, args)[2]
                        return ret
                    if objfunc == '//':
                        ret = object
                        for i in range(len(args)):
                            ret //= self.parse(i, line, f, sp, args)[2]
                        return ret

                    # applies methods to to the object, considering
                    # the method takes one argument
                    if objfunc == 'func':
                        ret = object
                        # apply the function to the object
                        for arg in args:
                            method = arg[0]
                            ret = self.interpret(f"{method}({ret})")
                        
                        return ret
        

                    # reverses the iterable
                    if objfunc == 'reverse':
                        self.vars[vname].value = self.vars[vname].value[::-1]
                        return self.vars[vname].value

                    # determines if this object is in the object passed
                    if objfunc == 'in':
                        return self.vars[vname].value in self.parse(0, line, f, sp, args)[2]

                    # variable type specific methods
                    # the isinstance branches below indicate mostly  DESCTRUCTIVE methods!
                    # so be sure to read the code

                    # number specific functions
                    if isinstance(object, int) or isinstance(object, float) or isinstance(object, complex):

                        # increases the value of the variable by 1
                        if objfunc == '++' or objfunc == 'inc':
                            self.vars[vname].value += 1
                            return self.vars[vname].value

                        elif objfunc == '--' or objfunc == 'dec':
                            self.vars[vname].value -= 1
                            return self.vars[vname].value

                        # determines if the number is even
                        elif objfunc == 'even':
                            return self.vars[vname].value % 2 == 0
                        elif objfunc == 'odd':
                            return self.vars[vname].value % 2 != 0

                        # all of the below methods take any amount of arguments
                        # and perform the operation on the variable
                        elif objfunc == 'add':
                            for i in range(len(args)):
                                self.vars[vname].value += self.parse(
                                    i, line, f, sp, args)[2]
                            return self.vars[vname].value

                        elif objfunc == 'sub':
                            for i in range(len(args)):
                                self.vars[vname].value -= self.parse(
                                    i, line, f, sp, args)[2]
                            return self.vars[vname].value

                        elif objfunc == 'mul':
                            for i in range(len(args)):
                                self.vars[vname].value *= self.parse(
                                    i, line, f, sp, args)[2]
                            return self.vars[vname].value

                        elif objfunc == 'div':
                            for i in range(len(args)):
                                self.vars[vname].value /= self.parse(
                                    i, line, f, sp, args)[2]
                            return self.vars[vname].value

                        # computes the absolute value of the number
                        elif objfunc == 'abs':
                            self.vars[vname].value = abs(
                                self.vars[vname].value)
                            return self.vars[vname].value

                        # rounds this number to the nearest place specified by the first argument
                        elif objfunc == 'round':

                            # decimal place
                            decplace = self.parse(0, line, f, sp, args)[2]

                            # round to the nearest decimal place
                            if args[0][0] != '':
                                self.vars[vname].value = round(
                                    self.vars[vname].value, decplace)

                            else:
                                self.vars[vname].value = round(
                                    self.vars[vname].value)
                            return self.vars[vname].value

                        elif objfunc == 'floor':
                            self.vars[vname].value = math.floor(
                                self.vars[vname].value)
                            return self.vars[vname].value

                        elif objfunc == 'ceil':
                            self.vars[vname].value = math.ceil(
                                self.vars[vname].value)
                            return self.vars[vname].value

                        # negates the value, if positive the value becomes negative and vice versa
                        elif objfunc == 'neg':
                            self.vars[vname].value = -self.vars[vname].value
                            return self.vars[vname].value

                        # all of the below methods should take any amount of arguments

                        # test if the variable is greater than all arguments
                        elif objfunc == 'greater' or objfunc == 'greaterthan' or objfunc == 'g':
                            for i in range(len(args)):
                                if self.vars[vname].value <= self.parse(i, line, f, sp, args)[2]:
                                    return False
                            return True

                        elif objfunc == 'less' or objfunc == 'lessthan' or objfunc == 'l':
                            for i in range(len(args)):
                                if self.vars[vname].value >= self.parse(i, line, f, sp, args)[2]:
                                    return False
                            return True

                        elif objfunc == 'greaterequal' or objfunc == 'ge':
                            for i in range(len(args)):
                                if self.vars[vname].value < self.parse(i, line, f, sp, args)[2]:
                                    return False
                            return True

                        elif objfunc == 'lessequal' or objfunc == 'le':
                            return self.vars[vname].value <= self.parse(0, line, f, sp, args)[2]

                        # more basic functions
                        return self.vars[vname].value

                    # set based functions
                    elif isinstance(object, set):

                        # adds all arguments to the set object
                        if objfunc == 'add' or objfunc == 'put':
                            for i in range(len(args)):
                                self.vars[vname].value.add(
                                    self.parse(i, line, f, sp, args)[2])
                            return self.vars[vname].value

                        if objfunc == 'pop':
                            return self.vars[vname].value.pop()

                        # removes all arguments from the set object
                        if objfunc == 'remove':
                            for i in range(len(args)):
                                self.vars[vname].value.remove(
                                    self.parse(i, line, f, sp, args)[2])
                            return self.vars[vname].value

                        # converts this set to a list
                        if objfunc == 'list':
                            return list(self.vars[vname].value)

                        # gets at an index in the set
                        # sets are not indexable
                        if objfunc == 'get':

                            # index to get at
                            ind = self.parse(0, line, f, sp, args)[2]

                            # get the index
                            for i in object:
                                if ind == 0:
                                    return i
                                ind -= 1

                    # array based functions
                    elif isinstance(object, list):

                        # adds all arguments to the first argument which should be a variable name
                        # as a string
                        if objfunc == 'push' or objfunc == 'append' or objfunc == 'add':
                            for i in range(len(args)):
                                self.vars[vname].value.append(
                                    self.parse(i, line, f, sp, args)[2])
                            return self.vars[vname].value

                        # pops a value from the array
                        elif objfunc == 'pop':
                            return self.vars[vname].value.pop()

                        # getting at an index
                        elif objfunc == 'get':
                            return self.vars[vname].value[self.parse(0, line, f, sp, args)[2]]

                        # sets at an index
                        elif objfunc == 'set':
                            self.vars[vname].value[self.parse(0, line, f, sp, args)[
                                2]] = self.parse(1, line, f, sp, args)[2]
                            return self.vars[vname].value

                        # gets the average of this array
                        elif objfunc == 'avg' or objfunc == 'average':
                            return sum(self.vars[vname].value) / len(self.vars[vname].value)

                        # inserts all values at an index
                        if objfunc == 'insert':

                            # index to insert
                            index = self.parse(0, line, f, sp, args)[2]

                            # inserts the rest of the arguments, one at a time
                            for i in range(len(args)):
                                self.vars[vname].value.insert(
                                    index, self.parse(i, line, f, sp, args)[2])
                            return self.vars[vname].value

                        # removes a certain amount of all arguments supplied
                        if objfunc == 'removen':
                            count = self.parse(0, line, f, sp, args)[2]

                            # removes count amount of the rest of the arguments from the object
                            for i in range(1, len(args)):
                                for j in range(count):
                                    del var[var.index(self.parse(
                                        i, line, f, sp, args)[2])]
                            return object

                        # removes all occurances of each argument from the list
                        if objfunc == 'remove':
                            for i in range(len(args)):
                                while self.parse(i, line, f, sp, args)[2] in var:
                                    del var[var.index(self.parse(
                                        i, line, f, sp, args)[2])]
                            return object

                        # gets a sorted copy of this array
                        if objfunc == 'sorted':
                            return sorted(self.vars[vname].value)

                        # sorts this array
                        if objfunc == 'sort':
                            self.vars[vname].value.sort()
                            return self.vars[vname].value

                        # gets the length of the array
                        if objfunc == 'len':
                            return len(self.vars[vname].value)

                        # determines if a list is empty
                        if objfunc == 'empty':
                            return len(self.vars[vname].value) == 0

                        # determines if this list contains an element
                        if objfunc == 'contains' or objfunc == 'has' or objfunc == 'includes':
                            return self.parse(0, line, f, sp, args)[2] in self.vars[vname].value

                        # finds an element in a list
                        # unlike index(), find returns -1 instead of throwing an
                        # error
                        if objfunc == 'find':
                            return self.vars[vname].value.find(self.parse(0, line, f, sp, args)[2])

                        # shuffles a list
                        if objfunc == 'shuffle':
                            random.shuffle(self.vars[vname].value)
                            return self.vars[vname].value

                        # performs a function for each element in the iterable
                        # map() is a destructive method
                        if objfunc == 'map':
                            # get the variable name
                            varname = self.parse(0, line, f, sp, args)[2]

                            # get the function
                            func = args[1][0]

                            for i in range(len(object)):
                                self.vars[varname] = Var(varname, object[i])
                                object[i] = self.interpret(func)
                            del self.vars[varname]
                            return object

                        # joins the array by the first argument
                        if objfunc == 'join' or objfunc == 'delimit':

                            # join the array
                            return str(self.parse(0, line, f, sp, args)[2]).join(map(str, self.vars[vname].value))

                        # converts this list to a set
                        if objfunc == 'toset':
                            return set(self.vars[vname].value)

                    # if the object is a string
                    elif isinstance(object, str):

                        if objfunc == 'add':
                            for i in range(len(args)):
                                self.vars[vname].value += self.parse(
                                    i, line, f, sp, args)[2]
                            return self.vars[vname].value

                        if objfunc == 'split':
                            arg = self.parse(0, line, f, sp, args)[2]
                            try:
                                return self.vars[vname].value.split(arg)
                            except:
                                return self.vars[vname].split(arg)

                        # gets the lines of this string
                        if objfunc == 'lines':
                            return self.vars[vname].value.split('\n')

                        # determines if the string is a digit
                        if objfunc == 'isdigit':
                            return self.vars[vname].value.isdigit()
                        # determines if the string is a letter
                        if objfunc == 'isalpha':
                            return self.vars[vname].value.isalpha()

                        # replaces all instances of the first argument with the second argument
                        if objfunc == 'replace':

                            # what to replace
                            replacing = self.parse(0, line, f, sp, args)[2]

                            # replacing with
                            wth = self.parse(1, line, f, sp, args)[2]

                            if len(args) == 2:
                                # replaces all instances of replacing with wth
                                self.vars[vname].value = self.vars[vname].value.replace(
                                    replacing, wth)
                            elif len(args) == 3:
                                self.vars[vname].value = self.vars[vname].value.replace(
                                    replacing, wth, self.parse(2, line, f, sp, args)[2])

                            # returns the new string
                            return self.vars[vname].value

                        # strips the value at the variable name
                        if objfunc == 'strip':
                            self.vars[vname].value = self.vars[vname].value.strip()
                            return self.vars[vname].value

                        # obtains a stripped version of itself
                        if objfunc == 'stripped':
                            return self.vars[vname].value.strip()

                        # obtains itself
                        if objfunc == 'self':
                            try:
                                return self.vars[vname].value
                            except:
                                return self.vars[vname]

                        # sets a character in this string
                        if objfunc == 'set':

                            # index to set
                            index = self.parse(0, line, f, sp, args)[2]

                            # what to set it to
                            to_set = self.parse(1, line, f, sp, args)[2]

                            # create a new string with the new character
                            self.vars[vname].value = self.vars[vname].value[:index] + \
                                to_set + self.vars[vname].value[index + 1:]

                            # returns the new string
                            return self.vars[vname].value

                        # gets a character in this string
                        if objfunc == 'get':
                            return self.vars[vname].value[self.parse(0, line, f, sp, args)[2]]

                        # uppercases the string
                        if objfunc == 'upper':
                            self.vars[vname].value = self.vars[vname].value.upper()
                            return self.vars[vname].value

                        # lowercases the string
                        if objfunc == 'lower':
                            self.vars[vname].value = self.vars[vname].value.lower()

                        # cuts a string to the two indices passed
                        if objfunc == 'cut':
                            self.vars[vname].value = self.vars[vname].value[self.parse(
                                0, line, f, sp, args)[2]:self.parse(1, line, f, sp, args)[2]]
                            return self.vars[vname].value

                        # gets a string containing a certain amount of characters left and 
                        # right of the first occurance of the string inside of a string
                        if objfunc == 'around':
                            
                            # keyword to search for
                            keyword = self.parse(0, line, f, sp, args)[2]
                            
                            # amount of characters to get to the left of the keyword
                            left = self.parse(1, line, f, sp, args)[2]
                            
                            # amount of characters to get to the right of the keyword
                            right = self.parse(2, line, f, sp, args)[2]
                            
                            # get the index of the keyword
                            index = object.find(keyword)
                            
                            # if not found
                            if index == -1:
                                return f"around(): Keyword '{keyword}' not found in string"
                            
                            # get the string
                            return object[index-left:index+len(keyword)+right]
                        
                        # startswith
                        if objfunc == 'startswith':
                            return object.startswith(self.parse(0, line, f, sp, args)[2])
                        
                        # endswith
                        if objfunc == 'endswith':
                            return object.endswith(self.parse(0, line, f, sp, args)[2])
                        
                    # working with Excel    
                    elif isinstance(object, self.Workbook):
                        
                        # active workbook
                        workbook = object.workbook
                        path = object.path
                        
                        # gets or creates a sheet in the workbook
                        if objfunc == 'sheet':
                            
                            # title of the new sheet
                            title = self.parse(0, line, f, sp, args)[2]

                            # if title is a string
                            if isinstance(title, str):

                                # if the sheet has already been created,
                                # return the created sheet
                                for name in workbook.sheetnames:
                                    if name.lower() == title.lower():
                                        return self.Sheet(workbook[name], name, workbook, path)
                                
                                # creates the sheet
                                sheet = workbook.create_sheet(title)
                                
                                # returns the sheet
                                return self.Sheet(sheet, title, workbook, path)
                            # title is integer,
                            # return the sheet at that index
                            else:
                                for i, sheet in enumerate(workbook.sheetnames):
                                    if i == title:
                                        return self.Sheet(workbook[sheet], sheet, workbook, path)
                            return None
                                
                            
                        # saves the workbook
                        if objfunc == 'save':
                            workbook.save(path)
                            return object
                        # closes the workbook
                        if objfunc == 'close' or objfunc == 'stop' or objfunc == 'kill':
                            workbook.close()
                            return object
                        
                        # otherwise return the object
                        return object
                        
                            
                    # GENERAL METHODS
                    # gets the immediate children of the parent window
                    def children(parent_window):
                        return [self.AppElement(child, child.window_text()) for child in window.children()]
                    # gets a child at an index
                    # prints the children

                    def child(parent_window, index):
                        child = children(parent_window)[index]
                        return self.AppElement(child, child.window_text())
                    # finds a child with subtext in its name

                    def find_children(parent_window, subtext):
                        subtext = subtext.lower()
                        return [self.AppElement(child, child.window_text()) for child in window.children()
                                if subtext in child.window_text().lower()]
                    
                        
                    # recursively searches the child tree for a certain object type
                    # dont allow ElementAmbiguousError
                    def recursive_search(parent_window, type, as_type, object_string_endswith=None):
                        found = []
                        # get the children
                        # use kwargs to avoid ElementAmbiguousError
                        # kwargs is a criteria to reduce a list by process, class_name, control_type, content_only and/or title.
                        kwargs = {'process': parent_window.process_id()}
                        c = parent_window.children(**kwargs)
                        for child in c:
                            if isinstance(child, type) or (object_string_endswith and str(child).endswith(object_string_endswith)):
                                found += [as_type(child, child.window_text())]
                            found += recursive_search(child, type, as_type, object_string_endswith)
                        return found
                    # prints all elements
                    def print_elements(parent_window, retrieve_elements):
                        for i, element in enumerate(retrieve_elements(parent_window)):
                            print(i, ":")
                            print(element)
                        return None
                    # finds an element containing the substring specified
                    def find_elements(parent_window, subtext, retrieve_elements):
                        elements = []
                        subtext = subtext.lower()
                        for element in retrieve_elements(parent_window):
                            if subtext in element.name.lower():
                                elements.append(self.AppElement(element, element.window_text()))
                        return elements
                    # finds the exact elements specified
                    def find_elements_exact(parent_window, text, retrieve_elements):
                        elements = []
                        for element in retrieve_elements(parent_window):
                            if text == element.name:
                                elements.append(self.AppElement(element, element.window_text()))
                        return elements
                    # waits for the first element to appear containing the substring specified
                    # is not case sensitive
                    def wait_for_element_subtext(parent_window, retrieve_elements, subtext, timeout=None):
                        subtext = subtext.lower()
                        # subfunction for locating the element
                        def find_element_():
                            try:
                                for element in retrieve_elements(parent_window):
                                    if subtext in element.name.lower():
                                        return self.AppElement(element, element.window_text())
                            except:
                                pass
                        if not timeout:
                            while True:
                                if (_ret := find_element_()) is not None:
                                    return _ret
                        else:
                            # get the current time
                            start_time = time.time()
                            # while the time elapsed is less than the timeout
                            while time.time() - start_time < timeout:
                                if (_ret := find_element_()) is not None:
                                    return _ret
                    # waits for the first element to appear with the exact text specified
                    def wait_for_element_exact(parent_window, retrieve_elements, text, timeout=None):
                        # subfunction for locating the element
                        def find_element_():
                            try:
                                for element in retrieve_elements(parent_window):
                                    if text == element.name:
                                        return self.AppElement(element, element.window_text())
                            except:
                                pass
                        if not timeout:
                            while True:
                                if (_ret := find_element_()) is not None:
                                    return _ret
                        else:
                            # get the current time
                            start_time = time.time()
                            # while the time elapsed is less than the timeout
                            while time.time() - start_time < timeout:
                                if (_ret := find_element_()) is not None:
                                    return _ret
                    # waits for the first element to appear in all children containing the substring specified with the type specified
                    def wait_for_type_subtext_all(parent_window, type, as_type, subtext, timeout=None):
                        return wait_for_element_subtext(parent_window, lambda parent_window: recursive_search(parent_window, type, as_type), subtext, timeout=timeout)
                    # wait for the first element to appear in all children with the exact text specified with the type specified
                    def wait_for_type_exact_all(parent_window, type, as_type, text, timeout=None):
                        return wait_for_element_exact(parent_window, lambda parent_window: recursive_search(parent_window, type, as_type), text, timeout=timeout)
                        
                                
                    # waits for a child to exist with text containing subtext
                    def wait_for_text(parent_window, subtext, timeout=None):
                        return wait_for_element_subtext(parent_window, children, subtext, timeout=timeout)
                    # waits for a child to exist in the entire child tree containing subtext
                    def wait_for_text_all(parent_window, subtext, timeout=None):
                        return wait_for_element_subtext(parent_window, all_children, subtext, timeout=timeout)
                    # waits for a child to exist with text exactly equal to text
                    def wait_for_text_exact(parent_window, text, timeout=None):
                        return wait_for_element_exact(parent_window, children, text, timeout=timeout)
                    # waits for a child to exist in the entire child tree with text exactly equal to text
                    def wait_for_text_exact_all(parent_window, text, timeout=None):
                        return wait_for_element_exact(parent_window, all_children, text, timeout=timeout)
                    # prints all children of a parent window
                    def print_children(parent_window):
                        return print_elements(parent_window, children)
                                        
                    # gets all children in the child tree
                    def all_children(parent_window):
                        found = []
                        for child in parent_window.children():
                            found.append(self.AppElement(child, child.window_text()))
                            found += all_children(child)
                        return found
                    # prints the child tree of a parent window
                    def print_all_children(parent_window):
                        return print_elements(parent_window, all_children)
                    # gets from all children at an index
                    def all_child(parent_window, index):
                        return all_children(parent_window)[index]
                    # finds all children with subtext in their name
                    def find_all_children(parent_window, subtext):
                        return find_elements(parent_window, subtext, all_children)
                    # finds all children from an exact text
                    def find_all_children_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, all_children)
                    
                    # ---------------------------
                    
                          
                    # NARROWING GENERAL METHODS
                    # recursively gets all menus existing in the parent_window tree
                    # accumulates all instances of pywinauto.controls.uia_controls.MenuWrapper
                    def menus(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.MenuWrapper, self.Menu)
                    # gets a single menu
                    def menu(parent_window, index):
                        return menus(parent_window)[index]
                    # prints all the menus
                    def print_menus(parent_window):
                        return print_elements(parent_window, menus)
                    # finds a menu with subtext in its name
                    def find_menus(parent_window, subtext):
                        return find_elements(parent_window, subtext, menus)
                    
                    # gets all toolbars
                    def toolbars(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.ToolbarWrapper, self.ToolBar)
                    def print_toolbars(parent_window):
                        return print_elements(parent_window, toolbars)
                    def toolbar(parent_window, index):
                        return toolbars(parent_window)[index]
                    def find_toolbars(parent_window, subtext):
                        return find_elements(parent_window, subtext, toolbars)
                    
                    # recursively gets all instances of pywinauto.controls.uia_controls.ButtonWrapper
                    def buttons(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.ButtonWrapper, self.Button)
                    def button(parent_window, index):
                        return buttons(parent_window)[index]
                    def print_buttons(parent_window):
                        return print_elements(parent_window, buttons)
                    def find_buttons(parent_window, subtext):
                        return find_elements(parent_window, subtext, buttons)
                    
                    # for hyperlinks
                    def links(parent_window):
                        return recursive_search(parent_window, int, self.Link, object_string_endswith="Hyperlink")
                    def link(parent_window, index):
                        return links(parent_window)[index]
                    def print_links(parent_window):
                        return print_elements(parent_window, links)
                    def find_links(parent_window, subtext):
                        return find_elements(parent_window, subtext, links)
                    def find_links_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, links)
                    
                    # for tabitems
                    def tabitems(parent_window):
                        return recursive_search(parent_window, int, self.TabItem, object_string_endswith="TabItem")
                    def tabitem(parent_window, index):
                        return tabitems(parent_window)[index]
                    def print_tabitems(parent_window):
                        return print_elements(parent_window, tabitems)
                    def find_tabitems(parent_window, subtext):
                        return find_elements(parent_window, subtext, tabitems)
                    def find_tabitems_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, tabitems)
                    
                    # for tabcontrols
                    def tabcontrols(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="TabControl")
                    def tabcontrol(parent_window, index):
                        return tabcontrols(parent_window)[index]
                    def print_tabcontrols(parent_window):
                        return print_elements(parent_window, tabcontrols)
                    def find_tabcontrols(parent_window, subtext):
                        return find_elements(parent_window, subtext, tabcontrols)
                    def find_tabcontrols_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, tabcontrols)
                    
                    # for EditWrapper
                    def inputs(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.EditWrapper, self.Input)
                    def input(parent_window, index):
                        return inputs(parent_window)[index]
                    def print_inputs(parent_window):
                        return print_elements(parent_window, inputs)
                    def find_inputs(parent_window, subtext):
                        return find_elements(parent_window, subtext, inputs)
                    def find_inputs_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, inputs)
                    
                    # for ButtonWrapper but endswith CheckBox
                    def checkboxes(parent_window):
                        return recursive_search(parent_window, int, self.Button, object_string_endswith="CheckBox")
                    def checkbox(parent_window, index):
                        return checkboxes(parent_window)[index]
                    def print_checkboxes(parent_window):
                        return print_elements(parent_window, checkboxes)
                    def find_checkboxes(parent_window, subtext):
                        return find_elements(parent_window, subtext, checkboxes)
                    def find_checkboxes_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, checkboxes)
                    
                    # for Image
                    def images(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="Image")
                    def image(parent_window, index):
                        return images(parent_window)[index]
                    def print_images(parent_window):
                        return print_elements(parent_window, images)
                    def find_images(parent_window, subtext):
                        return find_elements(parent_window, subtext, images)
                    def find_images_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, images)
                    
                    # for Tables
                    def tables(parent_window):
                        return recursive_search(parent_window, int, self.Table, object_string_endswith="Table")
                    def table(parent_window, index):
                        return tables(parent_window)[index]
                    def print_tables(parent_window):
                        return print_elements(parent_window, tables)
                    def find_tables(parent_window, subtext):
                        return find_elements(parent_window, subtext, tables)
                    def find_tables_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, tables)
                    
                    # for GroupBoxes
                    def groupboxes(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="GroupBox")
                    def groupbox(parent_window, index):
                        return groupboxes(parent_window)[index]
                    def print_groupboxes(parent_window):
                        return print_elements(parent_window, groupboxes)
                    def find_groupboxes(parent_window, subtext):
                        return find_elements(parent_window, subtext, groupboxes)
                    def find_groupboxes_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, groupboxes)
                    
                    # for Panes
                    def panes(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="Pane")
                    def pane(parent_window, index):
                        return panes(parent_window)[index]
                    def print_panes(parent_window):
                        return print_elements(parent_window, panes)
                    def find_panes(parent_window, subtext):
                        return find_elements(parent_window, subtext, panes)
                    def find_panes_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, panes)
                    
                    # for ListItems
                    def listitems(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.ListItemWrapper, self.AppElement, object_string_endswith="ListItem")
                    def listitem(parent_window, index):
                        return listitems(parent_window)[index]
                    def print_listitems(parent_window):
                        return print_elements(parent_window, listitems)
                    def find_listitems(parent_window, subtext):
                        return find_elements(parent_window, subtext, listitems)
                    def find_listitems_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, listitems)
                    
                    # for documents
                    def documents(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="Document")
                    def document(parent_window, index):
                        return documents(parent_window)[index]
                    def print_documents(parent_window):
                        return print_elements(parent_window, documents)
                    def find_documents(parent_window, subtext):
                        return find_elements(parent_window, subtext, documents)
                    def find_documents_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, documents)
                    
                    # for decendants
                    def descendants(parent_window):
                        return recursive_search(parent_window, int, self.AppElement)
                    # ---------------------------
                    # GENERALIZING METHOD CALLS FOR ELEMENT DISCOVERY
                    def callables(window, 
                                  # array elements
                                  objfunc1, objfunc1_method,
                                  # print the elements 
                                  objfunc2, objfunc2_method,
                                  # get a certain element
                                  objfunc3, objfunc3_method,
                                  # find elements with subtext in their names
                                  objfunc4, objfunc4_method,
                                  # find elements with exact text in their names
                                  objfunc5=None, objfunc5_method=None,
                                  # waits for the first element of a certain type with subtext in name
                                    objfunc6=None, objfunc6_method=None, type1=None, as_type1=None,
                                    # waits for the first element of a certain type with exact text in name
                                    objfunc7=None, objfunc7_method=None, type2=None, as_type2=None,
                                  ):
                        
                        # RETRIEVING CHILDREN
                        # gets the available child reference keywords
                        if objfunc == objfunc1:
                            return objfunc1_method(window)
                        # prints the children
                        if objfunc == objfunc2:
                            return objfunc2_method(window)
                        # gets a certain child
                        # first argument is the index of the child
                        if objfunc == objfunc3:
                            return objfunc3_method(window, self.parse(0, line, f, sp, args)[2])
                        # finds children with subtext in their names
                        if objfunc == objfunc4:
                            return objfunc4_method(window, self.parse(0, line, f, sp, args)[2])
                        if objfunc == objfunc5:
                            return objfunc5_method(window, self.parse(0, line, f, sp, args)[2])
                        
                        # waits for the first child of a certain type with exact text in its name
                        if objfunc == objfunc6:
                            # if 1 argument, there is no timeout
                            if len(args) == 1:
                                return wait_for_type_exact_all(window, type1, as_type1, self.parse(0, line, f, sp, args)[2])
                            elif len(args) == 2:
                                return wait_for_type_exact_all(window, type1, as_type1, self.parse(0, line, f, sp, args)[2], self.parse(1, line, f, sp, args)[2])
                        # waits for the first child of a certain type with subtext in its name
                        if objfunc == objfunc7:
                            # if 1 argument, there is no timeout
                            if len(args) == 1:
                                return wait_for_type_subtext_all(window, type2, as_type2, self.parse(0, line, f, sp, args)[2])
                            elif len(args) == 2:
                                return wait_for_type_subtext_all(window, type2, as_type2, self.parse(0, line, f, sp, args)[2], self.parse(1, line, f, sp, args)[2])
                            
                        
                        return '<msnint2 no callable>'
                        
                    
                    
                    # ---------------------------

                    
                    # moves the mouse to the center of an element, and clicks it
                    def clk(window, button='left', waittime=0):
                        # set the focus to this element
                        window.set_focus()
                        
                        # wait for the element to be ready
                        time.sleep(waittime)
                
                        # get the new coordinates of this element after the focus
                        coords = window.get_properties()['rectangle'].mid_point()
                        
                        # click the mouse
                        mouse.click(button=button, coords=coords)

                        # return the object
                        return object   
                    
                    # determines if a point is visible within a rectangle
                    def has_point(object, x, y):
                        try:
                            rect = object.get_properties()['rectangle']
                            # if implemented
                            return rect.top <= y <= rect.bottom and rect.left <= x <= rect.right
                        except:
                            print(str(object))
                            return True
                    # recursively get the first object that has the point
                    # the first object that has the point and no children
                    def rec(root, x, y):
                        
                        # if the root has children
                        if root.children():
                            # for each child
                            for child in root.children():
                                # if the child has the point
                                if has_point(child, x, y):
                                    # return the child
                                    return rec(child, x, y)
                        # if the root has no children
                        else:
                            # return the root
                            return self.AppElement(root, root.window_text())
                    
                    # get all objects that have the point
                    def get_all(root, x, y):
                        all = []
                        # if the root has children
                        if root.children():
                            # for each child
                            for child in root.children():
                                # if the child has the point
                                if has_point(child, x, y):
                                    # add the child to the list
                                    all.append(self.AppElement(child, child.window_text()))
                                    # get all of the child's children
                                    all += get_all(child, x, y)
                        # return the list
                        return all

                    # presses multiple keys at the same time
                    def press_simul(kys):
                        sending = ''
                        # keys down
                        for key in kys:
                            sending += '{' + key + ' down}'
                        # keys up
                        for key in kys:
                            sending += '{' + key + ' up}'
                        return sending
                    
                    # function for converting keys requiring a shift press
                    #   example: a '3' should be converted to {VK_SHIFT down}3{VK_SHIFT up}
                    #   example: a '"' should be converted to {VK_SHIFT down}'{VK_SHIFT up}
                    #   example: a 'E' should be converted to {VK_SHIFT down}e{VK_SHIFT up}
                    # this function is mainly for converting an exerpt of code to a typable
                    # string for pywinauto to type
                    def convert_keys(keystrokes):
                        new = ''
                        special = {
                            '!' : '1',
                            '@' : '2',
                            '#' : '3',
                            '$' : '4',
                            '%' : '5',
                            '^' : '6',
                            '&' : '7',
                            '*' : '8',
                            '(' : '9',
                            ')' : '0',
                            '_' : '-',
                            '+' : '=',
                            '{' : '[',
                            '}' : ']',
                            '|' : '\\',
                            ':' : ';',
                            '"' : "'",
                            '<' : ',',
                            '>' : '.',
                            '?' : '/',
                            '~' : '`'
                        }
                        # for each keystroke
                        for key in keystrokes:
                            if key in special:
                                # if the key is a special character
                                new += '{VK_SHIFT down}' + special[key] + '{VK_SHIFT up}'
                            elif key.isupper():
                                # if the key is uppercase
                                new += '{VK_SHIFT down}' + key.lower() + '{VK_SHIFT up}'
                            else:
                                # if the key is not a special character
                                new += key
                        return new
                    
                    # types keys with a delay between each key
                    def type_keys_with_delay(window, text, delay):
                        e = False
                        for char in text:
                            try:
                                window.type_keys(char, with_spaces=True)
                            except:
                                if not e:
                                    window.set_focus()
                                    e = True
                                pywinauto.keyboard.send_keys(char)
                            time.sleep(delay)
                            
                    # parses object functions for discovering types
                    # of elements
                    def search(window):
                        ret = '<msnint2 no callable>'
                        
                         # RETRIEVING CHILDREN
                        # gets the available child reference keywords
                        if  (chldrn := callables(window,
                                    'children', children,
                                    'print_children', print_children,
                                    'child', child,
                                    'find_children', find_children)) != '<msnint2 no callable>': 
                            ret = chldrn
                        
                        # working with the entire child tree
                        elif (all_chldrn := callables(window,
                                    'all_children', all_children,
                                    'print_all_children', print_all_children,
                                    'all_child', all_child,
                                    'find_all_children', find_all_children,
                                    'find_all_children_exact', find_all_children_exact, 
                                    objfunc6='wait_for_child', objfunc6_method=wait_for_type_exact_all,
                                        type1=pywinauto.controls.uiawrapper.UIAWrapper,
                                        as_type1=self.AppElement,
                                    objfunc7='wait_for_child_exact', objfunc7_method=wait_for_type_subtext_all,
                                        type2=pywinauto.controls.uiawrapper.UIAWrapper,
                                        as_type2=self.AppElement
                                        
                                    )) != '<msnint2 no callable>': 
                            ret = all_chldrn

                        # getting all menus
                        elif (mns := callables(window,
                                    'menus', menus,
                                    'print_menus', print_menus,
                                    'menu', menu,
                                    'find_menus', find_menus, 
                                    objfunc5=None, objfunc5_method=None,
                                    objfunc6='wait_for_menu_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=pywinauto.controls.uia_controls.MenuWrapper,
                                        as_type1=self.Menu,
                                    objfunc7='wait_for_menu', objfunc7_method=wait_for_type_subtext_all,
                                        type2=pywinauto.controls.uia_controls.MenuWrapper,
                                        as_type2=self.Menu
                                    )) != '<msnint2 no callable>': 
                            ret = mns
                        
                        # gets all toolbars
                        elif (tbrs := callables(window,
                                    'toolbars', toolbars,
                                    'print_toolbars', print_toolbars,
                                    'toolbar', toolbar,
                                    'find_toolbars', find_toolbars, 
                                    objfunc5=None, objfunc5_method=None,
                                    objfunc6='wait_for_toolbar_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=pywinauto.controls.uia_controls.ToolbarWrapper,
                                        as_type1=self.ToolBar,
                                    objfunc7='wait_for_toolbar', objfunc7_method=wait_for_type_subtext_all,
                                        type2=pywinauto.controls.uia_controls.ToolbarWrapper,
                                        as_type2=self.ToolBar
                                    )) != '<msnint2 no callable>': 
                            ret = tbrs
                        
                        # gets all buttons
                        elif (btns := callables(window,
                                    'buttons', buttons,
                                    'print_buttons', print_buttons,
                                    'button', button,
                                    'find_buttons', find_buttons,
                                    objfunc5=None, objfunc5_method=None,
                                    objfunc6='wait_for_button_exact', objfunc6_method=wait_for_type_exact_all, 
                                        type1=pywinauto.controls.uia_controls.ButtonWrapper, 
                                        as_type1=self.Button,
                                    objfunc7='wait_for_button', objfunc7_method=wait_for_type_subtext_all, 
                                        type2=pywinauto.controls.uia_controls.ButtonWrapper,
                                        as_type2=self.Button
                                    )) != '<msnint2 no callable>':  
                            ret = btns
                        
                        # gets all tabitems
                        elif (tbs := callables(window,
                                    'tabitems', tabitems,
                                    'print_tabitems', print_tabitems,
                                    'tabitem', tabitem,
                                    'find_tabitems', find_tabitems, 
                                    objfunc5=None, objfunc5_method=None,
                                    objfunc6='wait_for_tabitem_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=int,
                                        as_type1=self.TabItem,
                                    objfunc7='wait_for_tabitem', objfunc7_method=wait_for_type_subtext_all,
                                        type2=int,
                                        as_type2=self.TabItem  
                                    )) != '<msnint2 no callable>':
                            ret = tbs
                        
                        # gets all links
                        elif (lnks := callables(window,
                                    'links', links,
                                    'print_links', print_links,
                                    'link', link,
                                    'find_links', find_links,
                                    objfunc5=None, objfunc5_method=None,
                                    objfunc6='wait_for_link_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=int,
                                        as_type1=self.Hyperlink,
                                    objfunc7='wait_for_link', objfunc7_method=wait_for_type_subtext_all,
                                        type2=int,
                                        as_type2=self.Hyperlink
                                    )) != '<msnint2 no callable>': 
                            ret = lnks
                        
                        # gets all Inputs
                        elif (inpts := callables(window,
                                    'inputs', inputs,
                                    'print_inputs', print_inputs,
                                    'input', input,
                                    'find_inputs', find_inputs,
                                    objfunc6='wait_for_input_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=pywinauto.controls.uia_controls.EditWrapper,
                                        as_type1=self.Input,
                                    objfunc7='wait_for_input', objfunc7_method=wait_for_type_subtext_all,
                                        type2=pywinauto.controls.uia_controls.EditWrapper,
                                        as_type2=self.Input
                            )) != '<msnint2 no callable>': 
                            ret = inpts
                        
                        # gets all checkboxes
                        elif (chks := callables(window,
                                    'checkboxes', checkboxes,
                                    'print_checkboxes', print_checkboxes,
                                    'checkbox', checkbox,
                                    'find_checkboxes', find_checkboxes,
                                    objfunc6='wait_for_checkbox_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=pywinauto.controls.uia_controls.ButtonWrapper,
                                        as_type1=self.Button,
                                    objfunc7='wait_for_checkbox', objfunc7_method=wait_for_type_subtext_all,
                                        type2=pywinauto.controls.uia_controls.ButtonWrapper,
                                        as_type2=self.Button
                            )) != '<msnint2 no callable>': 
                            ret = chks
                        
                        # gets all images
                        elif (imgs := callables(window,
                                    'images', images,
                                    'print_images', print_images,
                                    'image', image,
                                    'find_images', find_images)) != '<msnint2 no callable>': 
                            ret = imgs
                            
                        # gets all tables
                        elif (tbls := callables(window,
                                    'tables', tables,
                                    'print_tables', print_tables,
                                    'table', table,
                                    'find_tables', find_tables,
                                    objfunc6='wait_for_table_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=pywinauto.controls.uia_controls.ListViewWrapper,
                                        as_type1=self.Table,
                                    objfunc7='wait_for_table', objfunc7_method=wait_for_type_subtext_all,
                                        type2=pywinauto.controls.uia_controls.ListViewWrapper,
                                        as_type2=self.Table
                                    )) != '<msnint2 no callable>': 
                            ret = tbls
                        
                        # get all GroupBoxes
                        elif (grps := callables(window,
                                    'groupboxes', groupboxes,
                                    'print_groupboxes', print_groupboxes,
                                    'groupbox', groupbox,
                                    'find_groupboxes', find_groupboxes,
                                    objfunc6='wait_for_groupbox_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=int,
                                        as_type1=self.AppElement,
                                    objfunc7='wait_for_groupbox', objfunc7_method=wait_for_type_subtext_all,
                                        type2=int,
                                        as_type2=self.AppElement
                                    )) != '<msnint2 no callable>': 
                            ret = grps
                            
                        # for Panes
                        elif (pns := callables(window,
                                    'panes', panes,
                                    'print_panes', print_panes,
                                    'pane', pane,
                                    'find_panes', find_panes,
                                    objfunc6='wait_for_pane_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=int,
                                        as_type1=self.AppElement,
                                    objfunc7='wait_for_pane', objfunc7_method=wait_for_type_subtext_all,
                                        type2=int,
                                        as_type2=self.AppElement
                                    )) != '<msnint2 no callable>':
                            ret = pns
                            
                        # for ListItems
                        elif (lsts := callables(window,
                                    'listitems', listitems,
                                    'print_listitems', print_listitems,
                                    'listitem', listitem,
                                    'find_listitems', find_listitems,
                                    objfunc6='wait_for_listitem_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=pywinauto.controls.uia_controls.ListItemWrapper,
                                        as_type1=self.AppElement,
                                    objfunc7='wait_for_listitem', objfunc7_method=wait_for_type_subtext_all,
                                        type2=pywinauto.controls.uia_controls.ListItemWrapper,
                                        as_type2=self.AppElement
                                    )) != '<msnint2 no callable>':
                            ret = lsts
                        
                        # for TabControls
                        elif (tabs := callables(window,
                                    'tabcontrols', tabcontrols,
                                    'print_tabcontrols', print_tabcontrols,
                                    'tabcontrol', tabcontrol,
                                    'find_tabcontrols', find_tabcontrols,
                                    objfunc6='wait_for_tabcontrol_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=int,
                                        as_type1=self.AppElement,
                                    objfunc7='wait_for_tabcontrol', objfunc7_method=wait_for_type_subtext_all,
                                        type2=int,
                                        as_type2=self.AppElement
                                    )) != '<msnint2 no callable>':
                            ret = tabs
                        
                        # for Documents
                        elif (docs := callables(window,
                                    'documents', documents,
                                    'print_documents', print_documents,
                                    'document', document,
                                    'find_documents', find_documents,
                                    objfunc6='wait_for_document_exact', objfunc6_method=wait_for_type_exact_all,
                                        type1=int,
                                        as_type1=self.AppElement,
                                    objfunc7='wait_for_document', objfunc7_method=wait_for_type_subtext_all,
                                        type2=int,
                                        as_type2=self.AppElement
                                    )) != '<msnint2 no callable>':
                            ret = docs
                            
                        return ret

                    # if the object is a pywinauto application
                    # KNOWN ISSUES:
                    #   - I've tested this on a Windows 11 laptop and it doesn't
                    #     work for some reason
                    if isinstance(object, self.App):
                        
                        # return for an app
                        ret = object

                        # path to the application to work with
                        path = object.path
                        # actual pwinauto application object
                        app = object.application
                        # window
                        window = app.window() if app else None

                        # thread based operation
                        p_thread = False
                        if objfunc.endswith(':lock'):
                            p_thread = True
                            objfunc = objfunc[:-5]
                            auto_lock.acquire()
                        
                        # element discovery with search()
                        if (srch := search(window)) != '<msnint2 no callable>':
                            ret = srch


                        # STARTING AND STOPPING APPLICATIONS
                        # creates and starts the application
                        # TODO
                        if objfunc == 'start':
                            # create and start the application
                            if not object.application:
                                object.application = Application(backend="uia").start(path)

                            # add to global apps
                            global apps
                            apps[len(apps) + 1] = object

                            ret = object.application
                        # kills the application
                        elif objfunc == 'stop' or objfunc == 'kill' or objfunc == 'close':
                            # kill the application
                            ret = app.kill()

                            
                        
                        
                        # gets the top_window
                        elif objfunc == 'print_tree':
                            ret = app.dump_tree()
                            
                        # gets a connection to this application
                        elif objfunc == 'connection':
                            ret = self.App(object.path, Application(backend="uia").connect(process=object.application.process))
                                                    

                        # gets information about this application
                        # gets the text of the window
                        elif objfunc == 'text':
                            ret = window.window_text()
                        # gets the window
                        elif objfunc == 'window':
                            ret = window
                        # gets the handle
                        elif objfunc == 'handle':
                            ret = window.handle
                        
                        # chrome based children collection
                        def chrome_children_():
                            chrome_window = app.window(title_re='.*Chrome.')
                            chrome_handle = chrome_window.handle
                            wd = app.window(handle=chrome_handle)
                            document = wd.child_window(found_index=0, class_name='Chrome_RenderWidgetHostHWND')
                            return document.descendants()
                        
                        # GOOGLE CHROME ONLY
                        if objfunc == 'chrome_children':
                            # if not arguments
                            if args[0][0] == '':
                                ret = chrome_children_()
                            # if one argument, check if the first argument is contained
                            elif len(args) == 1:
                                subtext = self.parse(0, line, f, sp, args)[2].lower()
                                ret = [self.AppElement(d, d.window_text()) for d in chrome_children_() if subtext in d.window_text().lower()]
                            # if two arguments, check if the first argument is exact
                            elif len(args) == 2:
                                subtext = self.parse(0, line, f, sp, args)[2]
                                ret = [self.AppElement(d, d.window_text()) for d in chrome_children_() if subtext == d.window_text()]
                        
                        # waits for a child containing text
                        elif objfunc == 'wait_for_text':
                            # if no timeout provided
                            if len(args) == 1:
                                ret = wait_for_text(window, self.parse(0, line, f, sp, args)[2])
                            # if timeout provided
                            elif len(args) == 2:
                                ret = wait_for_text(window, self.parse(0, line, f, sp, args)[2], 
                                                     timeout=self.parse(1, line, f, sp, args)[2])
                        # waits for a child containing text in the entire child tree
                        elif objfunc == 'wait_for_text_all':
                            # if no timeout provided
                            if len(args) == 1:
                                ret = wait_for_text_all(window, self.parse(0, line, f, sp, args)[2])
                            elif len(args) == 2:
                                ret = wait_for_text_all(window, self.parse(0, line, f, sp, args)[2], 
                                                     timeout=self.parse(1, line, f, sp, args)[2])
                                
                        # waits for a child containing the exact text
                        elif objfunc == 'wait_for_text_exact':
                            # if no timeout provided
                            if len(args) == 1:
                                ret = wait_for_text_exact(window, self.parse(0, line, f, sp, args)[2])
                            elif len(args) == 2:
                                ret = wait_for_text_exact(window, self.parse(0, line, f, sp, args)[2], 
                                                     timeout=self.parse(1, line, f, sp, args)[2])
                        # waits for a child containing the exact text in the entire child tree
                        elif objfunc == 'wait_for_text_exact_all':
                            # if no timeout provided
                            if len(args) == 1:
                                ret = wait_for_text_exact_all(window, self.parse(0, line, f, sp, args)[2])
                            elif len(args) == 2:
                                ret = wait_for_text_exact_all(window, self.parse(0, line, f, sp, args)[2], 
                                                     timeout=self.parse(1, line, f, sp, args)[2])

                        # APPLICATION ACTIONS
                        # sends keystrokes to the application
                        # takes one argument, being the keystrokes to send
                        elif objfunc == 'write':
                            writing = self.parse(0, line, f, sp, args)[2]
                            try:
                                # sends keystrokes to the application
                                ret = window.type_keys(writing, with_spaces=True)
                            except:
                                # with_spaces not allowed
                                ret = window.type_keys(writing)
                        # writes special characters into the console
                        # takes one argument, being the special characters to write
                        elif objfunc == 'write_special':
                            # keystrokes
                            keystrokes = self.parse(0, line, f, sp, args)[2]
                            # convert to special characters
                            ret = window.type_keys(convert_keys(keystrokes), with_spaces=True)
                            
                        # presses keys at the same time
                        elif objfunc == 'press':
                            kys = []
                            for i in range(len(args)):
                                kys.append(self.parse(i, line, f, sp, args)[2])
                            # presses the keys at the same time
                            ret = window.type_keys(press_simul(kys))
                        # sends keystrokes to the application
                        # takes one argument, being the keystrokes to send
                        elif objfunc == 'send_keys':
                            # sends keystrokes to the application
                            ret = pywinauto.keyboard.send_keys(convert_keys(self.parse(0, line, f, sp, args)[2]), with_spaces=True)
                        
                        # gets the element that is currently hovered over
                        # recurses through all children, determining which elements have
                        # the mouses position
                        elif objfunc == 'hovered':                            
                            # get the root window of this application
                            root = window.top_level_parent()
                            
                            # get the current mouse position
                            x, y = win32api.GetCursorPos()
                            
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = get_all(root, x, y)     
                        
                        # opens the developer tools
                        elif objfunc == 'inspect':
                            # presses the shortcut keys to open the developer tools
                            ret = window.type_keys('{F12}')
                            
                            # waits for the inspect window to appear
                            wait_for_text_all(window, 'Console')
                            
                        # closes the developer tools
                        elif objfunc == 'close_inspect':
                            # presses the shortcut keys to close the developer tools
                            ret = window.type_keys('{F12}')
                        
                        # refreshes the page
                        elif objfunc == 'refresh':
                            # presses the shortcut keys to refresh the page
                            ret = window.type_keys('{F5}')
                        
                        # presses the enter key
                        elif objfunc == 'enter':
                            # presses the enter key
                            ret = window.type_keys('{ENTER}')
                        # presses the escape key
                        elif objfunc == 'escape':
                            # presses the escape key
                            ret = window.type_keys('{ESC}')
                        # page down
                        elif objfunc == 'page_down':
                            # presses the page down key
                            ret = window.type_keys('{PGDN}')
                        # page up
                        elif objfunc == 'page_up':
                            # presses the page up key
                            ret = window.type_keys('{PGUP}')
                        
                        # # collects all children within the entire page
                        # # finds all scrollbars and scrolls throughout the entire page
                        # # in all directions, collecting all children
                        # if objfunc == 'collect_children':
                        #     chld = set()
                        #     # gets all scrollbars
                        #     scrlbrs = window.scrollbars()
                        #     # get the height of the screen
                        #     height = window.rectangle().height
                            
                        #     return scrlbrs
                        
                        # # gets the vertical scrollbar
                        # if objfunc == 'vertical_scrollbar':
                        #     wd = window.child_window(control_type="Scroll")
                        #     return self.ScrollBar(wd, 'vertical')
                                
                        # release auto_lock
                        if p_thread:
                            auto_lock.release()

                        # return the object
                        return ret

                    # if the object is a pywinauto window element
                    elif isinstance(object, self.AppElement):
                        # returning
                        ret = object
                        # get the window of the AppElement object
                        window = object.window
                        # get the text of the AppElement object
                        name = object.name

                        # function to move the mouse from start to end,
                        # with a speed of speed
                        def movemouse(start, end, speed):
                            # reverse the speed, so a speed of 50 gives
                            # end_range of 50, and a speed of 75 gives
                            # end_range of 25
                            # dragging the mouse
                            # presses the mouse down at the coordinates
                            mouse.press(coords=start)
                            end_range = 100 - speed
                            for i in range(0, end_range):
                                mouse.move(coords=(int(start[0] + (end[0] - start[0]) / 100 * i), 
                                                    int(start[1] + (end[1] - start[1]) / 100 * i)))
                                time.sleep(0.001)
                                    
                            # releases the mouse at the end coordinates
                            mouse.release(coords=end)


                        p_thread = False

                        # thread based functions
                        # if the function is a thread based function
                        if objfunc.endswith(':lock'):
                            p_thread = True
                            auto_lock.acquire()
                            objfunc = objfunc[:-5]

                        # OBTAINING DIFFERENT TYPES OF CHILDREN
                        # get the element window
                        if objfunc == 'window':
                            ret = window
                        
                        # element discovery with search()
                        if (srch := search(window)) != '<msnint2 no callable>':
                            ret = srch

                        # getting information about the current window
                        # gets the window text
                        elif objfunc == 'text':
                            ret = window.window_text()
                        # GETTING LOCATION OF THE WINDOW
                        elif objfunc == 'top':
                            ret = window.get_properties()['rectangle'].top
                        elif objfunc == 'bottom':
                            ret = window.get_properties()['rectangle'].bottom
                        elif objfunc == 'left':
                            ret = window.get_properties()['rectangle'].left
                        elif objfunc == 'right':
                            ret = window.get_properties()['rectangle'].right
                        elif objfunc == 'center' or objfunc == 'mid_point':
                            ret = window.get_properties()['rectangle'].mid_point()
                        # getting the rectangle overall
                        elif objfunc == 'rectangle':
                            ret = [window.get_properties()['rectangle'].top, window.get_properties()['rectangle'].bottom, window.get_properties()['rectangle'].left, window.get_properties()['rectangle'].right]

                        # computes the diameter of the window
                        elif objfunc == 'width':
                            try:
                                left = window.get_properties()['rectangle'].left
                                right = window.get_properties()['rectangle'].right
                                ret = right - left
                            except:
                                ret = None
                        # computes the height of the window
                        elif objfunc == 'height':
                            try:
                                top = window.get_properties()['rectangle'].top
                                bottom = window.get_properties()['rectangle'].bottom
                                ret = bottom - top
                            except:
                                ret = None 
                            
                        # getting adjacent elements
                        # could or could not be decendants
                        # operation is very slow, should be used mainly
                        # for element discovery
                        elif objfunc == 'element_above':
                            # pixels above
                            pixels = self.parse(0, line, f, sp, args)[2]
                            # get the root window of this application
                            root = object.top_level_parent()
                            
                            # get the top middle point of this element
                            top = object.get_properties()['rectangle'].top - pixels
                            mid = object.get_properties()['rectangle'].mid_point()[0]
                            # if there exist two arguments, move the mouse to that location
                            if len(args) == 2:
                                mouse.move(coords=(mid, top))
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = rec(root, mid, top)
                        elif objfunc == 'element_below':
                            # pixels above
                            pixels = self.parse(0, line, f, sp, args)[2]
                            # get the root window of this application
                            root = object.top_level_parent()
                            # get the top middle point of this element
                            bottom = object.get_properties()['rectangle'].bottom + pixels
                            mid = object.get_properties()['rectangle'].mid_point()[0]
                            if len(args) == 2:
                                mouse.move(coords=(mid, bottom))
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = rec(root, mid, bottom)
                        elif objfunc == 'element_left':
                            # pixels to the left
                            pixels = self.parse(0, line, f, sp, args)[2]
                            # get the root window of this application
                            root = object.top_level_parent()
                            
                            # get the left middle point of this element
                            left = object.get_properties()['rectangle'].left - pixels
                            mid = object.get_properties()['rectangle'].mid_point()[1]
                            if len(args) == 2:
                                mouse.move(coords=(left, mid))
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = rec(root, left, mid)
                        elif objfunc == 'element_right':
                            # pixels to the right
                            pixels = self.parse(0, line, f, sp, args)[2]
                            # get the root window of this application
                            root = object.top_level_parent()
                            
                            # get the right middle point of this element
                            right = object.get_properties()['rectangle'].right + pixels
                            mid = object.get_properties()['rectangle'].mid_point()[1]
                            if len(args) == 2:
                                mouse.move(coords=(right, mid))
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = rec(root, right, mid)
                        
                        # focus on the window
                        elif objfunc == 'focus':
                            ret = window.set_focus()
                        
                        # scrolls to the window
                        elif objfunc == 'scroll':
                            ret = mouse.scroll(coords=(window.get_properties()['rectangle'].mid_point()[0],
                                                       window.get_properties()['rectangle'].mid_point()[1]))
                        
                        # drags this element to either another AppElement
                        elif objfunc == 'drag':                            
                            # if one argument and that argument isinstance(AppElement)
                            first = self.parse(0, line, f, sp, args)[2]
                            # midpoint of the element to drag to
                            start = (window.get_properties()['rectangle'].mid_point()[0],
                                        window.get_properties()['rectangle'].mid_point()[1])
                            end = (first.get_properties()['rectangle'].mid_point()[0],
                                        first.get_properties()['rectangle'].mid_point()[1])

                            # slowly moves the mouse to the end coordinates
                            # this is to prevent the mouse from moving too fast
                            # and not dragging the object
                            # the farther the distance, the longer it takes
                            # to move the mouse
                            speed = 50
                            if len(args) == 2:
                                speed = self.parse(1, line, f, sp, args)[2]
                            # drags the mouse
                            movemouse(start, end, speed) 
                            ret = True
                            
                        # drags this AppElement to coordinates
                        elif objfunc == 'drag_coords':
                            
                            start = (window.get_properties()['rectangle'].mid_point()[0],
                                      window.get_properties()['rectangle'].mid_point()[1])
                            end = (self.parse(0, line, f, sp, args)[2], 
                                     self.parse(1, line, f, sp, args)[2])
                            
                            # gets the speed, if specified
                            speed = 50
                            if len(args) == 3:
                                speed = self.parse(2, line, f, sp, args)[2]
                            # drags the mouse
                            movemouse(start, end, speed)
                            ret = True
                            
                        # WINDOW ACTIONS
                        # sends keystrokes to the application
                        # takes one argument, being the keystrokes to send
                        elif objfunc == 'write':
                            writing = self.parse(0, line, f, sp, args)[2]
                            timeout = False
                            # if a timeout between keystrokes is offered
                            if len(args) == 2:
                                timeout = True
                            
                            if timeout:
                                ret = type_keys_with_delay(window, writing, self.parse(1, line, f, sp, args)[2])
                            else:
                                try:
                                    ret = window.type_keys(writing, with_spaces=True)
                                except:
                                    window.set_focus()
                                    ret = window.type_keys(writing)
                                    
                        # presses backspace
                        # if no arguments, presses it one time
                        # else, presses it the first argument many times
                        if objfunc == 'backspace':
                            window.set_focus()
                            # no argument
                            if args[0][0] == '':
                                ret = window.type_keys('{BACKSPACE}')
                            # else, send {BACKSPACE} that many times
                            else:
                                times = self.parse(0, line, f, sp, args)[2]
                                ret = window.type_keys('{BACKSPACE}' * times)

                        # presses the enter key
                        elif objfunc == 'enter':
                            ret = window.type_keys('{ENTER}')
                        
                        # hovers over the window
                        elif objfunc == 'hover':
                            # hovers the mouse over the window, using the mid point of the element
                            ret = mouse.move(coords=(window.get_properties()['rectangle'].mid_point()))                       
                            
                        # different types of AppElements
                        # if the appelement is a button
                        if isinstance(object, self.Button):
                            
                            # clicks the button
                            if objfunc == 'click' or objfunc == 'left_click':
                                ret = object.click()
                            # left clicks the button
                            elif objfunc == 'right_click':
                                ret = object.right_click()
                            ret = object

                        # working with Links
                        elif isinstance(object, self.Link):           
                            waittime = self.parse(0, line, f, sp, args)[2] if args[0][0] != '' else 1                 
                            # clicks the link
                            if objfunc == 'click' or objfunc == 'left_click':
                                ret = clk(window, waittime=waittime)
                            # right clicks the link
                            elif objfunc == 'right_click':
                                ret = clk(window, button='right', waittime=waittime)
                            ret = object
                            
                        # working with Tables
                        elif isinstance(object, self.Table):
                            # get table
                            table = object.window
                            
                            # gets a row by index, based on the above logic
                            def row(index):
                                row = []
                                items = []
                                try:
                                    cols = table.column_count()
                                except NotImplementedError:
                                    # not implemented
                                    cols = 5
                                    items = table.items()
                                for i in range(cols):
                                    try:
                                        try:
                                            wrapper = table.cell(row=index, column=i)
                                        except:
                                            # table.items() gets a 1D list of items,
                                            # compute the index of the item
                                            # based on 'i' and 'index'
                                            wrapper = items[i + index * cols]
                                            
                                        row.append(self.AppElement(wrapper, wrapper.window_text()))
                                    except:
                                        break
                                return row
                            # gets a column by index
                            def col(index):
                                col = []
                                for i in range(table.column_count()):
                                    try:
                                        wrapper = table.cell(row=i, column=index)
                                        col.append(self.AppElement(wrapper, wrapper.window_text()))
                                    except:
                                        break
                                return col
                            
                            
                            # gets a cell at a row and column
                            if objfunc == 'get':
                                
                                # get column
                                col = self.parse(0, line, f, sp, args)[2]
                                # get row
                                row = self.parse(1, line, f, sp, args)[2]
                                wrapper = table.cell(row=row, column=col)
                                # gets the cell
                                ret = self.AppElement(wrapper, wrapper.window_text())
                                
                            # try to accumulate all the rows
                            # up to sys.maxsize
                            elif objfunc == 'matrix':
                                matrix = []
                                for i in range(sys.maxsize):
                                    try:
                                        if (_r := row(i)):
                                            matrix.append(_r)
                                        else:
                                            break
                                    except:
                                        break
                                ret = matrix                                   
                            # gets a row
                            elif objfunc == 'row':
                                ret = row(self.parse(0, line, f, sp, args)[2])
                            # gets a column
                            elif objfunc == 'column':
                                ret = col(self.parse(0, line, f, sp, args)[2])
                                
                            

                            

                            
                        # working with ToolBars
                        elif isinstance(object, self.ToolBar):
                            toolbar_window = object.window
                            # gets the buttons of the toolbar
                            if objfunc == 'buttons':
                                ret = [toolbar_window.button(i) for i in range(toolbar_window.button_count())]
                            # prints the buttons of this toolbar
                            if objfunc == 'print_buttons':
                                for i in range(toolbar_window.button_count()):
                                    print(i, ':', toolbar_window.button(i))
                                ret = None
                            # gets a button at an index
                            if objfunc == 'button':
                                ret = toolbar_window.button(self.parse(0, line, f, sp, args)[2])
                            # finds all buttons with subtext in their names
                            if objfunc == 'find_buttons':
                                ret = find_buttons(toolbar_window, self.parse(0, line, f, sp, args)[2])
                            ret = object
                        
                        # working with scrollbars
                        elif isinstance(object, self.ScrollBar):
                            scrollbar_window = object.window
                            
                            if objfunc == 'scroll_down':
                                ret = scrollbar_window.scroll_down(amount='page', count=1)
                                       
                        # extra methods such that this AppElement requires different logic
                        if objfunc == 'click' or objfunc == 'left_click':
                            ret = clk(window, waittime=self.parse(0, line, f, sp, args)[2] if args[0][0] != '' else 1)              
                        elif objfunc == 'right_click':
                            ret = clk(window, button='right', waittime=self.parse(0, line, f, sp, args)[2] if args[0][0] != '' else 1)


                        # if thread based, release the lock
                        if p_thread:
                            auto_lock.release()
                        
                        return ret

                    # if the object is a dictionary
                    elif isinstance(object, dict):

                        # allows for repetitive setting on a multiple indexed dictionary
                        if objfunc == 'set':
                            self.vars[vname].value[self.parse(0, line, f, sp, args)[
                                2]] = self.parse(1, line, f, sp, args)[2]
                            return self.vars[vname].value

                        # first argument is what to set, should be called to_set
                        # rest of the arguments are the indices at which to index the object and set to_set
                        if objfunc == 'setn':

                            # what to set
                            to_set = self.parse(0, line, f, sp, args)[2]

                            # the rest of the arguments are the indices
                            # example: dict.setn('im being set', 'index1', 'index2', 'index3', ...)
                            # should equal: dict['index1']['index2']['index3'] = 'im being set'

                            # the object to set
                            obj = self.vars[vname].value

                            # iterates through the indices
                            for i in range(1, len(args)):

                                # if the index is the last one
                                if i == len(args) - 1:

                                    # sets the index to to_set
                                    obj[self.parse(i, line, f, sp, args)[
                                        2]] = to_set

                                # if the index is not the last one
                                else:

                                    # sets the object to the index
                                    obj = obj[self.parse(
                                        i, line, f, sp, args)[2]]

                            # returns the object
                            return self.vars[vname].value

                        # recursively gets a value in a dictionary
                        if objfunc == 'get':

                            # the object to get from
                            obj = self.vars[vname].value

                            # iterates through the indices
                            for i in range(len(args)):

                                # sets the object to the index
                                obj = obj[self.parse(i, line, f, sp, args)[2]]

                            # returns the object
                            return obj

                        # gets the keys of this dictionary
                        if objfunc == 'keys':
                            return self.vars[vname].value.keys()

                        # gets the values of this dictionary
                        if objfunc == 'values':
                            return self.vars[vname].value.values()

                        # gets the items of this dictionary
                        if objfunc == 'items':
                            return self.vars[vname].value.items()

                        # executes a function for each key-value pair
                        if objfunc == 'foreach':

                            # variable name of the key
                            keyname = self.parse(0, line, f, sp, args)[2]

                            # variable name of the value
                            valuename = self.parse(1, line, f, sp, args)[2]

                            # function to execute
                            function = args[2][0]

                            # loop through the dictionary
                            for key, value in self.vars[vname].value.items():

                                # set the key and value variables
                                self.vars[keyname] = Var(keyname, key)
                                self.vars[valuename] = Var(valuename, value)

                                # execute the function
                                self.interpret(function)

                            # return the dictionary
                            return self.vars[vname].value

                        # maps each value in the dictionary to the output of the function
                        if objfunc == 'map':

                            # map arguments
                            keyvarname = self.parse(0, line, f, sp, args)[2]
                            valuevarname = self.parse(1, line, f, sp, args)[2]
                            function = args[2][0]

                            new_dict = {}

                            # loop through the objects items, assigning the key to the key and
                            # value to the value
                            for key, value in self.vars[vname].value.items():

                                # log old key
                                old_key = key

                                # execute the function
                                self.vars[keyvarname] = Var(keyvarname, key)
                                self.vars[valuevarname] = Var(
                                    valuevarname, value)

                                # run the function
                                ret = self.interpret(function)

                                if self.vars[keyvarname].value == old_key:
                                    new_dict[old_key] = self.vars[valuevarname].value
                                else:
                                    new_dict[self.vars[keyvarname].value] = self.vars[valuevarname].value

                            self.vars[vname].value = new_dict

                            return self.vars[vname].value

                # the below conditions interpret a line based on initial appearances
                # beneath these conditions will the Interpreter then parse the arguments from the line as a method call
                # request for Interpreter redirect to a block of code
                # the first argument is
                if func == 'redirect':
                    line_vname = self.parse(0, line, f, sp, args)[2]
                    _block = args[1][0]
                    self.redirect_inside = []

                    # creates redirect for this interpreter
                    self.redirect = [line_vname, _block]
                    self.redirecting = True
                    return self.redirect

                # request for Interpreter redirect cancellation
                if func == 'stopredirect':
                    self.redirecting = False

                    return True

                # starts the redirection
                if func == 'startredirect':
                    ret = None
                    for _ins in self.redirect_inside:
                        ins = _ins[0]
                        _line = _ins[1]
                        self.vars[self.redirect[0]] = Var(
                            self.redirect[0], _line)
                        ret = self.interpret(ins)
                    return ret

                # alternative to '~' user defined method syntax
                # example call:
                # function('sum')
                elif func == 'function':

                    # obtain the name of the function
                    fname = self.parse(0, line, f, sp, args)[2]

                    # obtain line to be added to the method
                    block = args[1][0]

                    # function arguments

                    # return variable name
                    ret = '__unused'

                    # create the new Method
                    new_method = self.Method(fname, self)

                    # add the body
                    new_method.add_body(block)
                    new_method.add_return(f"{fname}__return__")
                    

                    # obtain the rest of the arguments as method args
                    for i in range(2, len(args)):
                        # adds variable name as an argument
                        # if any function specific argument is None, break
                        val = self.parse(i, line, f, sp, args)[2]
                        if val == None:
                            break
                        new_method.add_arg(val)
                    self.methods[fname] = new_method
                    return fname

                # performs modular arithmetic on the two arguments given
                elif func == 'mod':
                    return self.parse(0, line, f, sp, args)[2] % self.parse(1, line, f, sp, args)[2]

                # returns a value to a function
                # first argument is the function to return to
                # second argument is the value to return
                # ret() is used to create a return buffer for
                #   multiprogramming.
                elif func == 'ret':

                    # function to return to
                    fname = self.parse(0, line, f, sp, args)[2]

                    # value
                    value = self.parse(1, line, f, sp, args)[2]

                    vname = f"{fname}__return__"
                    self.vars[vname].value = value
                    return value

                # user method execution requested
                elif func in self.methods.keys():
                    method = self.methods[func]

                    # create func args
                    func_args = []

                    try:
                        for i in range(len(args)):
                            arguments = args[i]
                            line, as_s, arg = self.parse(i, line, f, sp, args)
                            func_args.append(arg)
                            meth_argname = method.args[i]
                            self.vars[meth_argname] = Var(meth_argname, arg)
                    except:
                        l = len(method.args)
                        if l != 0:
                            self.err("bad arg count",
                                     f"correct arg count is {l}", line)

                    # create return variable
                    ret_name = method.returns[0]

                    # add the return variable if not exists
                    if ret_name not in self.vars:
                        self.vars[ret_name] = Var(ret_name, None)
                        
                    # execute method
                    method.run(func_args, self, args)

                    # if its a variable
                    if ret_name in self.vars:
                        return self.vars[ret_name].value
                    
                    #                     # create return variable
                    # created_return_name = f"{func}__return__"
                    # if created_return_name not in self.vars:
                    #     self.vars[created_return_name].value = r
                                        
                    try:
                        return eval(str(self.vars[ret_name].value), {}, {})
                    except:
                        pass
                    try:
                        return str(self.vars[ret_name].value)
                    except:
                        return str(self.vars[ret_name])

                # creating a list
                if func == 'arr' or func == 'from':
                    arr = []
                    if args[0][0] == '':
                        return arr
                    for i in range(len(args)):
                        arr.append(self.parse(i, line, f, sp, args)[2])
                    return arr

                # creates a dictionary from its arguments
                # every two arguments
                # an odd quantity of arguments is impossible
                if func == 'dictfrom':
                    d = {}
                    if args[0][0] == '':
                        return d

                    # step over arguments in steps of 2
                    for i in range(0, len(args), 2):
                        cur = self.parse(i, line, f, sp, args)[2]
                        nxt = self.parse(i + 1, line, f, sp, args)[2]
                        d[cur] = nxt
                    return d

                # splits the first argument by the second argument
                if func == 'split':
                    to_split = self.parse(0, line, f, sp, args)[2]
                    splitting_by = self.parse(1, line, f, sp, args)[2]
                    return to_split.split(splitting_by)

                # obtains text between the first argument of the second argument
                if func == 'between':

                    # surrounding token
                    surrounding = self.parse(0, line, f, sp, args)[2]

                    # string to analyze
                    string = self.parse(1, line, f, sp, args)[2]

                    funccalls = []
                    try:
                        while string.count(surrounding) > 1:
                            string = string[string.index(
                                surrounding) + len(surrounding):]
                            funccalls.append(
                                string[:string.index(surrounding)])
                            string = string[string.index(
                                surrounding) + len(surrounding):]
                    except:
                        None

                    return funccalls

                # determines if the argument passed is of the type specified
                if func == 'isstr':
                    return isinstance(self.parse(0, line, f, sp, args)[2], str)
                elif func == 'islist':
                    return isinstance(self.parse(0, line, f, sp, args)[2], list)
                elif func == 'isfloat':
                    return isinstance(self.parse(0, line, f, sp, args)[2], float)
                elif func == 'isint':
                    return isinstance(self.parse(0, line, f, sp, args)[2], int)
                elif func == 'isdict':
                    return isinstance(self.parse(0, line, f, sp, args)[2], dict)

                # gets the sum of all arguments
                if func == 'sum':
                    total = 0
                    for i in range(len(args)):
                        total += sum(self.parse(i, line, f, sp, args)[2])
                    return total

                # creates / sets a variable
                if func == 'var':

                    # extract varname
                    varname = self.parse(0, line, f, sp, args)[2]

                    # extract value
                    value = self.parse(1, line, f, sp, args)[2]

                    # add / set variable
                    self.vars[varname] = Var(varname, value)
                    return value

                # converts the argument to a list
                elif func == 'list':
                    return list(self.parse(0, line, f, sp, args)[2])

                # determines if a variable exists or not
                elif func == 'exists':
                    return self.parse(0, line, f, sp, args)[2] in self.vars

                # gets the length of the first argument
                elif func == 'len':
                    line, as_s, arg = self.parse(0, line, f, sp, args)
                    return len(arg)

                # asserts each argument is True, prints and logs assertion error
                elif func == 'assert':
                    for i in range(len(args)):
                        arguments = args[i]
                        line, assertion = self.convert_arg(
                            arguments[0], line, f, sp, args)
                        if not assertion:
                            failed = ''
                            for arg in args:
                                failed += str(arg[0]) + ' '
                            err = self.err("assertion error", "", failed)
                            self.logg(err, line)
                            return False

                    return True

                # trace capabilities
                elif obj == 'trace':
                    if objfunc == 'before':
                        return lines_ran[len(lines_ran) - self.parse(0, line, f, sp, args)[2]:]
                    if objfunc == 'this':
                        return lines_ran[-1]
                    if objfunc == 'len':
                        return total_ints
                    return '<msnint2 class>'

                # casting
                elif func == 'int':
                    return int(self.parse(0, line, f, sp, args)[2])
                elif func == 'float':
                    return float(self.parse(0, line, f, sp, args)[2])
                elif func == 'str':
                    return str(self.parse(0, line, f, sp, args)[2])
                elif func == 'bool':
                    return bool(self.parse(0, line, f, sp, args)[2])
                elif func == 'complex':
                    return complex(self.parse(0, line, f, sp, args)[2])

                # gets the type of the argument
                elif func == 'type':
                    return type(self.parse(0, line, f, sp, args)[2])

                # gets the dir of the argument
                elif func == 'dir':
                    return dir(self.parse(0, line, f, sp, args)[2])

                # casting to iterables / sets / dicts
                elif func == 'set':

                    # creates a set from all arguments
                    if args[0][0] == '':
                        return set()

                    # creates a set from all arguments
                    s = set()

                    # adds all arguments to the set
                    for i in range(len(args)):
                        s.add(self.parse(i, line, f, sp, args)[2])
                    return s

                elif func == 'dict':
                    return dict(self.parse(0, line, f, sp, args)[2])
                elif func == 'tuple':
                    return tuple(self.parse(0, line, f, sp, args)[2])

                # conditional logic
                elif func == 'if':
                    # if condition and blocks arguments
                    ifcond_s = args[0][0]
                    true_block_s = args[1][0]

                    # false block is optional
                    try:
                        false_block_s = args[2][0]
                    except:
                        false_block_s = None

                    ifcond = self.parse(0, line, f, sp, args)[2]

                    # if condition is true
                    if ifcond:
                        return self.parse(1, line, f, sp, args)[2]

                    # otherwise false block is executed
                    if false_block_s:
                        return self.parse(2, line, f, sp, args)[2]
                    return False

                # while logic WIPWIPWIPWIPWIP
                elif func == 'while':

                    # while arguments as strings
                    whilecond_s = args[0][0]
                    while_block_s = args[1][0]

                    while (self.interpret(whilecond_s)):
                        self.interpret(while_block_s)
                    return True

                 # iteration
                elif func == 'for':

                    # block to execute
                    inside = args[3][0]

                    # times to loop
                    line, as_s, start = self.parse(0, line, f, sp, args)
                    line, as_s, end = self.parse(1, line, f, sp, args)
                    line, as_s, loopvar = self.parse(2, line, f, sp, args)
                    self.vars[loopvar] = Var(loopvar, start)
                    # regular iteration
                    if start < end:
                        for i in range(start, end):
                            if loopvar in self.vars and self.vars[loopvar].value >= end:
                                break
                            self.vars[loopvar] = Var(loopvar, i)
                            self.interpret(inside)

                    # reversed if requested
                    elif start > end:
                        for i in reversed(range(end, start)):
                            if loopvar in self.vars and self.vars[loopvar].value < end:
                                break
                            self.vars[loopvar] = Var(loopvar, i)
                            self.interpret(inside)
                    return self.vars[loopvar].value

                # executes a block of code for each element in an array
                elif func == 'each':

                    # get array argument
                    line, as_s, array = self.parse(0, line, f, sp, args)

                    # get element variable name
                    line, as_s, element_name = self.parse(1, line, f, sp, args)
                    block_s = args[2][0]

                    # prepare each element
                    self.vars[element_name] = Var(element_name, 0)

                    # execute block for each element
                    for i in range(len(array)):
                        self.vars[element_name].value = array[i]
                        self.interpret(block_s)
                    return array
                
                # sorting an array by an attribute of each element
                elif func == 'sortby':
                    
                    # iterable to sort
                    iterable = self.parse(0, line, f, sp, args)[2]
                    
                    # variable name
                    varname = self.parse(1, line, f, sp, args)[2]
                    
                    # block of code to interpret, the sorting
                    # is based on the interpretation of this block
                    block = args[2][0]
                    
                    # pairing elements to their interpretations
                    pairing = []
                    for i in range(len(iterable)):
                        self.vars[varname] = Var(varname, iterable[i])
                        pairing.append((self.interpret(block), iterable[i]))
                    # sort the pairing based on the first element of each pair
                    pairing.sort(key=lambda x: x[0])
                    # return the sorted array containing the second element
                    # of each pair
                    return [pair[1] for pair in pairing]
                                
                # performs list comprehension
                elif func == 'comp':
                    lst = []
                    
                    # array to comprehend
                    arr = self.parse(0, line, f, sp, args)[2]
                    
                    # varname for the element
                    varname = self.parse(1, line, f, sp, args)[2]
                    
                    # block to perform
                    block = args[2][0]
                    
                    # performs the list comprehension
                    for v in arr:
                        self.vars[varname] = Var(varname, v)
                        r = self.interpret(block)
                        if r != msn2_none:
                            lst.append(r)
                    return lst
                
                # returns the first argument, then performs the second argument as a block
                elif func == 'do':
                    ret = self.parse(0, line, f, sp, args)[2]
                    self.interpret(args[1][0])
                    return ret
                
                # special value for msn2 to return None
                elif func == 'None':
                    return msn2_none
                
                # filters an iterable to retain all elements that satisfy the second argument as
                # a block
                # the first argument is a string for a variable for each element
                # second argument is a block to run for each element
                elif func == 'filter':
                    
                    # iterable to filter
                    iterable = self.parse(0, line, f, sp, args)[2]
                    
                    # variable name
                    varname = self.parse(1, line, f, sp, args)[2]

                    # block to execute
                    block = args[2][0]
                    
                    # new array
                    filtered = []
                    
                    # iterate through each element
                    for v in iterable:
                            
                        # set the variable to the element
                        self.vars[varname] = Var(varname, v)
                        
                        # if the block returns true, add the element to the new array
                        if self.interpret(block):
                            filtered.append(v)

                    return filtered

                # unpacks the first argument into any amount of variables
                # specified by the remaining arguments provided as variable names
                # creates the variables if they don't exists
                # 5/22/2023
                elif func == 'unpack':

                    # iterable to unpack
                    iterable = self.parse(0, line, f, sp, args)[2]

                    # variable names to unpack into
                    for i in range(1, len(args)):
                        varname = self.parse(i, line, f, sp, args)[2]
                        self.vars[varname] = Var(varname, iterable[i - 1])

                    return iterable
                
                # determines if the array has all elements specified
                # takes any amount of arguments
                # first argument is the iterable
                # the rest are elements to check for
                elif func == 'has':
                    # iterable to check
                    iterable = self.parse(0, line, f, sp, args)[2]
                    
                    # iterate through each element
                    for i in range(len(args) - 1):
                        if self.parse(i + 1, line, f, sp, args)[2] not in iterable:
                            return False
                    return True
                
                # gets the first element in the iterable
                elif func == 'first' or func == 'head':
                    try:
                        return self.parse(0, line, f, sp, args)[2][0]
                    except:
                        return None
                
                # gets the last element in the iterable
                elif func == 'last' or func == 'tail':
                    try:
                        return self.parse(0, line, f, sp, args)[2][-1]
                    except:
                        return None

                # the following provide efficient variable arithmetic
                elif func == 'add':
                    line, as_s, first = self.parse(0, line, f, sp, args)
                    line, as_s, second = self.parse(1, line, f, sp, args)

                    # case array
                    if isinstance(self.vars[first].value, list):
                        self.vars[first].value.append(second)

                    # case string or number
                    else:
                        self.vars[first].value += second
                    return self.vars[first].value

                # performs basic operations on non-variable values
                elif obj == 'op':

                    # obtains the first argument
                    arg1 = self.parse(0, line, f, sp, args)[2]

                    # adds all arguments
                    if objfunc == 'append' or objfunc == 'push' or objfunc == 'add' or objfunc == 'plus' or objfunc == '+' or objfunc == 'concat' or objfunc == 'concatenate' or objfunc == 'join' or objfunc == 'merge' or objfunc == 'sum':

                        if isinstance(arg1, list):
                            for i in range(1, len(args)):
                                arg2 = self.parse(i, line, f, sp, args)[2]
                                arg1.append(arg2)
                            return arg1
                        else:
                            for i in range(1, len(args)):
                                arg2 = self.parse(i, line, f, sp, args)[2]
                                arg1 += arg2
                            return arg1

                    # subtracts all arguments
                    if objfunc == 'sub' or objfunc == 'minus' or objfunc == 'subtract' or objfunc == '-':
                        for i in range(1, len(args)):
                            arg2 = self.parse(i, line, f, sp, args)[2]
                            arg1 -= arg2
                        return arg1

                    if objfunc == 'mul' or objfunc == 'times' or objfunc == 'x' or objfunc == '*' or objfunc == 'multiply':
                        for i in range(1, len(args)):
                            arg2 = self.parse(i, line, f, sp, args)[2]
                            arg1 *= arg2
                        return arg1

                    if objfunc == 'div' or objfunc == 'divide' or objfunc == 'over' or objfunc == '/' or objfunc == '÷':
                        for i in range(1, len(args)):
                            arg2 = self.parse(i, line, f, sp, args)[2]
                            arg1 /= arg2
                        return arg1
                    
                    # integer division
                    if objfunc == 'idiv' or objfunc == 'intdiv' or objfunc == 'intdivide' or objfunc == 'intover' or objfunc == '//' or objfunc == '÷÷':
                        for i in range(1, len(args)):
                            arg2 = self.parse(i, line, f, sp, args)[2]
                            arg1 //= arg2
                        return arg1

                    if objfunc == 'mod' or objfunc == 'modulo' or objfunc == 'modulus' or objfunc == '%' or objfunc == 'remainder':
                        for i in range(1, len(args)):
                            arg2 = self.parse(i, line, f, sp, args)[2]
                            arg1 %= arg2
                        return arg1

                    if objfunc == 'pow' or objfunc == 'power' or objfunc == 'exponent' or objfunc == '**':
                        for i in range(1, len(args)):
                            arg2 = self.parse(i, line, f, sp, args)[2]
                            arg1 **= arg2
                        return arg1
                    
                    if objfunc == 'root' or objfunc == 'nthroot' or objfunc == 'nthrt' or objfunc == '√':
                        for i in range(1, len(args)):
                            arg2 = self.parse(i, line, f, sp, args)[2]
                            arg1 **= (1 / arg2)
                        return arg1
                    return '<msnint2 class>'
                
                # computes the maximum value from all arguments
                # takes any amount of arguments, all being
                # either numbers or lists
                elif func == 'maximum':
                    maxval = max(_f) if isinstance((_f := self.parse(0, line, f, sp, args)[2]), list) else _f
                    for i in range(1, len(args)):
                        val = self.parse(i, line, f, sp, args)[2]
                        # is a list argument
                        if isinstance(val, list):
                            maxval = max(maxval, max(val))
                        # is a number
                        else:
                            maxval = max(maxval, val)
                    return maxval
                elif func == 'minimum':
                    minval = min(_f) if isinstance((_f := self.parse(0, line, f, sp, args)[2]), list) else _f
                    for i in range(1, len(args)):
                        val = self.parse(i, line, f, sp, args)[2]
                        # is a list argument
                        if isinstance(val, list):
                            minval = min(minval, min(val))
                        # is a number
                        else:
                            minval = min(minval, val)
                    return minval

                # more support for functions
                elif obj == 'function':

                    fname = self.parse(0, line, f, sp, args)[2]

                    # adds a line of code to a function / method's body
                    if objfunc == 'addbody':
                        _body = self.parse(1, line, f, sp, args)[2]

                        self.methods[fname].add_body(_body)
                        return fname

                    # adds an argument to a function
                    if objfunc == 'addarg':

                        arg = self.parse(1, line, f, sp, args)[2]

                        self.methods[fname].add_arg(arg)
                        return fname

                    # adds a return variable to this function
                    if objfunc == 'addreturn':
                        retvar = self.parse(1, line, f, sp, args)[2]

                        self.methods[fname].add_return(retvar)
                        return fname

                    if objfunc == 'getbody':
                        return self.methods[fname].body

                    if objfunc == 'getargs':
                        return self.methods[fname].args

                    if objfunc == 'getreturn':
                        return self.methods[fname].returns[0]

                    # removes a function from the working context
                    if objfunc == 'destroy':
                        self.methods.pop(fname)
                        return fname

                    # simulates an execution of a function
                    if objfunc == 'run':

                        # form a string that is msn2 of the user defined function

                        args_str = ''
                        for i in range(1, len(args)):
                            arg = self.parse(i, line, f, sp, args)[2]
                            if i != len(args) - 1:
                                args_str += str(arg) + ', '
                            else:
                                args_str += str(arg)

                        inst = fname + '(' + args_str + ')'
                        return self.interpret(inst)

                    return '<msnint2 class>'

                elif func == 'sub':
                    line, as_s, first = self.parse(0, line, f, sp, args)
                    line, as_s, second = self.parse(1, line, f, sp, args)
                    self.vars[first].value -= second
                    return self.vars[first].value
                elif func == 'mul':
                    line, as_s, first = self.parse(0, line, f, sp, args)
                    line, as_s, second = self.parse(1, line, f, sp, args)
                    self.vars[first].value *= second
                    return self.vars[first].value
                elif func == 'div':
                    line, as_s, first = self.parse(0, line, f, sp, args)
                    line, as_s, second = self.parse(1, line, f, sp, args)
                    self.vars[first].value /= second
                    return self.vars[first].value

                # appends to an array variable
                elif func == 'append':

                    # varname
                    varname = self.parse(0, line, f, sp, args)[2]

                    # value to append
                    value = self.parse(1, line, f, sp, args)[2]

                    self.vars[varname].value.append(value)
                    return value

                # gets at the index specified
                elif func == '->':

                    # array
                    array = self.parse(0, line, f, sp, args)[2]

                    # index
                    index = self.parse(1, line, f, sp, args)[2]

                    return array[index]

                # gets the MSNScript version of this interpreter
                elif func == 'version':
                    return self.version

                # destroys a function or variable
                elif func == 'destroy':
                    varname = self.parse(0, line, f, sp, args)[2]

                    # deletes all variables or methods that start with '__'
                    if varname == '__':
                        for key in list(self.vars):
                            if key.startswith('__'):
                                del self.vars[key]
                        for key in list(self.methods):
                            if key.startswith('__'):
                                del self.methods[key]
                        return True

                    if varname in self.vars:
                        del self.vars[varname]

                    elif varname in self.methods:
                        del self.methods[varname]

                    return True

                # random capabilities
                elif func == 'random':
                    # gets a random number between 0 and 1
                    if len(args) == 1:

                        arg = self.parse(0, line, f, sp, args)[2]

                        return random.random()

                    # random number in range
                    elif len(args) == 2:
                        arg = self.parse(0, line, f, sp, args)[2]
                        arg2 = self.parse(1, line, f, sp, args)[2]

                        return (random.random() * (arg2 - arg)) + arg

                    # random int in range
                    elif len(args) == 3:
                        arg = self.parse(0, line, f, sp, args)[2]
                        arg2 = self.parse(1, line, f, sp, args)[2]

                        return math.floor((random.random() * (arg2 - arg)) + arg)

                    return '<msnint2 class>'

                # html parsing simplified
                elif obj == 'html':

                    url = self.parse(0, line, f, sp, args)[2]

                    # creates a BeautifulSoup object of a url
                    if objfunc == 'soup':

                        response = requests.get(url)
                        return BeautifulSoup(response.content, 'html5lib')

                    # scrapes all html elements from a url
                    if objfunc == 'from':

                        obj_to_add = []
                        all_elem = self.html_all_elements(url)
                        for elem in all_elem:
                            obj_to_add.append(
                                {'tag': elem.name, 'attrs': elem.attrs, 'text': elem.text})
                        return obj_to_add

                # ai specific usage
                elif obj == 'ai':

                    # verify existence of openai api key
                    if not openai.api_key:
                        try:
                            openai.api_key = os.environ['OPENAI_API_KEY']
                        except:
                            raise Exception(
                                'OpenAI API key not found. Please set your OPENAI_API_KEY environment variable to your OpenAI API key.')

                    # asks openai model a question
                    # simple ai, see top of file for definition
                    if objfunc == 'basic':

                        # generates an ai response with the basic model
                        return ai_response(models['basic']['model'], self.parse(0, line, f, sp, args)[2], models['basic']['creativity'])
                    return '<msnint2 class>'

                # defines new syntax, see tests/validator.msn2 for documentation
                elif func == 'syntax':

                    # gets the syntax token
                    token = self.parse(0, line, f, sp, args)[2]

                    # gets the variable name of the between
                    between = self.parse(1, line, f, sp, args)[2]

                    # function that should be executed when the syntax is found
                    function = args[2][0]

                    return self.add_syntax(token, between, function)

                # creates a new enclosed syntax that should execute the block
                # specified on the line by which it was created
                elif func == 'enclosedsyntax':

                    start = self.parse(0, line, f, sp, args)[2]
                    end = self.parse(1, line, f, sp, args)[2]
                    varname = self.parse(2, line, f, sp, args)[2]

                    index = str(start) + 'msnint2_reserved' + str(end)
                    enclosed[index] = [start, end, varname, args[3][0]]

                    if len(args) == 5:
                        enclosed[index].append(
                            self.parse(4, line, f, sp, args)[2])
                    return enclosed[index]

                # defines a new macro
                elif func == 'macro':

                    token = self.parse(0, line, f, sp, args)[2]

                    varname = self.parse(1, line, f, sp, args)[2]

                    code = args[2][0]

                    macros[token] = [token, varname, code]

                    # 4th argument offered as a return value from that macro
                    # as opposed to a block of code
                    if len(args) == 4:
                        macros[token].append(
                            self.parse(3, line, f, sp, args)[2])

                    return macros[token]

                # creates a macro that should be declared at the end of a line
                elif func == 'postmacro':

                    token = self.parse(0, line, f, sp, args)[2]

                    varname = self.parse(1, line, f, sp, args)[2]

                    code = args[2][0]

                    postmacros[token] = [token, varname, code]

                    # same as macro
                    if len(args) == 4:
                        postmacros[token].append(
                            self.parse(3, line, f, sp, args)[2])

                    return postmacros[token]

                # creates an inline syntax that allows for value replacement
                # within a line
                # this is different from the above system calls because this
                # syntax is invoked and returned as a value which will replace
                # the invocation within the line
                # arguments are the same as enclosedsyntax
                #
                # WIPWIPWIPWIPWIP
                elif func == 'inlinesyntax':
                    start = self.parse(0, line, f, sp, args)[2]
                    end = self.parse(1, line, f, sp, args)[2]
                    varname = self.parse(2, line, f, sp, args)[2]

                    index = str(start) + 'msnint2_reserved' + str(end)

                    inlines[index] = [start, end, varname, args[3][0]]

                    if len(args) == 5:
                        inlines[index].append(
                            self.parse(4, line, f, sp, args)[2])
                    return inlines[index]

                # performs object based operations
                elif obj == 'var':

                    # determines if all variables passed are equal
                    if objfunc == 'equals':
                        firstvar = self.vars[self.parse(
                            0, line, f, sp, args)[2]].value
                        for i in range(1, len(args)):
                            if firstvar != self.vars[self.parse(i, line, f, sp, args)[2]].value:
                                return False
                        return True
                    return '<msnint2 class>'

                # gets the value of a variable
                elif func == 'val':

                    # gets the variable name
                    varname = self.parse(0, line, f, sp, args)[2]

                    try:
                        return self.vars[varname].value
                    except:
                        return self.vars[varname]

                # gets a sorted version of the array
                elif func == 'sorted':
                    return sorted(self.parse(0, line, f, sp, args)[2])

                # performs file-specific operations
                elif obj == 'file':

                    # creates a file
                    if objfunc == 'create':
                        lock.acquire()
                        line, as_s, filename = self.parse(0, line, f, sp, args)
                        open(filename, 'w').close()
                        lock.release()
                        return True

                    # reads text from a file
                    if objfunc == 'read':
                        lock.acquire()
                        file = open(self.parse(0, line, f, sp, args)[2], "r")
                        contents = file.read()
                        file.close()
                        lock.release()
                        return contents

                    # writes to a file
                    if objfunc == 'write':
                        lock.acquire()
                        file = open(self.parse(0, line, f, sp, args)[2], "w")
                        towrite = self.parse(1, line, f, sp, args)[2]
                        file.write(towrite)
                        file.close()
                        lock.release()
                        return towrite

                    # writes the argument as code
                    if objfunc == 'writemsn':
                        lock.acquire()
                        file = open(self.parse(0, line, f, sp, args)[2], "w")
                        towrite = args[1][0]
                        file.write(towrite)
                        lock.release()
                        return towrite

                    # clears a file of all text
                    if objfunc == 'clear':
                        lock.acquire()
                        file = open(self.parse(0, line, f, sp, args)[2], "w")
                        file.write("")
                        file.close()
                        lock.release()
                        return True

                    # appends to a file
                    if objfunc == 'append':
                        lock.acquire()
                        file = open(self.parse(0, line, f, sp, args)[2], "a")
                        towrite = self.parse(1, line, f, sp, args)[2]
                        file.write(towrite)
                        file.close()
                        lock.release()
                        return towrite

                    # deletes a file
                    if objfunc == 'delete':
                        lock.acquire()
                        deleting = self.parse(0, line, f, sp, args)[2]
                        try:
                            os.remove(deleting)
                        except:
                            None
                        lock.release()
                        return deleting

                    # renames a file
                    if objfunc == 'rename':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        new = self.parse(1, line, f, sp, args)[2]
                        os.rename(old, new)
                        lock.release()
                        return new

                    # copies a file
                    if objfunc == 'copy':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        new = self.parse(1, line, f, sp, args)[2]
                        shutil.copy2(old, new)
                        lock.release()
                        return new

                    if objfunc == 'copy2':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        new = self.parse(1, line, f, sp, args)[2]
                        shutil.copy2(old, new)
                        lock.release()
                        return new

                    if objfunc == 'copyfile':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        new = self.parse(1, line, f, sp, args)[2]
                        shutil.copyfile(old, new)
                        lock.release()
                        return new

                    if objfunc == 'fullpath':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        fullpath = os.path.abspath(path)
                        lock.release()
                        return fullpath

                    # moves a file
                    if objfunc == 'move':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        new = self.parse(1, line, f, sp, args)[2]
                        shutil.move(old, new)
                        lock.release()
                        return new

                    # determines if a file exists
                    if objfunc == 'exists':
                        lock.acquire()
                        exists = os.path.exists(
                            self.parse(0, line, f, sp, args)[2])
                        lock.release()
                        return exists

                    # determines if a file is a directory
                    if objfunc == 'isdir':
                        lock.acquire()
                        isdir = os.path.isdir(
                            self.parse(0, line, f, sp, args)[2])
                        lock.release()
                        return isdir

                    # determines if a file is a file
                    if objfunc == 'isfile':
                        lock.acquire()
                        isfile = os.path.isfile(
                            self.parse(0, line, f, sp, args)[2])
                        lock.release()
                        return isfile

                    # lists files in a directory
                    if objfunc == 'listdir':
                        lock.acquire()
                        try:
                            listdir = os.listdir(
                                self.parse(0, line, f, sp, args)[2])
                            lock.release()
                            return listdir
                        except FileNotFoundError:

                            # directory doesn't exist
                            lock.release()
                            return None

                    # makes a directory
                    if objfunc == 'mkdir':
                        lock.acquire()
                        try:
                            made = os.mkdir(self.parse(
                                0, line, f, sp, args)[2])
                            lock.release()
                            return made
                        except FileExistsError:
                            lock.release()
                            return False

                    # removes a directory
                    if objfunc == 'rmdir':
                        lock.acquire()
                        try:
                            rm = os.rmdir(self.parse(0, line, f, sp, args)[2])
                            lock.release()
                            return rm
                        except OSError:
                            lock.release()
                            return None

                    # gets the current working directory
                    if objfunc == 'getcwd':
                        lock.acquire()
                        cwd = os.getcwd()
                        lock.release()
                        return cwd

                    # gets the size of a file
                    if objfunc == 'getsize':
                        lock.acquire()
                        size = os.path.getsize(
                            self.parse(0, line, f, sp, args)[2])
                        lock.release()
                        return size

                    # deletes all files and directories within a directory
                    if objfunc == 'emptydir':
                        lock.acquire()
                        directory = self.parse(0, line, f, sp, args)[2]
                        try:
                            for file in os.listdir(directory):
                                try:
                                    os.remove(os.path.join(directory, file))
                                except:
                                    shutil.rmtree(os.path.join(
                                        directory, file), ignore_errors=True)
                            lock.release()
                            return directory
                        except FileNotFoundError:

                            # directory doesn't exist
                            lock.release()
                            return None

                elif func == 'fileacquire':
                    lock.acquire()
                    return True

                elif func == 'filerelease':
                    lock.release()
                    return True

                # automation operations
                elif obj == 'auto':
                    
                    # gets the largest element from a list of elements
                    if objfunc == 'largest':
                        elements = self.parse(0, line, f, sp, args)[2]
                        if not elements:
                            return elements
                        largest = elements[0]
                        for element in elements:
                            try:
                                # element has width and height
                                if element.width() > largest.width() and element.height() > largest.height():
                                    largest = element
                            except:
                                # element does not have width and height
                                return element
                        return largest
                    return '<msnint2 class>'
                
                # # performs math operations
                elif obj == 'math':
                    if objfunc == 'add':
                        return self.parse(0, line, f, sp, args)[2] + self.parse(1, line, f, sp, args)[2]

                    if objfunc == 'subtract':
                        return self.parse(0, line, f, sp, args)[2] - self.parse(1, line, f, sp, args)[2]

                    if objfunc == 'multiply':
                        return self.parse(0, line, f, sp, args)[2] * self.parse(1, line, f, sp, args)[2]

                    if objfunc == 'divide':
                        return self.parse(0, line, f, sp, args)[2] / self.parse(1, line, f, sp, args)[2]

                    if objfunc == 'power':
                        return self.parse(0, line, f, sp, args)[2] ** self.parse(1, line, f, sp, args)[2]

                    if objfunc == 'root':
                        return self.parse(0, line, f, sp, args)[2] ** (1 / self.parse(1, line, f, sp, args)[2])

                    if objfunc == 'sqrt':
                        return self.parse(0, line, f, sp, args)[2] ** 0.5

                    if objfunc == 'mod':
                        return self.parse(0, line, f, sp, args)[2] % self.parse(1, line, f, sp, args)[2]

                    if objfunc == 'floor':
                        return math.floor(self.parse(0, line, f, sp, args)[2])

                    if objfunc == 'ceil':
                        return math.ceil(self.parse(0, line, f, sp, args)[2])

                    if objfunc == 'round':
                        return round(self.parse(0, line, f, sp, args)[2])

                    if objfunc == 'abs':
                        return abs(self.parse(0, line, f, sp, args)[2])

                    if objfunc == 'sin':
                        return math.sin(self.parse(0, line, f, sp, args)[2])

                    if objfunc == 'cos':
                        return math.cos(self.parse(0, line, f, sp, args)[2])

                    if objfunc == 'tan':
                        return math.tan(self.parse(0, line, f, sp, args)[2])

                    if objfunc == 'asin':
                        return math.asin(self.parse(0, line, f, sp, args)[2])
                    return '<msnint2 class>'

                # inserts a value into the iterable at the specified index
                elif func == 'map':
                    
                    # iterable
                    iterable = self.parse(0, line, f, sp, args)[2]
                    
                    # varname
                    varname = self.parse(1, line, f, sp, args)[2]
                    
                    # function
                    function = args[2][0]
                    
                    # map the function to each element in the iterable
                    for i, el in enumerate(iterable):
                        self.vars[varname] = Var(varname, el)
                        iterable[i] = self.interpret(function)
                    return iterable
                        
                        
                        
                    
                # inserts a value into the iterable at the specified index
                elif func == 'insert':
                    return self.parse(0, line, f, sp, args)[2].insert(self.parse(1, line, f, sp, args)[2], self.parse(2, line, f, sp, args)[2])

                # gets the type of the first argument passed
                elif func == 'type':
                    return type(self.parse(0, line, f, sp, args)[2])

                # gets the parent context
                elif func == 'parent':
                    return self.parent

                # gets the booting context
                elif func == 'boot':
                    while self.parent != None:
                        self = self.parent
                    return self

                # sets an index of an array
                # first argument is the variable to modify
                # second is the index to modify
                # third argument is the value to set
                elif func == 'set':

                    # obtain varname of array
                    varname = self.parse(0, line, f, sp, args)[2]

                    # index to set at
                    ind = self.parse(1, line, f, sp, args)[2]

                    # value to set
                    val = self.parse(2, line, f, sp, args)[2]

                    # perform set operation
                    self.vars[varname].value[ind] = val

                    return val

                # deletes a variable
                elif func == 'del':
                    for i in range(len(args)):
                        line, as_s, first = self.parse(i, line, f, sp, args)
                        del self.vars[first]
                    return True

                # concatinates two strings
                elif func == 'cat':

                    # first argument (required)
                    first = self.parse(0, line, f, sp, args)[2]

                    # second argument (required)
                    second = self.parse(1, line, f, sp, args)[2]

                    cat = str(first) + str(second)

                    # concatinate rest of arguments
                    for i in range(2, len(args)):
                        cat += str(self.parse(i, line, f, sp, args)[2])

                    return cat

                # determines equality of all arguments
                elif func == 'equals':
                    line, as_s, arg1 = self.parse(0, line, f, sp, args)
                    for i in range(1, len(args)):
                        line, as_s, curr_arg = self.parse(i, line, f, sp, args)
                        if curr_arg != arg1:
                            return False
                    return True

                # nots a bool
                elif func == 'not':
                    return not self.parse(0, line, f, sp, args)[2]

                # ands two bools
                elif func == 'and':
                    first = self.parse(0, line, f, sp, args)[2]
                    if not first:
                        return False
                    for i in range(1, len(args)):
                        line, as_s, curr_arg = self.parse(i, line, f, sp, args)
                        if not curr_arg:
                            return False
                    return True

                # ors two bools
                elif func == 'or':
                    return self.parse(0, line, f, sp, args)[2] or self.parse(1, line, f, sp, args)[2]

                # comparing numbers
                elif func == 'greater' or func == 'g':
                    return self.parse(0, line, f, sp, args)[2] > self.parse(1, line, f, sp, args)[2]
                elif func == 'less' or func == 'l':
                    return self.parse(0, line, f, sp, args)[2] < self.parse(1, line, f, sp, args)[2]
                elif func == 'greaterequal' or func == 'ge':
                    return self.parse(0, line, f, sp, args)[2] >= self.parse(1, line, f, sp, args)[2]
                elif func == 'lessequal' or func == 'le':
                    return self.parse(0, line, f, sp, args)[2] <= self.parse(1, line, f, sp, args)[2]

                # data structure for holding multiple items
                elif func == 'class':
                    # new interpreter
                    inter = Interpreter()

                    # log self
                    inter.parent = self

                    # extract class name
                    name = self.parse(0, line, f, sp, args)[2]

                    # block at which the class exists
                    block_s = args[1][0]

                    # execute the block in the private environment
                    inter.execute(block_s)

                    # creates a variable out of the new interpreters resources
                    obj_to_add = {}
                    for varname in inter.vars:
                        val = inter.vars[varname].value
                        obj_to_add[varname] = Var(varname, val)

                    for methodname in inter.methods:
                        obj_to_add[methodname] = Var(
                            methodname + "#method", inter.methods[methodname])

                    self.vars[name] = Var(name, obj_to_add)
                    return obj_to_add

                # gets the first argument at the second argument
                elif func == 'get':
                    iterable = self.parse(0, line, f, sp, args)[2]
                    index = self.parse(1, line, f, sp, args)[2]
                    return iterable[index]

                # get the keys of the first argument
                elif func == 'keys':
                    return self.parse(0, line, f, sp, args)[2].keys()

                # imports resources from another location
                elif func == 'import' or func == 'launch' or func == 'include' or func == 'using':

                    # for each import
                    for i in range(len(args)):
                        line, as_s, path = self.parse(i, line, f, sp, args)
                        if not path.endswith('.msn2'):
                            path += '.msn2'
                        if path in self.imports:
                            continue
                        self.imports.add(path)
                        contents = ''
                        with open(path) as f:
                            contents = f.readlines()
                            script = ''
                            for line in contents:
                                script += line
                            self.logg("importing library", str(args[0][0]))
                            self.execute(script)
                    return
                
                # imports values from an enclosing Python script
                elif func == 'in':
                    reserved_name = '_msn2_reserved_in__'
                    if reserved_name not in self.vars:
                        return None
                    inval = self.vars[reserved_name].value
                    # if no arguments, return the value
                    if args[0][0] == '':
                        return inval
                    # if 1 argument, get index of value from input
                    elif len(args) == 1:
                        return inval[self.parse(0, line, f, sp, args)[2]]
                    # if 2 arguments, get slice of input
                    elif len(args) == 2:
                        start = self.parse(0, line, f, sp, args)[2]
                        end = self.parse(1, line, f, sp, args)[2]
                        return inval[start:end]
                    return inval
                    
                # exports values to an enclosing Python script
                elif func == 'out':
                        
                    # variables to output
                    outting = []
                
                    # for each argument
                    for i in range(len(args)):
                        outting.append(self.parse(i, line, f, sp, args)[2])        
                        
                    _out = '_msn2_reserved_out__'
                    # create a variable for the enclosing Python script
                    # to access
                    self.vars[_out] = Var(_out, outting)
                    
                    # return outting
                    return outting

                # interpreter printing mechanism
                elif func == 'prnt':
                    for i in range(len(args)):
                        arguments = args[i]
                        first = self.interpret(arguments[0])
                        srep = str(first)
                        line = line[:f + sp + arguments[1] + 1] + \
                            srep + line[f + sp + arguments[2] + 1:]
                        if i != len(args) - 1:
                            self.out += srep + ' '
                        else:
                            self.out += srep + '\n'
                    return first

                # python print
                elif func == 'print':
                    ret = None
                    for i in range(len(args)):
                        ret = self.parse(i, line, f, sp, args)[2]
                        if i != len(args) - 1:
                            print(ret, end=" ", flush=True)
                        else:
                            print(ret, flush=True)
                    return ret

                # sleeps the thread for the first argument amount of seconds
                elif func == "sleep":
                    return time.sleep(self.parse(0, line, f, sp, args)[2])

                # returns this interpreter
                elif func == 'me':
                    return self.me()

                # provides a representation of the current environment
                elif func == 'env':
                    should_print_s = args[0][0]
                    line, should_print = self.convert_arg(
                        should_print_s, line, f, sp, args)
                    strenv = ''
                    strenv += "--------- environment ---------"
                    strenv += "\nout:\n" + self.out
                    strenv += "\nvariables:\n"
                    for varname, v in self.vars.items():
                        try:
                            strenv += "\t" + varname + \
                                " = " + str(v.value) + '\n'
                        except:
                            None

                    strenv += "\nmethods:\n"
                    # printing methods
                    for methodname, Method in self.methods.items():
                        strenv += "\t" + methodname + "("
                        for i in range(len(Method.args)):
                            arg = Method.args[i]
                            if i != len(Method.args) - 1:
                                strenv += "" + str(arg) + ", "
                            else:
                                strenv += "" + str(arg)
                        # add body line count
                        strenv += ") : " + str(len(Method.body)) + " inst\n"

                    # printing macros
                    strenv += "\nmacros:\n\t"

                    # adding regular macros
                    if len(macros) > 0:
                        strenv += "premacros:\n\t\t"
                        for macro in macros:
                            strenv += macro + "\n\t\t"

                    if len(postmacros) > 0:
                        strenv += "\n\tpostmacros:\n\t\t"
                        for macro in postmacros:
                            strenv += macro + "\n\t\t"

                    if len(syntax) > 0:
                        strenv += "\n\tsyntax:\n\t\t"
                        for macro in syntax:
                            strenv += macro + "\n\t\t"

                    if len(enclosed) > 0:
                        strenv += "\n\tenclosedsyntax:\n\t\t"
                        for macro in enclosed:
                            strenv += macro + "\n\t\t"

                    strenv += "\nlog:\n" + self.log
                    strenv += "\n-------------------------------"
                    if should_print:
                        print(strenv)
                    return strenv


                # arithmetic, equivalent to the op class

                # executes MSNScript2 from its string representation
                
                elif func == '-':
                    if len(args) == 1:
                        return self.interpret(self.parse(0, line, f, sp, args)[2])
                    
                    # subtracts all arguments from the first argument
                    else:
                        ret = self.parse(0, line, f, sp, args)[2]
                        for i in range(1, len(args)):
                            ret -= self.parse(i, line, f, sp, args)[2]
                        return ret
                elif func == '+':
                    ret = self.parse(0, line, f, sp, args)[2]
                    for i in range(1, len(args)):
                        ret += self.parse(i, line, f, sp, args)[2]
                    return ret
                elif func == 'x':
                    ret = self.parse(0, line, f, sp, args)[2]
                    for i in range(1, len(args)):
                        ret *= self.parse(i, line, f, sp, args)[2]
                    return ret
                elif func == '/':
                    ret = self.parse(0, line, f, sp, args)[2]
                    for i in range(1, len(args)):
                        ret /= self.parse(i, line, f, sp, args)[2]
                    return ret
                elif func == '//':
                    ret = self.parse(0, line, f, sp, args)[2]
                    for i in range(1, len(args)):
                        ret //= self.parse(i, line, f, sp, args)[2]
                    return ret
                elif func == '%':
                    ret = self.parse(0, line, f, sp, args)[2]
                    for i in range(1, len(args)):
                        ret %= self.parse(i, line, f, sp, args)[2]
                    return ret
                elif func == '^':
                    ret = self.parse(0, line, f, sp, args)[2]
                    for i in range(1, len(args)):
                        ret **= self.parse(i, line, f, sp, args)[2]
                    return ret

                # determines if a string is a digit
                elif func == 'isdigit':
                    return self.parse(0, line, f, sp, args)[2].isdigit()
                # determines if a string is alpha
                elif func == 'isalpha':
                    return self.parse(0, line, f, sp, args)[2].isalpha()

                # does something with a value as a temporary
                # variable
                elif func == 'as':

                    # temporary variable name
                    varname = self.parse(0, line, f, sp, args)[2]

                    # value to set
                    val = self.parse(1, line, f, sp, args)[2]

                    # block to execute
                    block = args[2][0]

                    # set the variable
                    self.vars[varname] = Var(varname, val)

                    # execute the block
                    ret = self.interpret(block)

                    # delete the variable
                    del self.vars[varname]

                    return ret

                # strips a str
                elif func == 'strip':
                    return self.parse(0, line, f, sp, args)[2].strip()

                # returns the MSNScript2 passed as a string
                elif func == 'async' or func == 'script' or func == 'HTML':
                    # inserts key tokens
                    return self.msn2_replace(args[0][0])

                # gets the current time
                elif func == 'now':
                    return time.time()

                # creates a private execution enviroment
                # private block will have read access to the enclosing Interpreter's
                # variables and methods
                elif func == 'private' or func == 'inherit:all':
                    block_s = args[0][0]
                    inter = Interpreter()
                    inter.parent = self
                    for vname, entry in self.vars.items():
                        try:
                            inter.vars[vname] = Var(vname, entry.value)
                        except:
                            inter.vars[vname] = Var(vname, entry)
                    for mname, entry in self.methods.items():
                        inter.methods[mname] = entry
                    return inter.interpret(block_s)

                # breaks out of the working context
                elif func == 'break':
                    self.breaking = True
                    return

                # reverses the first argument
                elif func == 'reverse':
                    return self.parse(0, line, f, sp, args)[2][::-1]

                # inherits methods only from the parent context
                elif func == 'inherit:methods':
                    for methodname in self.parent.methods:
                        self.methods[methodname] = self.parent.methods[methodname]
                    return True

                # inherits variables from parent context
                elif func == 'inherit:vars':
                    for varname in self.parent.vars:
                        self.vars[varname] = self.parent.vars[varname]
                    return True
                
                # inherits a single variable or function
                elif func == 'inherit:single':
                    name = self.parse(0, line, f, sp, args)[2]
                    if name in self.parent.vars:
                        self.vars[name] = self.parent.vars[name]
                    elif name in self.parent.methods:
                        self.methods[name] = self.parent.methods[name]
                    return True

                # creates a new execution environment
                # new env is executed by a fresh interpreter
                # nothing is inherited from parent context
                elif func == 'new' or func == 'inherit:none':
                    inter = Interpreter()
                    inter.parent = self
                    return inter.interpret(args[0][0])

                # sets the python alias
                elif func == 'alias':
                    python_alias = self.parse(0, line, f, sp, args)[2]
                    return python_alias

                # starts a new process with the first argument as the target
                elif func == 'process':

                    # path to the process to run
                    path = self.parse(0, line, f, sp, args)[2]

                    # if windows:
                    if os.name == 'nt':
                        import subprocess
                        # runs the process
                        sub = subprocess.run(
                            args=[python_alias, 'msn2.py', path], shell=True)
                        self.processes[path] = sub
                        return sub

                    # if linux
                    elif os.name == 'posix':
                        print('[-] posix not supported yet')
                        return None
                    return None
                
                # starts a process via MSN2 code
                elif func == 'proc':
                    
                    # name of the process
                    name = self.parse(0, line, f, sp, args)[2]
                    
                    # block for the process
                    block = args[1][0]
                                        
                    # import the processes library and
                    # create a new process
                    return self.interpret(f"(import('lib/processes'),fork('{name}', async({block})))")

                # gets the pid of the working process
                elif func == 'pid':
                    return os.getpid()

                # creates a new thread to execute the block, thread
                # starts on the same interpreter
                elif func == "thread":
                    
                    # name not provided
                    if len(args) == 1:
                        global thread_serial
                        name = f"__msn2_thread_id_{thread_serial}"
                        block = args[0][0]
                    # name provided (2 arguments provided)
                    else:
                        name = self.parse(0, line, f, sp, args)[2]
                        block = args[1][0]
                    thread = threading.Thread(
                        target=self.interpret, args=(block,))
                    thread.name = name
                    self.threads[name] = [thread, self]
                    thread.start()
                    return True
                
                # creates a thread pool to execute the block
                elif func == "threadpool":
                    # get the amount of threads to create
                    thread_count = self.parse(0, line, f, sp, args)[2]
                    # get the block to execute
                    block = args[1][0]
                    # create the thread pool
                    pool = concurrent.futures.ThreadPoolExecutor(thread_count)
                    # submit the block to the pool
                    pool.submit(self.interpret, block)
                    return True
                
                
                # creates or edits thread variable
                elif func == 'tvar':
                    
                    # thread name
                    name = str(self.parse(0, line, f, sp, args)[2])
                    
                    # variable name
                    varname = str(self.parse(1, line, f, sp, args)[2])
                    
                    # variable value
                    val = self.parse(2, line, f, sp, args)[2]
                    
                    # thread var name
                    tvarname = f"_msn2_tvar_{name}_{varname}"
                    
                    # sets a thread specific variable
                    self.vars[tvarname] = Var(varname, val)
                    return val
                    
                # gets a thread variable
                elif func == 'gettvar':
                    
                    # thread name
                    name = str(self.parse(0, line, f, sp, args)[2])
                    
                    # variable name
                    varname = str(self.parse(1, line, f, sp, args)[2])
                    
                    # thread var name
                    tvarname = f"_msn2_tvar_{name}_{varname}"
                    
                    # gets the variable
                    return self.vars[tvarname].value
                
                # creates a string variable name for functions
                # that require a string variable name
                elif func == 'tvarstr':
                        
                    # thread name
                    name = str(self.parse(0, line, f, sp, args)[2])
                    
                    # variable name
                    varname = str(self.parse(1, line, f, sp, args)[2])
                    
                    # returns the string
                    return f"_msn2_tvar_{name}_{varname}"
                
                # interprets a variable by variable name
                # a and a variable method
                elif func == 'varmethod':
                    
                    # variable name
                    varname = str(self.parse(0, line, f, sp, args)[2])
                    
                    # method block
                    block = args[1][0]
                    
                    return self.interpret(f"{varname}.{block}")
                    
                # acquires the global lock
                elif func == 'acquire':
                    return auxlock.acquire()

                # releases the global lock
                elif func == 'release':
                    return auxlock.release()
                
                # acquires the pointer lock
                elif func == 'acquire:pointer':
                    return pointer_lock.acquire()
                elif func == 'release:pointer':
                    return pointer_lock.release()

                # joins the current working thread with the thread name specified
                elif func == 'join':

                    for i in range(len(args)):
                        name = self.parse(i, line, f, sp, args)[2]
                        thread = self.thread_by_name(name)
                        while thread == None:
                            None
                        thread.join()
                    return True

                # exits the working thread
                elif func == 'stop':
                    return os._exit(0)

                # tries the first argument, if it fails, code falls to the catch/except block
                # there is no finally implementation
                elif func == 'try':
                    ret = None
                    try:
                        ret = self.interpret(args[0][0])
                    except:
                        try:
                            catch_block = args[1][0]
                            ret = self.interpret(catch_block)
                        except:
                            None
                    return ret

                # waits for a certain condition to be true
                elif func == 'wait':
                    ret = None
                    # no block per tick provided
                    if len(args) == 1:
                        while not (ret := self.interpret(args[0][0])):
                            None

                    # block provided
                    elif len(args) == 2:
                        while not (ret := self.interpret(args[0][0])):
                            self.interpret(args[1][0])

                    # block with tick provided
                    elif len(args) == 3:
                        s = self.parse(2, line, f, sp, args)[2]
                        while not (ret := self.interpret(args[0][0])):
                            self.interpret(args[1][0])
                            time.sleep(s)
                    return ret
                
                # performs a an action every certain amount of seconds
                # where the amount of seconds is the first argument
                # the block is the second argument
                # third argument is optional, and is the amount of seconds
                # the interval should last for, if not provided, the interval
                # will last forever
                if func == 'interval':
                    
                    # amount of seconds
                    seconds = self.parse(0, line, f, sp, args)[2]
                    
                    # block to execute
                    block = args[1][0]
                    
                    
                    # if the interval should last for a certain amount of seconds
                    # should account for the first argument to correctly wait
                    if len(args) == 3:
                        
                        extra = self.parse(2, line, f, sp, args)[2] 
                        
                        # if time is negative, we set it to infinity
                        if extra == -1:
                            extra = float('inf')
                        
                        end = time.time() + extra
                        while time.time() < end:
                            time.sleep(seconds)
                            self.interpret(block)
                            
                    else:
                        while True:
                            time.sleep(seconds)
                            self.interpret(block)
                    

                # exports a quantity of variables or methods from the working context to the parent context,
                # ex private context -> boot context
                elif func == 'export':
                    
                    # if last argument is True, 
                    # we add the variables to the parent context's variable
                    last_arg = self.parse(len(args) - 1, line, f, sp, args)[2]
                    
                    for i in range(len(args)):
                        varname = self.parse(i, line, f, sp, args)[2]
                        if varname in self.vars:
                            if isinstance(last_arg, bool):
                                # if self.vars[varname].value is any type of number
                                if isinstance(self.vars[varname].value, (int, float, complex)):
                                    self.parent.vars[varname].value += self.vars[varname].value
                                # otherwise add every element to the parent context's variable
                                elif isinstance(self.vars[varname].value, list):
                                    for element in self.vars[varname].value:
                                        self.parent.vars[varname].value.append(element)
                            else: 
                                self.parent.vars[varname] = self.vars[varname]
                        elif varname in self.methods:
                            self.parent.methods[varname] = self.methods[varname]
                    return True

                # exports a single value as the variable name
                # first argument is the new variable name
                # second is the value to export
                elif func == 'exportas':

                    # variable name
                    varname = self.parse(0, line, f, sp, args)[2]

                    # value
                    val = self.parse(1, line, f, sp, args)[2]

                    # export to parent context
                    self.parent.vars[varname] = Var(varname, val)
                    return val

                # exports all variables and methods to the parent context
                elif func == 'exportall':
                    for varname in self.vars:
                        self.parent.vars[varname] = self.vars[varname]
                    for methodname in self.methods:
                        self.parent.methods[methodname] = self.methods[methodname]
                    return True

                # sends a command to the console, console type depends on
                # executors software of course
                elif func == 'console':
                    # os.system all arguments,
                    # returns the console output
                    # of the last argument
                    ret = None
                    for i in range(len(args)):
                        ret = os.system(self.parse(i, line, f, sp, args)[2])
                    return ret
                # Execute the command and capture the output
                # only takes one argument
                elif func == 'console:read':
                    import subprocess
                    # returns the console output
                    # of the last argument
                    process = subprocess.run(self.parse(0, line, f, sp, args)[2], shell=True, capture_output=True, text=True)
                    if process.returncode == 0:
                        return process.stdout
                    else:
                        return process.stderr

                # performs a get request to an http server
                # first parameter is the URL
                # second parameter is a map of parameters to sent as a request
                elif func == 'request':

                    # get URL to request from
                    url = self.parse(0, line, f, sp, args)[2]

                    try:
                        # get parameters
                        params = self.parse(1, line, f, sp, args)[2]
                    except:
                        params = None

                    response = requests.get(url=url, params=params)

                    # return response
                    return response.json()

                # requires thread-safe context, see /demos/protected.msn2
                # simulates returning of the function currently running
                # should be used cautiously, if you dont know whether to use return() or var()
                # to return a value, use var()
                elif func == 'return':
                    method = self.methods[self.loggedmethod[-1]]

                    # evaluate returning literal
                    line, as_s, ret = self.parse(0, line, f, sp, args)

                    # set return variable
                    ret_name = method.returns[0]

                    # if not a variable
                    if ret_name not in self.vars:
                        self.vars[ret_name] = Var(ret_name, None)

                    self.vars[ret_name].value = ret
                    return ret

                # gets the public IP address of the machine
                elif func == 'pubip':

                    # asks an api server for this address
                    return requests.get('https://api.ipify.org').text

                # gets the private ips of this machine
                elif func == 'privips':
                    return socket.gethostbyname_ex(socket.gethostname())[2]

                # starts an api endpoint
                elif func == 'ENDPOINT':

                    # initial API endpoint data
                    path = None
                    init_data = {}
                    port = 5000
                    host = '127.0.0.1'
                    last_arg = None

                    # 1 argument, defaults to 127.0.0.1:5000/path = {}
                    if len(args) == 1:

                        # path to endpoint
                        path = self.parse(0, line, f, sp, args)[2]
                        last_arg = path

                    # 2 arguments, defaults to 127.0.0.1:5000/path = init_data
                    elif len(args) == 2:

                        # path to endpoint
                        path = self.parse(0, line, f, sp, args)[2]

                        # json to initialize at the endpoint
                        init_data = self.parse(1, line, f, sp, args)[2]
                        
                        last_arg = init_data

                    # 3 arguments, defaults to host:port/path = init_data
                    else:

                        # host to endpoint as first argument
                        host = self.parse(0, line, f, sp, args)[2]

                        # port to endpoint as second argument
                        port = self.parse(1, line, f, sp, args)[2]

                        # path to endpoint
                        path = self.parse(2, line, f, sp, args)[2]

                        # json to initialize at the endpoint
                        init_data = self.parse(3, line, f, sp, args)[2]
                        
                        last_arg = init_data
                        if len(args) == 5:
                            last_arg = self.parse(4, line, f, sp, args)[2]

                    # prepare endpoint
                    print('serving on http://' + host + ':' + str(port) + path)
                    app = Flask(__name__)
                    cors = False
                    # if the last argument is a string with 'CORS' in it
                    # then enable CORS
                    if isinstance(last_arg, str) and 'CORS' in last_arg:
                        # enable CORS
                        print('starting with cors')
                        cors = True
                        CORS(app)


                    # disable flask messages that aren't error-related
                    log = logging.getLogger('werkzeug')
                    log.disabled = True
                    app.logger.disabled = True

                    # gets Flask Api
                    api = Api(app)     
                    curr_endpoint = self.EndPoint.make_api(init_data)

                    # logs newly created endpoint
                    self.endpoints[path] = api

                    # adds class EndPoint as a Resource to the Api with the specific path
                    # passes arg2 alongside
                    api.add_resource(curr_endpoint, path)

                    # starting flask server
                    try:

                        # if internal
                        app.run(host=host, port=port,
                                debug=False, use_reloader=False)

                    except:
                        # if external
                        try:
                            if __name__ == '__main__':
                                app.run(host='0.0.0.0', port=port,
                                        debug=False, use_reloader=False)
                        except:
                            None
                    return api

                # posts to an api endpoint
                elif func == 'POST':

                    # url to post to, defaults to localhost
                    host = self.parse(0, line, f, sp, args)[2]

                    # port to post to
                    port = self.parse(1, line, f, sp, args)[2]

                    # path after url
                    path = self.parse(2, line, f, sp, args)[2]

                    # data to post
                    data = self.parse(3, line, f, sp, args)[2]

                    # if local network
                    if host == '0.0.0.0':
                        response = requests.post(
                            url=('http://127.0.0.1:' + str(port) + path), json=data)

                    # if localhost
                    else:
                        # post to endpoint
                        response = requests.post(
                            url=('http://' + host + ':' + str(port) + path), json=data)

                    # get response
                    return response.json()

                # gets from an api endpoint
                elif func == 'GET':

                    # url to get from, defaults to localhost
                    host = self.parse(0, line, f, sp, args)[2]

                    # port to get from
                    port = self.parse(1, line, f, sp, args)[2]

                    # path after url
                    path = self.parse(2, line, f, sp, args)[2]

                    # if local network
                    if host == '0.0.0.0':
                        return requests.get(url=('http://127.0.0.1:' + str(port) + path)).json()

                    # if localhost
                    else:
                        return requests.get(url=('http://' + host + ':' + str(port) + path)).json()

                # deletes from an api endpoint
                elif func == 'DELETE':

                    # url to delete from, defaults to localhost
                    host = self.parse(0, line, f, sp, args)[2]

                    # port to delete from
                    port = self.parse(1, line, f, sp, args)[2]

                    # path after url
                    path = self.parse(2, line, f, sp, args)[2]

                    if host == '0.0.0.0':
                        response = requests.delete(
                            url=('http://127.0.0.1:' + str(port) + path))
                    else:
                        # delete from endpoint
                        response = requests.delete(
                            url=('http://' + host + ':' + str(port) + path))

                    return response.json()

                # determines if the system is windows or not
                elif func == 'windows':
                    return os.name == 'nt'

                # determines if system is linux
                elif func == 'linux':
                    return os.name == 'posix'

                # determines if system is mac
                elif func == 'mac':
                    return sys.platform == 'darwin'

                # simulates function closure
                elif func == 'end':
                    method = self.methods[self.loggedmethod[-1]]
                    self.loggedmethod.pop()
                    method.ended = True
                    return True

                # gets the value from a Var object
                elif func == 'static':
                    return self.parse(0, line, f, sp, args)[2].value

                # object instance requested
                # if the function is in the variables
                # and the variable is a class
                elif func in self.vars and isinstance(self.vars[func].value, dict):

                    # get classname to create
                    classname = func

                    # template Var obj to create from
                    var_obj = self.vars[classname].value

                    # new class instance
                    instance = {}

                    curr_arg_num = 0

                    # attributes to apply
                    for name in var_obj:

                        # if attribute is a method
                        if isinstance(var_obj[name].value, self.Method):

                            # add the method to the instance
                            instance[name] = var_obj[name].value

                            # if the method's name is 'const'
                            if var_obj[name].value.name == 'const':
                                # run the function with the argument being
                                # this instance
                                var_obj[name].value.run([instance], self, actual_args=args[1:])
                            continue

                        # if attribute is a variable
                        # value can be None
                        try:
                            instance[name] = self.parse(
                                curr_arg_num, line, f, sp, args)[2]
                            if instance[name] == None:
                                instance[name] = self.vars[classname].value[name].value
                        # if not specified, field is default value
                        except:
                            try:
                                instance[name] = var_obj.value[name].copy()
                            except:
                                var = var_obj[name]
                                instance[name] = var.value
                        curr_arg_num += 1

                    return instance

                # gets an attribute of an instance of a class
                elif func == 'getattr':

                    # name of the object to pull from
                    name = self.parse(0, line, f, sp, args)[2]

                    # current working object
                    o = self.vars[name].value

                    # get attribute to pull
                    attr = self.parse(1, line, f, sp, args)[2]
                    return o[attr]

                # sets an attribute of an instance of a class
                elif func == 'setattr':

                    # name of the object to set at
                    name = self.parse(0, line, f, sp, args)[2]

                    # current working object
                    o = self.vars[name].value

                    # name of attribute to set
                    attr = self.parse(1, line, f, sp, args)[2]

                    # value to set
                    val = self.parse(2, line, f, sp, args)[2]

                    # set the value
                    o[attr] = val
                    return val

                # PRACTICAL FUNCTIONALITY
                # started 5/20/2023

                # starts and retrieves an instance
                # of an application on the local machine
                # only properly implemented for Windows
                # uses pywinauto to do all of this
                # one argument: path to application
                # any more arguments: the existing application doesn't close
                elif func == 'app':
                                        
                    # get the path to the application
                    path = self.parse(0, line, f, sp, args)[2]            

                    # if there is not second argument, we do not kill any
                    # existing instances of the application
                    name = None
                    extension = None
                    if len(args) == 1:
                        # get the name and extension of the application
                        _sp = path.split('\\')
                        name = _sp[-1].split('.')[0]
                        extension = _sp[-1].split('.')[1]
                        
                        # use taskkill to kill the application
                        # taskkill should end the program by name, and should kill
                        # all child processes forcefully, it should also not print
                        # anything to the console                    
                        os.system(f'taskkill /f /im {name}.{extension} >nul 2>&1')
                    
                    # creates an App variable
                    return self.App(path=path, name=name, extension=extension)
            
                # connects to the first argument given that 
                # the first argument is an instance of self.App
                elif func == 'connect':
                    # connecting to
                    appl = self.parse(0, line, f, sp, args)[2]
                    a = Application(backend="uia").connect(process=appl.application.process)
                    # connect to the application
                    return self.App(path=appl.path)
                
                # starts and retrieves an instance of an Excel workbook
                # using the openpyxl library
                #
                # this method works better than the app() system call
                # as it utilizes an Excel-specific library for stability
                # and speed
                #
                # creation of a workbook can be done with the 'file' msn2 class
                elif func == 'excel':
                    path = self.parse(0, line, f, sp, args)[2]

                    # creates and returns a Workbook
                    return self.Workbook(openpyxl.load_workbook(path), path)


                # quicker conditional operator as functional prefix
                elif len(func) > 0 and func[0] == '?':
                    func = func[1:]
                    ret = None
                    if self.interpret(func):
                        ret = self.interpret(args[0][0])
                    else:

                        # else block is optional
                        try:
                            ret = self.interpret(args[1][0])
                        except:
                            None
                    return ret
                
                # executes C code and retrieves the environment
                elif func == 'C':
                    
                    # get the C code
                    c_code = self.msn2_replace(args[0][0])

                    # create a directory for the C code
                    # if it does not exist
                    exec_folder_path = '_exec'
                    
                    # if the folder does not exist, create it
                    if not os.path.exists(exec_folder_path):
                        os.mkdir(exec_folder_path)
                    
                    # create a file for the C code
                    # and write the C code to it

                    # get the amount of files in the directory
                    # and use that as the file name
                    file_num = len(os.listdir(exec_folder_path))
                    file_name = f'{exec_folder_path}/c{file_num}.c'
                    if len(args) == 2:
                        c_code = self.parse(1, line, f, sp, args)[2]
                    with open(file_name, 'w') as f:
                        f.write(c_code)

                    
                    # creates a new process
                    # and executes the C code
                    # returns the environment
                    # including the out and variables
                    def retrieve_c_environment(c_code):
                        import subprocess

                        # executable
                        executable = f'{exec_folder_path}/c{file_num}.exe'

                        # create a new process
                        # and execute the C code
                        compiled_code = subprocess.run(
                            ['gcc', file_name, '-o', executable],
                            
                            # capture the output
                            capture_output=True,
                            text=True
                        )
                        
                        # if there's an error, print it
                        if len(compiled_code.stderr) > 0:
                            return {'out': '', 'err': compiled_code.stderr}

                        # run the executable
                        compiled_code = subprocess.run(
                            [executable],
                            # capture the output
                            capture_output=True,
                            text=True
                        )

                        # get the output and error
                        out = compiled_code.stdout
                        err = compiled_code.stderr
                        
                        # get the environment
                        # env = out.split('\n')[-2]
                        # env = env.replace('\'', '"')
                        # env = json.loads(env)
                        return {'out': out, 'err': err}

                    # execute the C code
                    return retrieve_c_environment(c_code)
                
                # executes JavaScript code and retrieves the environment
                # no compilation is needed, the code is executed via
                # node __filename__.js
                elif func == 'JS':
                    
                    # get the JavaScript code
                    js_code = self.msn2_replace(args[0][0])

                    # create a directory for the JavaScript code
                    # if it does not exist
                    exec_folder_path = '_exec'
                    
                    # if the folder does not exist, create it
                    if not os.path.exists(exec_folder_path):
                        os.mkdir(exec_folder_path)
                    
                    # create a file for the JavaScript code
                    # and write the JavaScript code to it

                    # get the amount of files in the directory
                    # and use that as the file name
                    file_num = len(os.listdir(exec_folder_path))
                    file_name = f'{exec_folder_path}/js{file_num}.js'
                    
                    # if JS() has two arguments, the second is the name of
                    # the file, excluding .js
                    if len(args) == 2:
                        file_name = f'{exec_folder_path}/{self.parse(1, line, f, sp, args)[2]}.js'
                    if len(args) == 3:
                        js_code = self.parse(2, line, f, sp, args)[2]
                    with open(file_name, 'w') as f:
                        f.write(js_code)

                    
                    # creates a new process
                    # and executes the JavaScript code
                    # returns the environment
                    # including the out and variables
                    def retrieve_js_environment(js_code):
                        import subprocess

                        # executable
                        executable = file_name
        
                        # create a new process
                        # and execute the JavaScript code
                        compiled_code = subprocess.run(
                            ['node', file_name],
                            
                            # capture the output
                            capture_output=True,
                            text=True
                        )

                        # get the output and error
                        out = compiled_code.stdout
                        err = compiled_code.stderr

                        # # remove '/temp.exe'
                        # os.remove('temp.exe')
                        # if there is an error, print it
                        if len(err) > 0:
                            print(err)
                        
                        # get the environment
                        # env = out.split('\n')[-2]
                        # env = env.replace('\'', '"')
                        # env = json.loads(env)
                        
                        # remove a succeeding newline
                        # if it exists
                        if len(out) > 0 and out[-1] == '\n':
                            out = out[:-1]
                        
                        return {'out': out, 'err': err}

                    # execute the JavaScript code
                    return retrieve_js_environment(js_code)
                
                # compiles and executes Java code and retrieves the environment
                elif func == 'JAVA':
                    
                    java_code = self.msn2_replace(args[0][0])
                    
                    # create a directory for the Java code
                    # if it does not exist
                    exec_folder_path = '_exec'
                    
                    # if the folder does not exist, create it
                    if not os.path.exists(exec_folder_path):
                        os.mkdir(exec_folder_path)
                        
                    # create a file for the Java code
                    # and write the Java code to it

                    # get the amount of files in the directory
                    # and use that as the file name
                    file_num = len(os.listdir(exec_folder_path))
                    file_name = f'{exec_folder_path}/java{file_num}.java'

                    # if JAVA() has two arguments, the second is the name of
                    # the file, excluding .java
                    if len(args) == 2:
                        file_name = f'{exec_folder_path}/{self.parse(1, line, f, sp, args)[2]}.java'
                    if len(args) == 3:
                        java_code = self.parse(2, line, f, sp, args)[2]
                    with open(file_name, 'w') as f:
                        f.write(java_code)

                        
                    # creates a new process
                    # and executes the Java code
                    # returns the environment
                    # including the out and variables
                    def retrieve_java_environment(java_code):
                        import subprocess

                        # create a new process
                        # and execute the Java code
                        compiled_code = subprocess.run(
                            ['javac', file_name],
                            
                            # capture the output
                            capture_output=True,
                            text=True,
                            shell=True
                        )
                        
                        # run the code
                        compiled_code = subprocess.run(
                            ['java', '-cp', exec_folder_path, f'{file_name}'],
                            
                            # capture the output
                            capture_output=True,
                            text=True,
                            shell=True
                        )

                        # get the output and error
                        out = compiled_code.stdout
                        err = compiled_code.stderr

                        # # remove '/temp.exe'
                        # os.remove('temp.exe')
                        # if there is an error, print it
                        if len(err) > 0:
                            print(err)
                        
                        # get the environment
                        # env = out.split('\n')[-2]
                        # env = env.replace('\'', '"')
                        # env = json.loads(env)
                        
                        # remove a succeeding newline
                        # if it exists
                        if len(out) > 0 and out[-1] == '\n':
                            out = out[:-1]
                        
                        return {'out': out, 'err': err}
                    
                    # execute the Java code
                    return retrieve_java_environment(java_code)
                    
                    
                    
                
                # inline function, takes any amount of instructions
                # returns the result of the last instruction
                elif func == "=>" or (func == '' and objfunc == ''):
                    ret = None
                    for i in range(len(args)):
                        arguments = args[i]

                        # current instruction
                        ins_s = arguments[0]

                        line, ret = self.convert_arg(ins_s, line, f, sp, args)
                    return ret
                
                # if the function, when parsed, is an integer,
                # then it is a loop that runs func times
                elif (_i := get_int(func)) != None:
                    ret = None
                    for _ in range(_i):
                        for arguments in args:
                            ins_s = arguments[0]
                            line, ret = self.convert_arg(ins_s, line, f, sp, args)
                    return ret

                # # if the function is a variable name
                elif func in self.vars:
                    # value
                    val = self.vars[func].value
                    # if the variable is an integer,
                    # run the arguments as blocks inside
                    # that many times
                    if isinstance(val, int):
                        ret = None
                        for _ in range(val):
                            for arguments in args:
                                ins_s = arguments[0]
                                line, ret = self.convert_arg(ins_s, line, f, sp, args)
                        return ret
                    
                    
                    # otherwise return the value
                    return val
                
                # mouse pointer operations
                elif obj.startswith('pointer'):
                    
                    # thread based action?
                    p_thread = False
                    
                    # return
                    ret = '<msnint2 class>'
                    
                    # determine if thread based pointer action has been requested
                    if obj.endswith(':lock'):
                        p_thread = True
                        pointer_lock.acquire()
                    
                    # gets the current position of the mouse
                    if objfunc == 'getpos' or objfunc == 'pos' or objfunc == 'position':
                        ret = win32api.GetCursorPos()
                    
                    # moves the mouse to an x, y position
                    elif objfunc == 'move' or objfunc == 'hover':
                        ret = mouse.move(coords=(self.parse(0, line, f, sp, args)[2], self.parse(1, line, f, sp, args)[2]))
                    # right clicks the mouse
                    elif objfunc == 'click' or objfunc == 'left_click':
                        # if args are provided
                        if len(args) == 2:
                            ret = mouse.click(coords=(self.parse(0, line, f, sp, args)[2], self.parse(1, line, f, sp, args)[2]))
                        # if no args are provided
                        else:
                            ret = mouse.click(coords=win32api.GetCursorPos())
                            
                    # right clicks the mouse
                    elif objfunc == 'right_click':
                        # if args are provided
                        if len(args) == 2:
                            ret = mouse.right_click(coords=(self.parse(0, line, f, sp, args)[2], self.parse(1, line, f, sp, args)[2]))
                        # if no args are provided
                        else:
                            ret = mouse.right_click(coords=win32api.GetCursorPos())
                    # double clicks the mouse
                    elif objfunc == 'double_click':
                        # if args are provided
                        if len(args) == 2:
                            ret = mouse.double_click(coords=(self.parse(0, line, f, sp, args)[2], self.parse(1, line, f, sp, args)[2]))
                        # if no args are provided
                        else:
                            ret = mouse.double_click(coords=win32api.GetCursorPos())
                    # scrolls the mouse wheel to the bottom of the page
                    elif objfunc == 'scroll_bottom':
                        ret = mouse.scroll(wheel_dist=9999999, coords=win32api.GetCursorPos())
                    # scrolls the mouse wheel to the top of the page
                    elif objfunc == 'scroll_top':
                        ret = mouse.scroll(wheel_dist=-9999999, coords=win32api.GetCursorPos())
                    elif objfunc == 'scroll':
                        ret = mouse.scroll(wheel_dist=self.parse(0, line, f, sp, args)[2], coords=win32api.GetCursorPos())
                        
                    # determines if the left mouse button is down
                    elif objfunc == 'left_down':
                        ret = win32api.GetKeyState(0x01) < 0   
                    # determines if the right mouse button is down
                    elif objfunc == 'right_down':
                        ret = win32api.GetKeyState(0x02) < 0
                    # waits for the left button to be pressed
                    elif objfunc == 'wait_left':
                        while True:
                            if win32api.GetKeyState(0x01) < 0:
                                break
                        ret = True
                    # waits for the right button to be pressed
                    elif objfunc == 'wait_right':
                        while True:
                            if win32api.GetKeyState(0x02) < 0:
                                break
                        ret = True
                    # waits for a click
                    # waits for the left button to be pressed down
                    # then waits for it to be released
                    elif objfunc == 'wait_left_click':
                        while True:
                            if win32api.GetKeyState(0x01) < 0:
                                break
                        while True:
                            if win32api.GetKeyState(0x01) >= 0:
                                break
                        ret = True
                    # waits for the right button to be pressed down
                    # then waits for it to be released
                    elif objfunc == 'wait_right_click':
                        while True:
                            if win32api.GetKeyState(0x02) < 0:
                                break
                        while True:
                            if win32api.GetKeyState(0x02) >= 0:
                                break
                        ret = True
                        
                    # DIRECTIONAL MOVEMENTS
                    # moves the mouse down from its current location
                    elif objfunc == 'down':
                        curr_x, curr_y = win32api.GetCursorPos()
                        ret = mouse.move(coords=(curr_x, curr_y + self.parse(0, line, f, sp, args)[2]))
                    # moves the mouse up from its current location
                    elif objfunc == 'up':
                        curr_x, curr_y = win32api.GetCursorPos()
                        ret = mouse.move(coords=(curr_x, curr_y - self.parse(0, line, f, sp, args)[2]))
                    # moves the mouse left from its current location
                    elif objfunc == 'left':
                        curr_x, curr_y = win32api.GetCursorPos()
                        ret = mouse.move(coords=(curr_x - self.parse(0, line, f, sp, args)[2], curr_y))
                    # moves the mouse right from its current location
                    elif objfunc == 'right':
                        curr_x, curr_y = win32api.GetCursorPos()
                        ret = mouse.move(coords=(curr_x + self.parse(0, line, f, sp, args)[2], curr_y))
                    
                    # drags the mouse
                    # takes 4 or 5 arguments
                    # the first two are the starting coordinates
                    # the second two are the ending coordinates
                    # 5th argument is speed from 0-100
                    elif objfunc == 'drag':
                        start = (self.parse(0, line, f, sp, args)[2], 
                                    self.parse(1, line, f, sp, args)[2])
                        end = (self.parse(2, line, f, sp, args)[2], 
                                     self.parse(3, line, f, sp, args)[2])
                        # presses the mouse down at the coordinates
                        mouse.press(coords=start)

                        # slowly moves the mouse to the end coordinates
                        # this is to prevent the mouse from moving too fast
                        # and not dragging the object
                        # the farther the distance, the longer it takes
                        # to move the mouse
                        
                        speed = 50
                        if len(args) == 5:
                            speed = self.parse(4, line, f, sp, args)[2]
                        
                        # reverse the speed, so a speed of 50 gives
                        # end_range of 50, and a speed of 75 gives
                        # end_range of 25
                        end_range = 100 - speed
                        for i in range(0, end_range):
                            mouse.move(coords=(int(start[0] + (end[0] - start[0]) / 100 * i), 
                                                int(start[1] + (end[1] - start[1]) / 100 * i)))
                            time.sleep(0.001)

                        # releases the mouse at the end coordinates
                        mouse.release(coords=end)
                        ret = True
                        
                    # release the lock
                    if p_thread:
                        pointer_lock.release()
                    return ret
                
                # clipboard operations
                # if no arguments, the clipboard
                # is returned
                # 
                # if one argument, the text is copied
                # uses pyperclip
                elif func == 'clipboard':
                    
                    # if no arguments
                    if args[0][0] == '':
                        return pyperclip.paste()
                    # if one argument
                    else:
                        copying = self.parse(0, line, f, sp, args)[2]
                        pyperclip.copy(copying)
                        return copying

                # functional syntax I decided to add to make loops a tiny bit faster,
                # cannot receive non literal arguments
                # syntax:     3|5|i (prnt(i))
                # prnts 3\n4\n5
                elif func.count('|') == 2:
                    loop_args = func.split('|')
                    start = self.interpret(loop_args[0])
                    end = self.interpret(loop_args[1])
                    loopvar = loop_args[2]

                    # prepare loop variable
                    self.vars[loopvar] = Var(loopvar, start)

                    # obtain loop block
                    block_s = args[0][0]

                    if start < end:
                        for i in range(start, end):
                            self.vars[loopvar].value = i
                            self.interpret(block_s)

                    # reversed iteration
                    else:
                        for i in reversed(range(end, start)):
                            self.vars[loopvar].value = i
                            self.interpret(block_s)
                    return

                # fallback
                else:
                    try:
                        line = self.replace_vars2(line)
                        return eval(line, {}, {})
                    except:
                        # maybe its a variable?
                        try:
                            return self.vars[line].value
                        except:
                            # ok hopefully its a string lol
                            return line

            if obj != '':
                objfunc += c
            else:
                func += c
                
        # try a variable
        if line in self.vars:
            return self.vars[line].value

        # otherwise nothing
        # try replacing variables 2
        try:
            line = self.replace_vars2(line)
        except:
            None
        # get value of line
        try:
            return eval(line, {}, {})
        except:
            try:
                return eval(str(self.replace_vars(line)), {}, {})
            except:
                return None
    # adds a new program wide syntax
    def add_syntax(self, token, between, function):
        syntax[token] = [between, function]
        return [token, between, function]

    # replaces tokens in the string with certain
    # characters or values
    # TODO: implement linear interpretation
    def msn2_replace(self, script):
                    
        tag = '<msn2element>'
        endtag = '</msn2element>'
        # replaces whats in between the tags
        # with the interpretation of whats between the tags
        #
        # interpretation  is with self.interpret(script) 
        #
        #script(
        #     <msn2element>'hello1'</msn2element>

            

        #     <msn2element>
        #         cat('hello', 
        #             <msn2element>'hi there'</msn2element>
        #         )
        #     </msn2element>
        # )
        #
        # correct output of script() = hello1hellohi there
        
        # interpret the tags similar to the way
        # parantheticals are interpreted
        # this is a recursive function
        # open paren = tag
        # close paren = endtag
        
        def recurse_tags(scr, force_string=False, secondary_exec=False):
            # get the first tag
            # if there is no tag, return the script
            if (first := scr.find(tag)) == -1:
                return scr

            # find the matching end tag
            stack = []
            i = first + len(tag)
            while i < len(scr):
                if scr[i:i+len(endtag)] == endtag:
                    if len(stack) == 0:
                        break
                    stack.pop()
                    i += len(endtag)
                elif scr[i:i+len(tag)] == tag:
                    stack.append(tag)
                    i += len(tag)
                else:
                    i += 1

            # recursively interpret the code between the tags
            inner_code = scr[first+len(tag):i]
            inner_code = recurse_tags(inner_code)
           # scr = recurse_tags(inner_code)
            interpreted_code = self.interpret(inner_code)
            
            if force_string:
                interpreted_code = f'"{interpreted_code}"'
            if secondary_exec:
                interpreted_code = '{=' + interpreted_code + '=}'
                            
            try:
                # replace the tags with the interpreted code
                new_scr = scr[:first] + interpreted_code + scr[i+len(endtag):]
            except:
                # interpreted code is a string
                new_scr = scr[:first] + str(interpreted_code) + scr[i+len(endtag):]
            
            
            # recursively continue replacing tags in the remaining script
            return recurse_tags(new_scr)
        
        # applying <msn2element> tags
        with_msn2elements = recurse_tags(script)
        
        # switch tags
        tag = '<msn2>'
        endtag = '</msn2>'

        # applying <msn2> tags
        # for string based needs
        with_msn2 = recurse_tags(with_msn2elements, force_string=True)
        
        # applying '{=' '=}' tags
        # does the same thing as <msn2element> tags
        
        tag = '{='
        endtag = '=}'
        with_msn2 = recurse_tags(with_msn2)

        # replace hashtag marker with a hashtag
        with_msn2 = with_msn2.replace('<tag>', '#')
        with_msn2 = with_msn2.replace('<nl>', '\n')
        with_msn2 = with_msn2.replace('<rp>', ')')
        with_msn2 = with_msn2.replace('<lp>', '(')
        with_msn2 = with_msn2.replace('<rcb>', '}')
        with_msn2 = with_msn2.replace('<lcb>', '{')
        with_msn2 = with_msn2.replace('(,)', ',')
        with_msn2 = with_msn2.replace('<or>', '||')

        return with_msn2

    def run_syntax(self, key, line):
        # get everything between after syntax and before next index of syntax
        # or end of line
        inside = line[len(key):line.rindex(key)]

        # variable name
        invarname = syntax[key][0]

        # function to be run
        function = syntax[key][1]

        # store the in between for the user function
        self.vars[invarname] = Var(invarname, inside)

        return self.interpret(function)

    def replace_vars2(self, line):
        for varname, var in self.vars.items():
            try:
                val = var.value
            except:
                try:
                    val = eval(str(var), {}, {})
                except:
                    val = str(var)
            if isinstance(val, str):
                val = f'"{val}"'
            line = line.replace(f'?{varname}?', str(val))
        return line

    def thread_by_name(self, name):
        try:
            # thread exists
            return self.env_by_name(name)[0]
        except:

            # thread does not exist (yet)
            return None

    def env_by_name(self, name):
        for threadname in self.threads.keys():
            if threadname == name:
                return self.threads[threadname]
        return None

    def parse(self, arg_number, line, f, sp, args):
        as_s = args[arg_number][0]
        line, ret = self.convert_arg(as_s, line, f, sp, args)
        return line, as_s, ret

    def convert_arg(self, ins_s, line, f, sp, args):
        ret = self.interpret(ins_s)
        return line[:f + sp + args[0][1] + 1] + str(ret) + line[f + sp + args[0][2] + 1:], ret

    def shell(self):
        ip = None
        while ip != 'exit':
            ip = input(">>> ")
            self.interpret(ip)

    def logg(self, msg, line):
        self.log += "[*] " + msg + " : " + line + "\n"

    def err(self, err, msg, line):
        if msg == '':
            errmsg = "[-] " + err + " : " + line
        else:
            errmsg = "[-] " + err + " : " + msg + " : " + line
        self.out += errmsg + "\n"
        self.log += errmsg + "\n"
        print(errmsg)
        return errmsg

    def __del__(self):
        None

    def method_args(self, line, j):
        argstring = ''
        instring = False
        for k in range(j + 1, len(line)):
            if line[k] == '"' and not instring:
                instring = True
            elif line[k] == '"' and instring:
                instring = False
            if not instring:
                if line[k] != ' ':
                    if line[k] == ')':
                        break
                    argstring += line[k]
            else:
                argstring += line[k]
        return argstring.split(','), k

    def thread_split(self, line):
        self.interpret(line)

    # splits a process
    def process_split(self, line):
        inter = Interpreter()
        return inter.interpret(line)

    def var_exists(self, varname):
        if varname in self.vars:
            return True
        return False

    def var_names(self):
        names = []
        for key in self.vars:
            names.append(key)
        return names

    def is_py_str(self, line):
        try:
            return line[0] == '"' and line[len(line) - 1] == '"'
        except:
            return False

    def is_str(self, line):
        return line[0] == '<' and line[len(line) - 1] == '>'

    # gets the variable value from the variable name
    def get_var(self, name):
        return self.vars[name].value

    # extracts the argument lines from the merged arguments passed
    def get_args(self, line):
        args = []
        l = len(line)
        arg = ''
        start = 0
        p = 0
        a = 0

        s = 0
        indouble = False

        s2 = 0
        insingle = False

        b = 0
        for i in range(l + 1):
            c = ''
            try:
                c = line[i]
            except:
                None

            if c == '[' and not s2 > 0 and not s > 0:
                a += 1
            if c == ']' and not s2 > 0 and not s > 0:
                a -= 1

            if c == '(' and not s2 > 0 and not s > 0:
                p += 1
            if c == ')' and not s2 > 0 and not s > 0:
                p -= 1

            if not self.in_string(s, s2):
                if c == '{':
                    b += 1
                if c == '}':
                    b -= 1

            if not indouble and not s2 > 0 and c == '"':
                s += 1
                indouble = True
            elif indouble and c == '"':
                s -= 1
                indouble = False

            if not insingle and not s > 0 and c == "'":
                s2 += 1
                insingle = True
            elif insingle and c == "'":
                s2 -= 1
                insingle = False

            # print(line)
            # print (f"""
            # c: {c}
            # p: {p}
            # a: {a}
            # s: {s}
            # s2: {s2}

            # """)

            if c == ',' and s == 0 and s2 == 0 and p == 0 and a == 0 and b == 0:
                args.append([arg, start, start + len(arg)])
                start = i + 1
                arg = ''
                continue
            elif i == l:
                args.append([arg, start, start + len(arg)])
            arg += c
        return args

    def in_string(self, s, s2):
        return s > 0 or s2 > 0

    def interpret_msnscript_1(self, line):

        # parse all text in the line for text surrounded by %
        funccalls = []
        infunc = False
        func = ''
        for i in range(0, len(line)):
            if line[i] == '|' and not infunc:
                infunc = True
            elif line[i] == '|' and infunc:
                infunc = False
                funccalls.append(func)
                func = ''
            elif infunc:
                func += line[i]

        # for each instance of an msn2 reference
        for call in funccalls:
            line = line.replace('|' + call + '|', str(self.interpret(call)))
        element = ''
        variable = ''

        for i in range(0, len(line)):
            c = line[i]
            if c != ' ':
                if c == '+' and line[i + 1] == '=':
                    variable = element
                    element = ''
                    for j in range(i + 2, len(line)):
                        element += line[j]

                    # if element is a number
                    if isinstance(element, float) or isinstance(element, int):
                        self.vars[variable].value += self.interpret(element)
                    # if element is a string
                    elif isinstance(element, str):
                        try:
                            self.vars[variable].value += self.interpret(
                                element)
                        except:
                            self.vars[variable].value += self.interpret(
                                element)
                    return self.vars[variable].value
                elif c == '-' and line[i + 1] == '=':
                    variable = element
                    element = ''
                    for j in range(i + 2, len(line)):
                        element += line[j]
                    self.vars[variable].value -= self.interpret(element)
                    return self.vars[variable].value
                elif c == '*' and line[i + 1] == '=':
                    variable = element
                    element = ''
                    for j in range(i + 2, len(line)):
                        element += line[j]
                    self.vars[variable].value *= self.interpret(element)
                    return self.vars[variable].value
                elif c == '/' and line[i + 1] == '=':
                    variable = element
                    element = ''
                    for j in range(i + 2, len(line)):
                        element += line[j]
                    self.vars[variable].value /= self.interpret(element)
                    return self.vars[variable].value
                elif c == '=':
                    variable = element
                    element = ''
                    string = False
                    array = False
                    for j in range(i+1, len(line)):
                        if line[j] == '"':
                            string = True
                        if line[j] == '[':
                            array = True
                        element += line[j]
                    self.vars[variable] = Var(
                        variable, self.interpret(element))
                    return self.vars[variable].value

                elif "{" in element and "}" in element:
                    if ":" not in element:
                        expression = element.replace("{", '').replace("}", '')
                        willrun = self.boolean(expression)

                        if willrun:
                            line = line[line.index("}") + 1:]
                            self.interpret(line)
                    else:
                        times = self.loop(element)
                        first = times[0]
                        last = times[1]
                        optvar = times[2]
                        try:
                            first = self.vars[first].value
                        except:
                            None
                        try:
                            last = self.vars[last].value
                        except:
                            None
                        first = int(first)
                        last = int(last)
                        self.vars[optvar] = Var(optvar, first)
                        if first <= last:
                            for i in range(first, last):
                                if optvar in self.vars:
                                    self.vars[optvar].value = i
                                self.interpret_msnscript_1(
                                    line.replace(' ', '').replace(element, ''))
                        else:
                            for i in range(last, first):
                                if optvar != '':
                                    self.vars[optvar].value = i
                                self.interpret_msnscript_1(
                                    line.replace(' ', '').replace(element, ''))
                    break
                elif c == "&":
                    None

                else:
                    element += c

    def boolean(self, expression):
        first = ''
        last = ''
        op = ''
        for i in range(0, len(expression)):
            if expression[i] == '&' and expression[i + 1] == '&' and expression[i + 2] != '-' and expression[i + 2] != '+':
                for j in range(i + 2, len(expression)):
                    last += expression[j]
                op = '=='
                break
            elif expression[i] == '!' and expression[i + 1] == '&' and expression[i + 2] == '&' and expression[i + 3] != '-' and expression[i + 3] != '+':
                for j in range(i + 3, len(expression)):
                    last += expression[j]
                op = '!='
                break
            elif expression[i] == '&' and expression[i + 1] == '&' and expression[i + 2] == '-' and expression[i + 3] != 'e':
                for j in range(i + 3, len(expression)):
                    last += expression[j]
                op = '<'
                break
            elif expression[i] == '&' and expression[i + 1] == '&' and expression[i + 2] == '+' and expression[i + 3] != 'e':
                for j in range(i + 3, len(expression)):
                    last += expression[j]
                op = '>'
                break
            elif expression[i] == '&' and expression[i + 1] == '&' and expression[i + 2] == '-' and expression[i + 3] == 'e':
                for j in range(i + 4, len(expression)):
                    last += expression[j]
                op = '<='
                break
            elif expression[i] == '&' and expression[i + 1] == '&' and expression[i + 2] == '+' and expression[i + 3] == 'e':
                for j in range(i + 4, len(expression)):
                    last += expression[j]
                op = '>='
                break
            else:
                first += expression[i]

        try:
            firsteval = self.vars[first].value
        except KeyError:
            firsteval = self.evaluate(first, 'string')

        try:
            lasteval = self.vars[last].value
        except KeyError:
            lasteval = self.evaluate(last, 'string')

        if op == '==':
            return firsteval == lasteval
        if op == '!=':
            return firsteval != lasteval
        if op == '<':
            return firsteval < lasteval
        if op == '>':
            return firsteval > lasteval
        if op == '<=':
            return firsteval <= lasteval
        if op == '>=':
            return firsteval >= lasteval
        return False

    # scrapes all html elements from a URL
    def html_all_elements(self, url):

        # obtains a response from the URL
        response = requests.get(url)

        soup = BeautifulSoup(response.content, 'html5lib')

        # obtains all html elements
        return soup.find_all()

    def loop(self, section):
        first = ''
        last = ''
        optvar = ''
        optvarfound = False
        for i in range(1, len(section)):
            if section[i] == ':':
                for j in range(i + 1, len(section)):
                    if section[j] == ':':
                        for k in range(j + 1, len(section)):
                            if section[k] == '}':
                                optvarfound = True
                                break
                            optvar += section[k]
                    if section[j] == '}':
                        break
                    if optvarfound:
                        break
                    last += section[j]
                break
            first += section[i]
        return first, last, optvar

    def evaluate(self, postop, type):
        new = postop
        try:
            return eval(new, {}, {})
        except:
            None
        if type == 'number':
            new = self.replace_vars(new)
            try:
                return eval(new, {}, {})
            except:
                return eval(str(new), {}, {})
        elif type == 'string':
            return self.string(postop)
        elif type == 'array':
            return self.array(postop)
        elif type == 'unknown':
            new = self.replace_vars(new)
            try:
                evaluation = eval(str(self.string(new)), {}, {})
            except:
                try:
                    evaluation = self.array(self.vars[new])
                except:
                    try:
                        evaluation = self.vars[self.string(eval(new, {}, {}))]
                    except:
                        try:
                            evaluation = eval(new, {}, {})
                        except:
                            evaluation = new

            return evaluation

    def me(self):
        return str(self).replace(' ', '').replace('<', '').replace('>', '').replace('Interpreter', '')

    def string(self, string):

        strn = ''
        isv = False
        try:
            string = self.vars[string]
            isv = True
        except KeyError:
            None

        # if isv and '"' not in string:
        #     string = '"' + string + '"'
        try:
            strn = eval(string, {}, {})
        except:
            None
        try:
            for var in self.vars:
                strn = strn.replace("{" + var + "}", str(self.vars[var].value))
            for method in self.methods.keys():
                toprint = ''
                body = self.methods[method].body
                for line in body:
                    toprint += line + '\\n'
                strn = strn.replace("{" + method + "}", toprint)
        except:
            None
        return strn

    def array(self, postop):
        array = []
        try:
            array = eval(postop, {}, {})
        except:
            None
        return array

    # prints text with a box around it
    def bordered(text):
        lines = text.splitlines()
        width = max(len(s) for s in lines)
        res = ['┌' + '─' * width + '┐']
        for s in lines:
            res.append('│' + (s + ' ' * width)[:width] + '│')
        res.append('└' + '─' * width + '┘')
        return '\n'.join(res)

    class Method:
        def __init__(self, name, interpreter):
            self.name = name
            self.args = []
            self.body = []
            self.returns = []
            self.ended = False
            self.interpreter = interpreter

        def add_arg(self, arg):
            self.args.append(arg)

        def add_body(self, body):
            self.body.append(body)

        def add_return(self, ret):
            self.returns.append(ret)

        def run(self, args, inter, actual_args=None):

            def try_var(func_var, func_insert):
                try:
                    inter.vars[func_var] = Var(func_var, func_insert)
                except:
                    inter.vars[func_var] = func_insert 
            for i in range(len(self.args)):
                if actual_args:
                    try:
                        if self.is_str(actual_args[i][0]):
                            try_var(self.args[i], args[i])
                            continue 
                        
                    # index out of bounds     
                    except IndexError:
                        pass  
                try:
                    inter.vars[self.args[i]] = inter.vars[args[i]]
                except:
                    try_var(self.args[i], args[i])
            method_ret = None
            for line in self.body:
                method_ret = inter.interpret(line)
            return method_ret
                
        def is_str(self, value):
            return (value[0] == '"' and value[-1] == '"') or (value[0] == "'" and value[-1] == "'")

    class EndPoint(Resource):

        @classmethod
        def make_api(cls, response):
            cls.response = response
            return cls

        # GET
        def get(self):
            return self.response

        # POST
        def post(self):

            # obtains current endpoint data
            current_data = self.response

            # updates current data with data to post
            current_data.update(request.get_json())

            # updates next response
            self.make_api(current_data)
            return current_data

        # DELETE
        def delete(self):
            self.make_api({})
            return self.response

    # class for an App

    class App:
        # constructor
        def __init__(self, path, application=None, name=None, extension=None):

            # path of application being launched
            self.path = path
            
            if name:
                self.name = name
                self.extension = extension
            else:
                _spl = path.split('\\')[-1].split('.')
                # extension of the application
                self.extension = _spl[-1]
                self.name = _spl[0]
            
            # pwinauto application object
            self.application = application

    # element for an application
    class AppElement:

        # constructor
        def __init__(self, window, name):

            # creates a modified window
            self.window = window
            # # set the window
            # self.window = window
            # set the name
            self.name = name
            
        # gets the text of the window
        def window_text(self):
            return self.name

        # gets all children of the window
        def children(self):
            return self.window.children()

        # sets the focus to the window
        def set_focus(self):
            self.window.set_focus()
            
        # gets the properties of the window
        def get_properties(self):
            return self.window.get_properties()
        
        # gets the highest level parent of this element
        def top_level_parent(self):
            return self.window.top_level_parent()
        
        # computes the height of the window
        def height(self):
            try:
                return self.window.get_properties()['rectangle'].bottom - self.window.get_properties()['rectangle'].top
            except:
                return
        # computes the width of the window
        def width(self):
            try:
                return self.window.get_properties()['rectangle'].right - self.window.get_properties()['rectangle'].left
            except:
                return
        # string
        def __str__(self):
            return Interpreter.bordered(f'Text: {self.name if self.name else "[No Text Found]"}\nSize:\
{f"{self.width()}x{self.height()}"}\nObject:\n{self.window}')
    
    # class for a button
    class Button(AppElement):
            
        # constructor
        def __init__(self, window, name):

            # call super constructor
            super().__init__(window, name)

        # clicks the button
        def click(self):
            self.window.click()
            
        # right clicks the button
        def right_click(self):
            self.window.click_input(button='right')
    
    # class for a Link
    class Link(AppElement):
                    
        # constructor
        def __init__(self, window, name):

            # call super constructor
            super().__init__(window, name)
            
    # class for a Menu
    class Menu(AppElement):
                
        # constructor
        def __init__(self, window, name):

            # call super constructor
            super().__init__(window, name)
    

    # class for a ToolBar
    class ToolBar(AppElement):
                    
        # constructor
        def __init__(self, window, name):

            # call super constructor
            super().__init__(window, name)
            
    # class for a scrollbar
    class ScrollBar(AppElement):
                        
        # constructor
        def __init__(self, window, name):

            # call super constructor
            super().__init__(window, name)  

    # class for TabItems
    class TabItem(AppElement):
        # constructor
        def __init__(self, window, name):

            # call super constructor
            super().__init__(window, name)  
            
    # class for Hyperlink
    class Hyperlink(AppElement):
        # constructor
        def __init__(self, window, name):

            # call super constructor
            super().__init__(window, name)
    
    # class for Inputs
    class Input(AppElement):
        # constructor
        def __init__(self, window, name):

            # call super constructor
            super().__init__(window, name)  
            
        # types text into the input
        def type_keys(self, text):
            self.window.type_keys(text)
            
    # class for Tables
    class Table(AppElement):
        # constructor
        def __init__(self, window, name):

            # call super constructor
            super().__init__(window, name)  
            
            
    # ------------------------------------
    # working with Excel
    class Workbook:
        
        # constructor
        def __init__(self, workbook, path) -> None:
            self.workbook = workbook
            self.path = path
            
    # sheet class
    class Sheet(Workbook):
        
        def __init__(self, sheet, title, workbook, path) -> None:
            super().__init__(workbook, path)
            self.sheet = sheet
            self.title = title