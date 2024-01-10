import os
import json
import sys
import pandas as pd
import numpy as np
from dateutil import parser as date_parser
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from pdfminer.converter import PDFPageAggregator
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfdevice import PDFDevice
import pdfminer
import PyPDF2
import tabula
from tabula import read_pdf
import streamlit as st
from langchain.chat_models import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.agents import create_pandas_dataframe_agent
import csv
import datetime
from datetime import datetime
import PyPDF2
from tokenize import Number
import pandas
import numpy as np
import requests
import time
from openpyxl import Workbook
import math
import io
import re
import boto3
import json
import tensorflow
from transformers import AutoTokenizer
import nltk
from nltk.corpus import stopwords
nltk.download("stopwords")
# nltk.download('omw-1.4')
nltk.download('punkt')
from nltk.tokenize import word_tokenize

def Average(lst):
    return sum(lst) / len(lst)

def getUnit(x):
    if x<10:
        return str(0)+str(x)
    else:
        return x

def set_report(df):
    workbook = Workbook()
    sheet = workbook.active
    
    no_trans_type={'credit':{'total_amount':0 ,'total_count':0},'debit':{'total_amount':0 ,'total_count':0}}

    # totalcredit.groupby('Txn Date')
    grp = df.groupby('Transaction Type')
    for trn_type, group in grp:
    #     print(trn_type)
        if(trn_type =='credit'):
            no_trans_type['credit']['total_amount']=sum(group['Transaction Amount'])
            no_trans_type['credit']['total_count']=len(group['Transaction Amount'])
            no_trans_type['credit']['highest']=max(group['Transaction Amount'])
            no_trans_type['credit']['lowest']=round(Average(group['Transaction Amount']),2)
        else:
            no_trans_type['debit']['total_amount']=sum(group['Transaction Amount'])
            no_trans_type['debit']['total_count']=len(group['Transaction Amount'])
            no_trans_type['debit']['highest']=max(group['Transaction Amount'])
            no_trans_type['debit']['lowest']=round(Average(group['Transaction Amount']),2)
    #     print(sum(group['Transaction Amount']))
    #     print(len(group['Transaction Amount']))
    
    # category wise transactions
    cate_tran={}
    grp2 = df.groupby('Category')
    for trn_type, group in grp2:
        cate_tran[trn_type]=len(group)
        
    df['month']=''
    for x in df.index:
        df.loc[x,'month']=df.loc[x,'Transaction Date'][0:-3]    
    
    newdata = df.sort_values(['Transaction Amount'],ascending=False).groupby(['month','Transaction Type'])
    # top five credit,debit monthly
    monthaly_data={}
    grp3 = df.groupby(['month','Transaction Type'])
    for trn_type, group in grp3:  
        print(trn_type)
        monthaly_data[trn_type]={'count':len(group),'amount':sum(group['Transaction Amount']),'max':max(group['Transaction Amount']),'min':min(group['Transaction Amount']),'avg':round(Average(group['Transaction Amount']),2)}
    
    # top five creditand debit by months 
    for trn_type, group in newdata:
        print(trn_type)
        print(group.head())
        
    eod=pd.DataFrame()
    
    start_Date=df.loc[0,'Transaction Date']
    end_Date=df.loc[len(df)-1,'Transaction Date']
    print(end_Date)
    
    # eod['date']=''
    eod['Transaction Date']=pd.date_range(start=start_Date,end=end_Date)
    eod['Balance']=0
    
    di_mon={}
    for put in df.index:
        if(df.loc[put,'Transaction Date'] not in di_mon):
            di_mon[df.loc[put,'Transaction Date']]=df.loc[put,'Balance']
        else:
            di_mon[df.loc[put,'Transaction Date']]=df.loc[put,'Balance']

    for x in eod.index:
        cuur=eod.loc[x][0]
        uni=str(cuur.year)+'-'+str(getUnit(cuur.month)) +'-'+str(getUnit(cuur.day))
        if(uni not in di_mon):
            di_mon[uni]=0
        
        
        # Create a new workbook
    workbook = Workbook()

    # Get the active sheet
    sheet = workbook.active
    sheet2 = workbook.create_sheet("monthly top five credit")  # Create a new sheet with a specific name
    sheet3 = workbook.create_sheet("monthly top five debit")
    # Define your dictionary

    mon_overview_credit=[
    '',
    'Total No. of Credit Transactions',
    'Total Amount of Credit Transactions',
    'Highest Transaction Value',
    'Lowest Transaction Value',
    'Average Balance',
    ]

    mon_overview_debit=[
    '',
    'Total No. of Debit Transactions',
    'Total Amount of Debit Transactions',
    'Highest Transaction Value',
    'Lowest Transaction Value',
    'Average Balance',
    ]


    for row in range(1,len(mon_overview_credit)+1):
        sheet.cell(row=row, column=1, value=mon_overview_credit[row-1])

    # iterate through month data 
    itt=2
    for key, value in monthaly_data.items(): #itrerate through col
        if(key[-1]=='credit'):
            #fill column

            sheet.cell(row=1, column=itt, value=key[0])
            sheet.cell(row=2, column=itt, value=value['count'])
            sheet.cell(row=3, column=itt, value=value['amount'])
            sheet.cell(row=4, column=itt, value=value['max'])
            sheet.cell(row=5, column=itt, value=value['min'])
            sheet.cell(row=6, column=itt, value=value['avg'])
            itt+=1

    new_row=10 #new table begin
    for row in range(1,len(mon_overview_credit)+1):
        sheet.cell(row=row+new_row, column=1, value=mon_overview_debit[row-1])        
    itt2=2

    for key, value in monthaly_data.items(): #itrerate through col
        if(key[-1]=='debit'):
            #fill column

            sheet.cell(row=1+new_row, column=itt2, value=key[0])
            sheet.cell(row=2+new_row, column=itt2, value=value['count'])
            sheet.cell(row=3+new_row, column=itt2, value=value['amount'])
            sheet.cell(row=4+new_row, column=itt2, value=value['max'])
            sheet.cell(row=5+new_row, column=itt2, value=value['min'])
            sheet.cell(row=6+new_row, column=itt2, value=value['avg'])
            itt2+=1    

    # Iterate over the dictionary and write key-value pairs to cells
    sheet2 = workbook.create_sheet("Sheet 2") 
    # Create a new sheet with a specific name
    itt3=1
    itt4=1

    inrow=2
    inrow2=2

    sheet2.cell(row=1, column=1, value='month')   #set headings
    sheet2.cell(row=1, column=2, value='Account Number')
    sheet2.cell(row=1, column=3, value='Bank Name')
    sheet2.cell(row=1, column=4, value='Transaction Date')
    sheet2.cell(row=1, column=5, value='Transaction Amount')
    sheet2.cell(row=1, column=6, value='Transaction Type')
    sheet2.cell(row=1, column=7, value='Description')
    sheet2.cell(row=1, column=8, value='Category')


    sheet3.cell(row=1, column=1, value='month')   #set headings
    sheet3.cell(row=1, column=2, value='Account Number')
    sheet3.cell(row=1, column=3, value='Bank Name')
    sheet3.cell(row=1, column=4, value='Transaction Date')
    sheet3.cell(row=1, column=5, value='Transaction Amount')
    sheet3.cell(row=1, column=6, value='Transaction Type')
    sheet3.cell(row=1, column=7, value='Description')
    sheet3.cell(row=1, column=8, value='Category')
    for trn_type, group in newdata:
        print(trn_type)
    #     print(group.head())
            #fill column

        cu=group.head()    
        if 'credit' in trn_type[-1]:
            for x in cu.index:
                #Account Number   Bank Name  balance        date  transactionValue  Transaction Type  description  \
                sheet2.cell(row=inrow, column=1, value=trn_type[0])
                sheet2.cell(row=inrow, column=2, value=cu.loc[x,'Account Number'])
                sheet2.cell(row=inrow, column=3, value=cu.loc[x,'Bank Name'])
                sheet2.cell(row=inrow, column=4, value=cu.loc[x,'Balance'])
                sheet2.cell(row=inrow, column=5, value=cu.loc[x,'Transaction Date'])
                sheet2.cell(row=inrow, column=6, value=cu.loc[x,'Transaction Amount'])
                sheet2.cell(row=inrow, column=7, value=cu.loc[x,'Transaction Type'])
                sheet2.cell(row=inrow, column=8, value=cu.loc[x,'Description'])
                sheet2.cell(row=inrow, column=9, value=cu.loc[x,'Category'])
                inrow+=1
        else:       
            for y in cu.index:
                #Account Number   Bank Name  Balance        date  transactionValue  Transaction Type  description  \
                sheet3.cell(row=inrow2, column=1, value=trn_type[0])
                sheet3.cell(row=inrow2, column=2, value=cu.loc[y,'Account Number'])
                sheet3.cell(row=inrow2, column=3, value=cu.loc[y,'Bank Name'])
                sheet3.cell(row=inrow2, column=4, value=cu.loc[y,'Balance'])
                sheet3.cell(row=inrow2, column=5, value=cu.loc[y,'Transaction Date'])
                sheet3.cell(row=inrow2, column=6, value=cu.loc[y,'Transaction Amount'])
                sheet3.cell(row=inrow2, column=7, value=cu.loc[y,'Transaction Type'])
                sheet3.cell(row=inrow2, column=8, value=cu.loc[y,'Description'])
                sheet3.cell(row=inrow2, column=9, value=cu.loc[y,'Category'])
                inrow2+=1




    # Save the workbook
    workbook.save("overall_report.xlsx")

def validateJSON(jsonData):
    try:
        result = json.loads(jsonData) and type(json.loads(jsonData)) is dict
        if result:
            return True
        else:
            return False
    except ValueError as err:
        return False
    
# get bank name
def get_bankname(file_path):
    pdfFileObj = open(file_path, 'rb')
    pdfReader = PyPDF2.PdfReader(pdfFileObj)
    pageObj = pdfReader.pages[0]
    pagelen = len(pdfReader.pages[0])
    pageData = pageObj.extract_text()
    bank_name = {'ICIC':'ICICI Bank', 'SBIN':'SBI Bank','HDFC':'HDFC Bank','UTIB':'Axis Bank','IDFB':'IDFC Bank'}
    listbyline = pageData.split('\n')
    name = ''
    for x in listbyline:
        if re.search(r"IFS Code|IFSC|IFSC Code", x):
            if re.search(r"([A-Z]{4})0[0-9]{6}", x):# checkfor ifsc code
                name = bank_name[re.search(r"([A-Z]{4})0[0-9]{6}", x).group()[0:4]] # bank name extraction
                break
    return name

def sonata(y):
        x=date_parser.parse(y,dayfirst=True)
        return x


def getnumber(text):
    isNegative = False

    if isinstance(text, int)==True or isinstance(text, float)==True:
        result=float(text)
    else:
        getVals=[]
        for val in text:
            if "-" in text:
                isNegative = True
            if val.isnumeric()==True or val==".":
                getVals.append(val)
        result = "".join(getVals)
        result=float(result)
    if isNegative:
        return result * -1
    return result

# get Account Number for any bank out 5
def get_account_number(text):
    listbyline=text.split('\n')
    account_no=''
    for x in listbyline:
        if re.search(r"ACCOUNT NO :|A/C No:|Account No : ", x):
            if " " not in x.split(":")[1]:
                account_no = x.split(":")[1]
            else:
                account_no = x.split(":")[1].split(" ")[1]
        elif re.search(r"Account Number", x):
            account_no = x.split(":")[-1].replace("\t", '')
        elif re.search(r"Account No :", x):
            account_no = x.split(":")[1].split(" ")[0]
    return int(account_no)

# HDFC
def getcoord(path,tol_diff):
    fp = open(path, 'rb')
    # Create a PDF parser object associated with the file object.
    parser = PDFParser(fp)
    # Create a PDF document object that stores the document structure.
    # Password for initialization as 2nd parameter
    # password = ''
    document = PDFDocument(parser)
    # Check if the document allows text extraction. If not, abort.
    # Create a PDF resource manager object that stores shared resources.
    rsrcmgr = PDFResourceManager()
    # Create a PDF device object.
    device = PDFDevice(rsrcmgr)
    # BEGIN LAYOUT ANALYSIS
    # Set parameters for analysis.
    laparams = LAParams()
    # Create a PDF page aggregator object.
    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    # Create a PDF interpreter object.
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    output=[]
    def parse_obj(lt_objs):
        # loop over the object list
        for obj in lt_objs:
            # if it's a textbox, print text and location
            if isinstance(obj, pdfminer.layout.LTTextBoxHorizontal):
                # print(obj.bbox[0], obj.bbox[1], obj.get_text().replace('\n', '_'))
                output.append([obj.bbox[0], obj.bbox[1], obj.bbox[2], obj.bbox[3],obj.get_text().replace('\n', '_')])
            # if it's a container, recurse
            elif isinstance(obj, pdfminer.layout.LTFigure):
                parse_obj(obj._objs)
                output.append(parse_obj(obj._objs))
    # loop over all pages in the document
    for page in PDFPage.create_pages(document):
        # read the page into a layout object
        interpreter.process_page(page)
        layout = device.get_result()
        # extract text from this object
        parse_obj(layout._objs)

    x0=0
    x1=0
    x2=0
    x3=0
    x4=0
    x5=0
    x6=0
    x7=0
    # print(output)

    for i in output:
        if i!=None:
            if 'Date_' == i[4]:
                x0=i[0] + tol_diff + 10
            if 'Narration_' == i[4]:
                x1=i[0] + tol_diff + 110
            if 'Chq./Ref.No._' == i[4]:
                x2=i[2] + tol_diff + 50
            if "Value Dt Withdrawal Amt._" == i[4]:
                x3=i[2] + tol_diff
            if "Deposit Amt._" == i[4]:
                x4=i[2] + tol_diff
            if "Closing Balance_" == i[4]:
                x5=i[2] + tol_diff

    print([x0,x1,x2,x3,x4,x5])

    return [x0,x1,x2,x3,x4,x5]

def parse_hdfc(path):
    print("File name: ", path)
    tolerance=10
    columns=getcoord(path,tolerance)

    
    pdf = PyPDF2.PdfReader(open(path,'rb'))
    # pdf.decrypt(b'CONS866823038')
    pages=len(pdf.pages)

    text = pdf.pages[0].extract_text()


    a_num=get_account_number(text)
    arr= []
    for p in range(1,pages+1):

        x=tabula.read_pdf(path,guess=False,stream=False,columns=columns,pages=p,multiple_tables=True,pandas_options={'header':None})

        for i in x:
            i=i.fillna(0)

            # with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
            #     print(i)
            # print(i)
            for j in range(0,len(i)):
                    flag=0

                    try:
                        sonata(i[0][j])
                        total=i[5][j]

                        flag=1

                        if total==0:
                            flag=0

                    except Exception: pass


                    if flag!=0:

                        date=sonata(i[0][j])
                        # print(date)

                        transaction_value=0

                        transaction_type=""

                        # print(i[4][j])
                        # print(type (i[4][j]))

                        if getnumber(i[3][j])>0 and getnumber(i[4][j]) == 0:
                            transaction_value=getnumber(i[3][j])
                            transaction_type="debit"

                        if getnumber(i[4][j])>0 and getnumber(i[3][j]) == 0:
                            transaction_value=getnumber(i[4][j])
                            transaction_type="credit"


                        desc = i[1][j]
                        for k in range(1,4):
                            try:
                                val_1=float(i[3][j+k])+float(i[4][j+k])+float(i[5][j+k])
                                if val_1==0:
                                    desc=desc + i[1][j+k]
                            except Exception: break

                        balance=getnumber(i[5][j])

                        final_res={
                            "Account Number": a_num,
                            "Bank Name": "HDFC Bank",
                            "Balance": float(balance),
                            "Transaction Date": date,
                            "Transaction Amount":float(transaction_value),
                            "Transaction Type": transaction_type,
                            "Description": desc.upper()
                          }
                        
                        arr.append(final_res)

    # print(arr)
    df = pd.DataFrame(arr)
    print(df.info())
    return df

# ICICI
def getcoord1(path,tol_diff):
    fp = open(path, 'rb')
    # Create a PDF parser object associated with the file object.
    parser = PDFParser(fp)
    # Create a PDF document object that stores the document structure.
    # Password for initialization as 2nd parameter
    # password = ''
    document = PDFDocument(parser)
    # Check if the document allows text extraction. If not, abort.
    # Create a PDF resource manager object that stores shared resources.
    rsrcmgr = PDFResourceManager()
    # Create a PDF device object.
    device = PDFDevice(rsrcmgr)
    # BEGIN LAYOUT ANALYSIS
    # Set parameters for analysis.
    laparams = LAParams()
    # Create a PDF page aggregator object.
    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    # Create a PDF interpreter object.
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    output=[]
    def parse_obj(lt_objs):
        # loop over the object list
        for obj in lt_objs:
            # if it's a textbox, print text and location
            if isinstance(obj, pdfminer.layout.LTTextBoxHorizontal):
                # print(obj.bbox[0], obj.bbox[1], obj.get_text().replace('\n', '_'))
                output.append([obj.bbox[0], obj.bbox[1], obj.bbox[2], obj.bbox[3],obj.get_text().replace('\n', '_')])
            # if it's a container, recurse
            elif isinstance(obj, pdfminer.layout.LTFigure):
                parse_obj(obj._objs)
                output.append(parse_obj(obj._objs))
    # loop over all pages in the document
    for page in PDFPage.create_pages(document):
        # read the page into a layout object
        interpreter.process_page(page)
        layout = device.get_result()
        # extract text from this object
        parse_obj(layout._objs)

    x0=0
    x1=0
    x2=0
    x3=0
    x4=0
    x5=0
    x6=0
    x7=0
    # print(output)

    for i in output:
        if i!=None:
            if 'Sl_No_1_' in i[4]:
                x0=i[0] + tol_diff + 70
            if 'Transaction_' == i[4]:
                x1=i[0] + tol_diff
            if 'Transaction_Posted Date_' in i[4]:
                x2=i[0] - tol_diff
                x3=i[2] + tol_diff + 40
            # if 'Cheque no /_' == i[4]:
            if 'Remarks_SI/' in i[4]:
                x4=i[0] + tol_diff + 50
            if "Withdra_wal (Dr)_" == i[4]:
                x5=i[0] - tol_diff
            if "Deposit_" == i[4]:
                x6=i[0] - tol_diff
            if "Balance_" == i[4]:
                x7=i[0] - tol_diff

    print([x0,x1,x2,x3,x4,x5,x6,x7])

    return [x0,x1,x2,x3,x4,x5,x6,x7]

def parse_icici(path):

    tolerance=10
    columns=getcoord1(path,tolerance)
    columns.sort()
    # print("Columns ICICI: ", columns)
    
    pdf = PyPDF2.PdfReader(open(path,'rb'))
    # pdf.decrypt(b'CONS866823038')
    pages=len(pdf.pages)

    text = pdf.pages[0].extract_text()


    a_num=get_account_number(text)
    arr= []
    for p in range(1,pages+1):

        x=tabula.read_pdf(path,guess=False,stream=False,columns=columns,pages=p,multiple_tables=True,pandas_options={'header':None})

        for i in x:
            i=i.fillna(0)

            # with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
            #     print(i)
    #         # print(i)
            for j in range(0,len(i)):
                    flag=0

                    try:
                        sonata(i[1][j])
                        total=i[8][j]

                        flag=1

                        if total==0:
                            flag=0

                    except Exception: pass


                    if flag!=0:

                        date=sonata(i[1][j])
                        # print(date)

                        transaction_value=0

                        transaction_type=""

                        # print(i[4][j])
                        # print(type (i[4][j]))

                        if getnumber(i[6][j])>0 and getnumber(i[7][j])==0:
                            transaction_value=getnumber(i[6][j])
                            transaction_type="debit"

                        if getnumber(i[7][j])>0 and getnumber(i[6][j])==0:
                            transaction_value=getnumber(i[7][j])
                            transaction_type="credit"


                        desc = str(i[3][j]) + str(i[4][j])
                        for k in range(1,7):
                            try:
                                val_1=float(i[6][j+k])+float(i[7][j+k])
                                if val_1==0:
                                    desc=desc + str(i[3][j+k]) + str(i[4][j+k])
                            except Exception: break

                        balance=getnumber(i[8][j]) + getnumber(i[8][j+1])

                        final_res={
                            "Account Number": a_num,
                            "Bank Name": "ICICI Bank",
                            "Balance": float(balance),
                            "Transaction Date": date,
                            "Transaction Amount":float(transaction_value),
                            "Transaction Type": transaction_type,
                            "Description": desc.upper()
                          }
                        
                        # print(final_res)
                        arr.append(final_res)

    df = pd.DataFrame(arr)
    return df
# SBI 
def parse_sbi(filename):
    print("file name", filename)
    df = read_pdf(filename, pages="all") #address of pdf file
    for page in range(len(df)):
        descrip=''
        loc_to_update=0; #inital location to update;
        for row in range(len(df[page])):
            if(str(df[page].loc[row][0])=='nan'):
                if(str(df[page].loc[row][2])!='nan'):
                    descrip+=str(df[page].loc[row][2])
            else:
        #         this block is for updating value
        #         loc_to_update
                df[page].loc[loc_to_update][2]=descrip
                loc_to_update=row
                descrip=str(df[page].loc[row][2])
            if(row+1 == len(df[page])):
                df[page].loc[loc_to_update][2]=descrip
          
    #    
    for pg in range(len(df)):
        for x in df[pg].index:
            if(str(df[pg].loc[x][0]) == 'nan'):
                df[pg].drop(x, inplace = True)

    for p in range(len(df)):
        df[p].fillna(0, inplace = True)    

    df_big = pd.concat(df)  #make single df from list of df 
    remo_column=[]
    for col in df_big.columns:
        if 'Unnamed' in col:
            remo_column.append(col)

    new_ind=[]
    inc=0
    for x in range(len(df)):
        for i in df[x].index:
            new_ind.append(inc)
            inc+=1

    df_big.set_index(pd.Index(new_ind),inplace=True) #set new index for df
    df_big.drop(remo_column, axis=1, inplace=True) #unwanted column removed
    df_big.drop(['Txn Date','Ref No./Cheque','Branch'], axis=1, inplace=True) #unwanted column removed
    df_big['Transaction Amount']=0
    df_big['Transaction Type']=''
    # add two column tansaction type and value
    for x in df_big.index:
        if(df_big.loc[x,'Debit']==0):
            df_big.loc[x, "Transaction Amount"] =df_big.loc[x,'Credit']
            df_big.loc[x, "Transaction Type"] ='credit'
        else:    
            df_big.loc[x, "Transaction Amount"] =df_big.loc[x,'Debit']
            df_big.loc[x, "Transaction Type"] ='debit'

    df_big.drop(['Debit','Credit'], axis=1, inplace=True) #unwanted column removed
    df_big = df_big.set_axis(['Transaction Date','Description', 'Balance', 'Transaction Amount', 'Transaction Type'], axis=1)
    df_big['Bank Name']='SBI Bank'
    df_big['Transaction Date']=pd.to_datetime(df_big['Transaction Date'],dayfirst=True, format='%d/%m/%Y')
    df_big['Transaction Amount'] = df_big['Transaction Amount'].str.replace(',','').astype(float)
    df_big['Balance'] = df_big['Balance'].str.replace(',','').astype(float)

    print(df_big.info())
    return df_big
# IDFC
def getcoord2(path,tol_diff):
    fp = open(path, 'rb')
    # Create a PDF parser object associated with the file object.
    parser = PDFParser(fp)
    # Create a PDF document object that stores the document structure.
    # Password for initialization as 2nd parameter
    # password = ''
    document = PDFDocument(parser)
    # Check if the document allows text extraction. If not, abort.
    # Create a PDF resource manager object that stores shared resources.
    rsrcmgr = PDFResourceManager()
    # Create a PDF device object.
    device = PDFDevice(rsrcmgr)
    # BEGIN LAYOUT ANALYSIS
    # Set parameters for analysis.
    laparams = LAParams()
    # Create a PDF page aggregator object.
    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    # Create a PDF interpreter object.
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    output=[]
    def parse_obj(lt_objs):
        # loop over the object list
        for obj in lt_objs:
            # if it's a textbox, print text and location
            if isinstance(obj, pdfminer.layout.LTTextBoxHorizontal):
                # print(obj.bbox[0], obj.bbox[1], obj.get_text().replace('\n', '_'))
                output.append([obj.bbox[0], obj.bbox[1], obj.bbox[2], obj.bbox[3],obj.get_text().replace('\n', '_')])
            # if it's a container, recurse
            elif isinstance(obj, pdfminer.layout.LTFigure):
                parse_obj(obj._objs)
                output.append(parse_obj(obj._objs))
    # loop over all pages in the document
    for page in PDFPage.create_pages(document):
        # read the page into a layout object
        interpreter.process_page(page)
        layout = device.get_result()
        # extract text from this object
        parse_obj(layout._objs)

    x0=0
    x1=0
    x2=0
    x3=0
    x4=0
    x5=0
    x6=0
    x7=0
    # print(output)

    for i in output:
        if i!=None:
            if 'Transaction Date Value Date_' == i[4]:
                x0=i[0] + tol_diff + 70
                x1=i[2] + tol_diff
            if 'Particulars_' == i[4]:
                x2=i[2] + tol_diff + 50
            if 'Cheque _No_' == i[4]:
                x3=i[2] + tol_diff + 10
            if "Debit_" == i[4]:
                x4=i[2] + tol_diff + 10
            if "Credit_" == i[4]:
                x5=i[2] + tol_diff + 5
            # if "Balance_" == i[4]:
            #     x6=i[2] + tol_diff

    print([x0,x1,x2,x3,x4,x5])

    return [x0,x1,x2,x3,x4,x5]

def parse_idfc(path):
    print("File name: ", path)
    tolerance=10
    columns=getcoord2(path,tolerance)

    
    pdf = PyPDF2.PdfReader(open(path,'rb'))
    # pdf.decrypt(b'CONS866823038')
    pages=len(pdf.pages)

    text = pdf.pages[0].extract_text()


    a_num=get_account_number(text)
    arr= []
    for p in range(1,pages+1):

        x=tabula.read_pdf(path,guess=False,stream=False,columns=columns,pages=p,multiple_tables=True,pandas_options={'header':None},encoding= 'unicode_escape')

        for i in x:
            i=i.fillna(0)

            # with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
            #     print(i)
            # print(i)
            for j in range(0,len(i)):
                    flag=0

                    try:
                        sonata(i[0][j])
                        total=i[6][j]

                        flag=1

                        if total==0:
                            flag=0

                    except Exception: pass


                    if flag!=0:

                        date=sonata(i[0][j])
                        # print(date)

                        transaction_value=0

                        transaction_type=""

                        # print(i[4][j])
                        # print(type (i[4][j]))

                        if getnumber(i[4][j])>0 and getnumber(i[5][j]) == 0:
                            transaction_value=getnumber(i[4][j])
                            transaction_type="debit"

                        if getnumber(i[5][j])>0 and getnumber(i[4][j])==0:
                            transaction_value=getnumber(i[5][j])
                            transaction_type="credit"


                        desc = i[2][j]
                        for k in range(1,4):
                            try:
                                val_1=float(i[4][j+k])+float(i[5][j+k])
                                if val_1==0:
                                    desc=desc + i[2][j+k]
                            except Exception: break

                        balance=getnumber(i[6][j])

                        final_res={
                            "Account Number": a_num,
                            "Bank Name":"IDFC Bank",
                            "Balance": float(balance),
                            "Transaction Date": date,
                            "Transaction Amount":float(transaction_value),
                            "Transaction Type": transaction_type,
                            "Description": desc.upper()
                          }
                        
                        # print(final_res)
                        arr.append(final_res)

    # print(arr)
    df = pd.DataFrame(arr)
    print(df.info())
    return df

# AXIS
def parse_axis(filename):
    
    print("file name",filename)
    df = read_pdf(filename,pages="all") #address of pdf file
    # set columns to all df  till second last page
    for page in range(len(df)-2):
        df[page] = df[page].set_axis(df[0].columns, axis=1)
        
    for pg in range(len(df)):
        for x in df[pg].index:
            if str(df[pg].loc[x][0]) == 'nan':
                df[pg].drop(x, inplace = True)
            
    for p in range(len(df)-2):
        df[p].fillna(0, inplace = True)
    
    # import pandas as pd

    df_big = pd.concat(df[0:-2])  #make single df from list of df
    
    new_ind=[]
    inc=0
    for x in range(len(df)-2):

        for i in df[x].index:
            new_ind.append(inc)
            inc+=1
            
    df_big = df_big.set_index(pd.Index(new_ind)) #set new index for df        
    
    df_big.drop(['Chq No','Branch Name','Chq No','Value Date'], axis=1, inplace=True) #unwanted column removed
    df_big["Transaction Type"] = ""
    df_big.loc[df_big["DR/CR"] == 'DR', "transactionType"] = "debit"
    df_big.loc[df_big["DR/CR"] == 'CR', "transactionType"] = "credit"

    df_big.drop('DR/CR',axis=1, inplace=True)
        
    df_big = df_big.set_axis(['Transaction Date','Description', 'Balance', 'Transaction Amount', 'Transaction Type'], axis=1)
    df_big['Bank Name']='AXIS Bank'
    
    removepoint=0
    for x in df_big.index:
#         print(df_big.loc[x][0])
        if(df_big.loc[x][0]=='Sr. No.'):
            removepoint=x
            break
    
    df_big=df_big.loc[0:removepoint-1]
    
    df_big['Transaction Date']=pd.to_datetime(df_big['Transaction Date'],dayfirst=True, format='%d-%m-%Y')
    df_big['Transaction Amount'] = df_big['Transaction Amount'].astype(float)
    df_big['Balance'] = df_big['Balance'].astype(float)
#     df_big['date']=pd.to_datetime(df_big['date'],dayfirst=True)
    print(df_big.info())
    
    return df_big
    
def preprocess_text(sentence):
    sentence = str(sentence)     # Convert sentence to into string
    sentence = sentence.upper()  # Uppercase
    sentence = sentence.replace('{html}',"")  # Replace html to blank
    cleanr = re.compile('<.*?>')   # Special characters
    cleantext = re.sub(cleanr, '', sentence) 
    tokens = word_tokenize(cleantext)
    filtered_words = [w for w in tokens if len(w) > 2 if not w in stopwords.words('english')]  # Words that are not in stop words & occured     more than 2 times
    return " ".join(filtered_words)  # Join the filtered tokens

def postprocess(df_new):
    df_new.loc[(df_new["Category"] == "Direct Expense") & (df_new["Transaction Type"] == "credit"), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Transfer-in/Revenue-other") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Adjustment/reversal") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Agency/Vendor Expense") & (df_new["Transaction Type"] == "credit"), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Bank charges") & (df_new["Transaction Type"] == "credit"), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Cash/cheque deposit") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Investment New") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Investment/FD deposit") & (df_new["Transaction Type"] == "credit"), "Category"] = "Investment/FD redeem"
    df_new.loc[(df_new["Category"] == "Investment/FD redeem") & (df_new["Transaction Type"] == "debit"), "Category"] = "Investment/FD deposit"
    df_new.loc[(df_new["Category"] == "Loan repayment") & (df_new["Transaction Type"] == "credit"), "Category"] = "Loan-in"
    df_new.loc[(df_new["Category"] == "Loan-in") & (df_new["Transaction Type"] == "debit"), "Category"] = "Loan repayment"
    df_new.loc[(df_new["Category"] == "Outward bounce") & (df_new["Transaction Type"] == "credit"), "Category"] = "Inward bounce"
    df_new.loc[(df_new["Category"] == "Inward bounce") & (df_new["Transaction Type"] == "debit"), "Category"] = "Outward bounce"
    df_new.loc[(df_new["Category"] == "Rental expense") & (df_new["Transaction Type"] == "credit"), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Revenue PG Lender Escrow") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Revenue PG Non split") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Revenue UPI") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Salary/Emp/Consultant") & (df_new["Transaction Type"] == "credit"), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Tax") & (df_new["Transaction Type"] == "credit"), "Category"] = "Tax/other-credit"
    df_new.loc[(df_new["Category"] == "Tax/other-credit") & (df_new["Transaction Type"] == "debit"), "Category"] = "Tax"
    df_new.loc[(df_new["Category"] == "Utilities/Bill") & (df_new["Transaction Type"] == "credit"), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Cash Expense") & (df_new["Transaction Type"] == "credit"), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "OD/CC Repayment") & (df_new["Transaction Type"] == "credit"), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Interest income") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Deposit credit") & (df_new["Transaction Type"] == "debit"), "Category"] = "Deposit debit"
    df_new.loc[(df_new["Category"] == "Deposit debit") & (df_new["Transaction Type"] == "credit"), "Category"] = "Deposit credit"
    df_new.loc[(df_new["Category"] == "Revenue POS") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Revenue PG split") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Revenue COD") & (df_new["Transaction Type"] == "debit") & (df_new["Description"].str.strip().str.upper().str.contains("DELHIV", case=False)), "Category"] = "Saas/Tech"
    df_new.loc[(df_new["Category"] == "Revenue COD") & (df_new["Transaction Type"] == "debit") & (df_new["Description"].str.strip().str.upper().str.contains("DELHIV", case=False) == False), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Marketing") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("AMAZ", case=False)), "Category"] = "Revenue Marketplace"
    df_new.loc[(df_new["Category"] == "Marketing") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("AMAZ", case=False) == False), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Saas/Tech") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("AMAZ", case=False)), "Category"] = "Revenue Marketplace"
    df_new.loc[(df_new["Category"] == "Saas/Tech") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("DELHIV", case=False)), "Category"] = "Revenue COD"
    df_new.loc[(df_new["Category"] == "Saas/Tech") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("DELHIV", case=False) == False) & (df_new["Description"].str.strip().str.upper().str.contains("AMAZ") == False), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Revenue Marketplace") & (df_new["Transaction Type"] == "debit") & (df_new["Description"].str.strip().str.upper().str.contains("AMAZ", case=False) == False), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Revenue Marketplace") & (df_new["Transaction Type"] == "debit") & (df_new["Description"].str.strip().str.upper().str.contains("AMAZ", case=False)), "Category"] = "Marketing"
    df_new.loc[(df_new["Category"] == "Revenue Marketplace") & (df_new["Transaction Type"] == "debit") & (df_new["Description"].str.strip().str.upper().str.contains("AMAZONSE", case=False)), "Category"] = "Saas/Tech"
    df_new.loc[(df_new["Category"] == "Revenue Marketplace") & (df_new["Transaction Type"] == "debit") & (df_new["Description"].str.strip().str.upper().str.contains("AMAZ", case=False)), "Category"] = "Marketing"
    df_new.loc[(df_new["Category"] == "Refund") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("AMAZ", case=False)), "Category"] = "Revenue Marketplace"
    df_new.loc[(df_new["Category"] == "Refund") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("AMAZ", case=False) == False), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Nach payment") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("RAZ", case=False)), "Category"] = "Revenue PG split"
    df_new.loc[(df_new["Category"] == "Nach payment") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("LOAN|INCRED|BAJAJ", regex=True, case=False)), "Category"] = "Loan-in"
    df_new.loc[(df_new["Category"] == "Nach payment") & (df_new["Transaction Type"] == "credit") & (df_new["Description"].str.strip().str.upper().str.contains("LOAN|INCRED|BAJAJ|RAZ", regex=True, case=False) == False), "Category"] = "Transfer-in/Revenue-other"
    df_new.loc[(df_new["Category"] == "Revenue Escrow") & (df_new["Transaction Type"] == "debit"), "Category"] = "Direct Expense"
    df_new.loc[(df_new["Category"] == "Trading/Investment") & (df_new["Transaction Type"] == "credit"), "Category"] = "Transfer-in/Revenue-other"
    
    return df_new

def bank_classifier_predict(df):
    # grab environment variables
    tokenizer = AutoTokenizer.from_pretrained('bert-base-cased')

    ENDPOINT_NAME = os.environ['ENDPOINT_NAME']
    
    start = time.time()
    runtime = boto3.client('sagemaker-runtime', region_name='ap-south-1', aws_access_key_id=os.environ["SAGEMAKER_ACCESS_KEY_ID"], aws_secret_access_key=os.environ["SAGEMAKER_SECRET_ACCESS_KEY"])
    print("EndpointName: ", ENDPOINT_NAME)

    print("Running classifier predict...")

    cat_dict = {0: 'Cash/cheque deposit', 1: 'Outward bounce', 2: 'Transfer-in/Revenue-other', 3: 'Revenue UPI', 4: 
        'Direct Expense', 5: 'Bank charges', 6: 'Tax', 7: 'Revenue COD', 8: 'Trading/Investment', 
        9: 'Revenue PG Non split', 10: 'Marketing', 11: 'Utilities/Bill', 12: 'Revenue PG split', 
        13: 'Salary/Emp/Consultant', 14: 'Rental expense', 15: 'Saas/Tech', 16: 'Cash Expense', 
        17: 'Revenue Marketplace', 18: 'Refund', 19: 'Investment New', 20: 'Adjustment/reversal', 
        21: 'Inward bounce', 22: 'Loan-in', 23: 'Agency/Vendor Expense', 24: 'Investment/FD redeem', 
        25: 'Nach payment', 26: 'OD/CC Repayment', 27: 'Interest income', 28: 'Investment/FD deposit', 
        29: 'Revenue Escrow', 30: 'Revenue PG Lender Escrow', 31: 'Loan repayment', 32: 'Tax/other-credit', 
        33: 'Deposit debit', 34: 'Deposit credit', 35: 'Revenue POS'}

    if df.shape[0] > 0:
        df_new = df.dropna()
        df_new["Description"] = df_new["Description"].astype('string')
        print("Info: ", df_new.info())

        df_clean = df_new.copy()
        df_clean["Description"] = df_new["Description"].map(lambda s: preprocess_text(s))

        max_len = 200
        X = tokenizer(
            text=df_clean['Description'].tolist(),
            add_special_tokens=True,
            max_length=max_len,
            truncation=True,
            padding=True,
            return_tensors='tf',
            return_token_type_ids=False,
            return_attention_mask=True,
            verbose=True
        )
        print("X: ", X)

        input_id = X["input_ids"].numpy()
        input_id = json.dumps(input_id.tolist())
        attention_mask = X["attention_mask"].numpy()
        attention_mask = json.dumps(attention_mask.tolist())

        inp_data = []
        for inpt, attn in zip(json.loads(input_id), json.loads(attention_mask)):
            if len(inpt) < max_len:
                diff = max_len - len(inpt)
                n = [0] * diff
                inpt.extend(n)
            if len(attn) < max_len:
                diff = max_len - len(attn)
                n = [0] * diff
                attn.extend(n)
            dic = {'input_ids': inpt, 'attention_mask': attn}
            inp_data.append(dic)
    

        if df_clean.shape[0] > 360:
            predictions = []
            
            num = math.ceil(len(inp_data) / 360)
            print("No. of Batches: ", num)
            json_list = np.array_split(inp_data, num)
            
            for idx, res in enumerate(json_list):
                print("Batch: ", idx + 1)
                json_df = json.dumps(res.tolist())
                # df_parq = res[["Description"]].to_parquet()
                # res["Description"] = res["Description"].astype('string')
                start1 = time.time()
                response = runtime.invoke_endpoint(EndpointName=ENDPOINT_NAME, Body=json_df, ContentType='application/json')
                # print("Response: ", response)
                end1 = time.time()
                print("Endpoint time taken: {}".format(end1 - start1))

                result = json.loads(response['Body'].read().decode())
                # print("Result: ", result)
                        
                final_predicted = [cat_dict[val] for val in result["Predictions"]]
                predictions.extend(final_predicted)
                
            df_new["Category"] = predictions
            df_new = postprocess(df_new)
            df_new.to_csv("bank_data.csv", index=False)

            end = time.time()
            print("Successfully predicted...Time taken: {} secs".format(end - start))
            return "bank_data.csv"
        else:
            print("Batches: 1")
            # df_new["Description"] = df_new["Description"].astype('string')
            # df_parq = df_new.to_parquet()
            json_df = json.dumps(inp_data)
            start1 = time.time()
            response = runtime.invoke_endpoint(EndpointName=ENDPOINT_NAME, Body=json_df, ContentType='application/json')
            # print("Response: ", response)
            end1 = time.time()
            print("Endpoint time taken: {}".format(end1 - start1))

            result = json.loads(response['Body'].read().decode())
            # print("Result: ", result)
                    
            final_predicted = [cat_dict[val] for val in result["Predictions"]]
            # ids = payload_df["_id"].values.tolist()
            
            df_new["Category"] = final_predicted
            df_new = postprocess(df_new)
            df_new.to_csv("bank_data.csv", index=False)

            end = time.time()
            print("Successfully predicted...Time taken: {} secs".format(end - start))
            return "bank_data.csv"
    else:
        return "No data found..."