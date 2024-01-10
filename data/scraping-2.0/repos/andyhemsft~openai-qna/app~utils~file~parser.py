import os
import logging
from openpyxl import load_workbook
from abc import abstractmethod

from langchain.document_loaders import PyPDFLoader
from app.utils.file.storage import BlobStorageClient
from app.config import Config

logger = logging.getLogger(__name__)

class Parser():
    """This class represents a Parser."""

    def __init__(self, config: Config):
        """
        Initialize the Parser.

        Args:
            config: the config object
        """

        self.config = config

    @abstractmethod
    def analyze_read(self, source_url: str) -> str:
        """
        Analyze and read source url.

        Args:
            source_url: the source url
        Returns:
            the parsed text
        """
        

    def write(self, dest_url: str, text: str):
        """
        Write the parsed text to a file.

        Args:
            dest_url: the destination url
            text: the parsed text
        """

        # Write to local file
        if self.config.DOCUMENT_DEST_LOCATION == 'local':
            with open(dest_url, 'w', encoding='utf-8') as f:
                f.write(text)

        # Write to Azure Blob Storage
        elif self.config.DOCUMENT_DEST_LOCATION == 'azure':
            blob_storage_client = BlobStorageClient(self.config)
            blob_storage_client.upload_blob(text, dest_url)

        return None

def get_parser(document_type: str) -> Parser:
    """This function returns a parser based on the type."""

    config = Config()
    if document_type == 'pdf':
        return PDFParser(config=config)
    
    elif document_type == 'excel':
        return ExcelParser(config=config)

    else:
        raise ValueError('Parser type not supported')

class ExcelParser(Parser):
    """This class represents an Excel Parser."""

    def __init__(self, config: Config):
        super().__init__(config)

    def analyze_read(self, source_url: str) -> str:
        """
        Analyze and read source url for an Excel using openpyxl.
        Unmerging the cells and assigning the value as top-left cell.

        Args:
            source_url: the source url
        Returns:
            the parsed text
        """

        wb = load_workbook(source_url)

        sheets = wb.sheetnames  ##['Sheet1', 'Sheet2']
        for i, sheet in enumerate(sheets):
            ws = wb[sheets[i]]
            
            # you need a separate list to iterate on (see explanation #2 below)
            mergedcells =[]  
            for group in ws.merged_cells.ranges:
                mergedcells.append(group)
            
            for group in mergedcells:
                min_col, min_row, max_col, max_row = group.bounds 
                top_left_cell_value = ws.cell(row=min_row, column=min_col).value
                ws.unmerge_cells(str(group))   # you need to unmerge before writing (see explanation #1 below)
                for irow in range(min_row, max_row+1):
                    for jcol in range(min_col, max_col+1): 
                        ws.cell(row = irow, column = jcol, value = top_left_cell_value)
        
        content = ''
        for i, sheet in enumerate(sheets):
            ws = wb[sheets[i]]
            
            for row in ws.values:

                l = list(row)
                for i in range(len(l)):
                    if l[i] == None:
                        continue
                    if i == len(l) - 1:
                        content = content + (str(l[i]).replace('\n', ' '))
                        break
                    else:
                        content = content + (str(l[i]).replace('\n', ' ') + ' | ')
                content = content + '\n'
            
        return content
    
    

class PDFParser(Parser):
    """This class represents a PDF Parser."""

    def __init__(self, config: Config):
        """
        Initialize the PDF Parser.

        Args:
            config: the config object
        """

        self.config = config

        if self.config.PDF_PARSER_TYPE == 'pdfloader':
            pass

        elif self.config.PDF_PARSER_TYPE == 'formrecognizer':
            raise ValueError('Form Recognizer not supported yet')

        else:
            raise ValueError('PDF Parser type not supported')
        
    def _extract_tables(self, source_url: str) -> str:
        """Extract table from source url.
        
        Args:
            source_url: the source url
        Returns:
            the tables
        """

        # # tables = tabula.read_pdf(source_url, pages="all")
        # tables = camelot.read_pdf(source_url)

        # logger.info("Total tables extracted: {}".format(len(tables)))

        # for i, table in enumerate(tables, start=1):
        #     # logger.info(f'Table {i}:\n{table.to_string()}')
        #     logger.info(f'Table {i}:\n{table.df.to_string()}')

        return "Not implemented yet"


    def analyze_read(self, source_url: str) -> str:
        """
        Analyze and read source url for a PDF.

        Args:
            source_url: the source url
        Returns:
            the parsed text
        """

        # Load PDF using langchain pdfloader
        if self.config.PDF_PARSER_TYPE == 'pdfloader':
            self.loader = PyPDFLoader(source_url)
            pages = self.loader.load()
  
            # Convert to UTF-8 encoding for non-ascii text
            for page in pages:
                try:
                    if page.page_content.encode("iso-8859-1") == page.page_content.encode("latin-1"):
                        page.page_content = page.page_content.encode("iso-8859-1").decode("utf-8", errors="ignore")
                except:
                    pass
            
            # Concatenate all pages
            text = '\n'.join([page.page_content for page in pages])

            return text

        elif self.config.PDF_PARSER_TYPE == 'formrecognizer':
            raise ValueError('Form Recognizer not supported yet')


