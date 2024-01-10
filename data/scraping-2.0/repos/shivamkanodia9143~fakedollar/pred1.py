import openai
import openpyxl

from flask import Flask, request

# Set up OpenAI API key
openai.api_key = "YOUR_API_KEY_HERE"

# Initialize Flask app
app = Flask(__name__)


# Define endpoint for handling POST requests
@app.route('/predict', methods=['POST'])
def predict():
    # Load input Excel files
    tasks_file = request.files['tasks_file']
    activities_file = request.files['activities_file']
    tasks_workbook = openpyxl.load_workbook(tasks_file)
    activities_workbook = openpyxl.load_workbook(activities_file)
    tasks_sheet = tasks_workbook.active
    activities_sheet = activities_workbook.active

    # Extract data from input Excel files
    tasks = []
    for row in tasks_sheet.iter_rows(min_row=2, values_only=True):
        tasks.append({
            'element_name': row[0],
            'description': row[1],
            'element_category': row[2],
            'section': row[3],
            'item_type': row[4]
        })

    activities = []
    for row in activities_sheet.iter_rows(min_row=2, values_only=True):
        activities.append({
            'activity': row[0],
            'category': row[1]
        })

    # Generate prompts for each task and activity combination
    prompt_responses = []
    for task in tasks:
        for activity in activities:
            prompt = f"Which activity does the task '{task['element_name']}' belong to? Activity: '{activity['activity']}' Category: '{activity['category']}'"
            response = openai.Completion.create(engine="davinci", prompt=prompt, max_tokens=1024)
            prompt_responses.append({
                'task_element_name': task['element_name'],
                'task_description': task['description'],
                'task_element_category': task['element_category'],
                'task_section': task['section'],
                'task_item_type': task['item_type'],
                'activity': activity['activity'],
                'category': activity['category'],
                'response': response.choices[0].text.strip()
            })

    # Generate output Excel file containing mapping of tasks to activities
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.append([
        'Task Element Name',
        'Task Description',
        'Task Element Category',
        'Task Section',
        'Task Item Type',
        'Activity',
        'Category',
        'Prediction'
    ])
    for prompt_response in prompt_responses:
        output_sheet.append([
            prompt_response['task_element_name'],
            prompt_response['task_description'],
            prompt_response['task_element_category'],
            prompt_response['task_section'],
            prompt_response['task_item_type'],
            prompt_response['activity'],
            prompt_response['category'],
            prompt_response['response']
        ])

    output_file = 'output.xlsx'
    output_workbook.save(output_file)

    return f'Output saved to {output_file}'


# Start Flask app
if __name__ == '__main__':
    app.run()
