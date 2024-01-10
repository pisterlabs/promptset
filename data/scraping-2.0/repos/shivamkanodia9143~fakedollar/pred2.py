import openai
import openpyxl

# Set up OpenAI API credentials
openai.api_key = "YOUR_API_KEY_HERE"

# Load the ChatGPT model
model = "text-davinci-002"
prompt = "Generate a mapping of elements to activities in retail fitouts."

# Define a function to generate the prompt
def generate_prompt(element_names, element_descs, element_cats, element_sections, element_item_types, activity_names, activity_cats):
    prompt_template = "An element in the construction of retail fitouts contains the name and a description that specifies how to install the element, and also has a category. These elements are grouped in several activities to be performed, each activity can contain one or more elements based on the element's 'name', 'description', 'category', 'section', and 'item type'.\n\nPlease generate a mapping of elements to activities based on the following information:\n\n"
    prompt_template += f"Element {1}:\n- Name: {element_names[0]}\n- Description: {element_descs[0]}\n- Category: {element_cats[0]}\n- Section: {element_sections[0]}\n- Item Type: {element_item_types[0]}\n\n"
    prompt_template += "List of activities:\n"
    for i in range(len(activity_names)):
        prompt_template += f"- Activity {i+1}: {activity_names[i]}, Category: {activity_cats[i]}\n"
    prompt_template += "\nGenerate the mapping of elements to activities."
    return prompt_template

# Define a function to generate the output mapping
def generate_mapping(prompt):
    # Generate the response from OpenAI's API
    response = openai.Completion.create(
        engine=model,
        prompt=prompt,
        max_tokens=2048,
        n=1,
        stop=None,
        temperature=0.5,
    )
    # Extract the generated mapping from the response
    mapping = response.choices[0].text.strip()
    # Return the mapping as a string
    return mapping

# Define a function to load Excel files and extract the necessary data
def load_data(element_file, activity_file):
    # Load the Excel files using openpyxl
    element_wb = openpyxl.load_workbook(element_file)
    activity_wb = openpyxl.load_workbook(activity_file)
    # Extract the necessary data from the Excel files
    element_names = []
    element_descs = []
    element_cats = []
    element_sections = []
    element_item_types = []
    for row in element_wb.active.iter_rows(min_row=2, values_only=True):
        element_names.append(row[0])
        element_descs.append(row[1])
        element_cats.append(row[2])
        element_sections.append(row[3])
        element_item_types.append(row[4])
    activity_names = []
    activity_cats = []
    for row in activity_wb.active.iter_rows(min_row=2, values_only=True):
        activity_names.append(row[0])
        activity_cats.append(row[1])
    # Return the extracted data as a tuple
    return (element_names, element_descs, element_cats, element_sections, element_item_types, activity_names, activity_cats)

# Define the Flask app and endpoints
from flask import Flask, request, jsonify

app = Flask(__name__)

@app.route('/generate_mapping', methods=['POST'])
def generate_mapping_endpoint():
    # Load the input files and extract the necessary data
    if 'tasks_file' and 'activities_file' not in request.files:
        return 'No tasks_file and activities_file file provided', 400

    file = request.files['tasks_file']
    if file.filename == '':
        return 'No file provided', 400

    if not file.filename.endswith('.xlsx'):
        return 'Invalid file format', 400

    # Load input Excel files
    tasks_file = request.files['tasks_file']
    activities_file = request.files['activities_file']
    data = load_data(tasks_file, activities_file)
    prompt = generate_prompt(data)
    mapping = generate_mapping(prompt)
