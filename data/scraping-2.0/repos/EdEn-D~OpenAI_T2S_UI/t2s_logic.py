import os
import openpyxl
import csv
from pathlib import Path
from openai import OpenAI
from pydub import AudioSegment
from dotenv import load_dotenv, find_dotenv


load_dotenv(find_dotenv())

client = OpenAI(api_key="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")

import openpyxl
import os

class T2SConverter:
    def __init__(self, file_path, output_dir, callback_func):
        self.file_path = file_path
        self.output_dir = output_dir
        self.callback_func = callback_func

    @staticmethod
    def load_excel_rows(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        rows = []

        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            description, text_for_speech = row[0], row[1]
            if description and text_for_speech:  # Check if both cells are not empty
                rows.append(description)

        return rows


    # Create new folder based on filename and voice
    def process_rows(self, selected_rows, selected_option):
        # Extract the base filename without extension
        base_name = os.path.splitext(os.path.basename(self.file_path))[0]

        # Create a new directory name combining base name and selected option
        new_dir_name = f"{base_name}_{selected_option}"
        new_dir_path = os.path.join(self.output_dir, new_dir_name)

        if not os.path.exists(new_dir_path):
            os.makedirs(new_dir_path, exist_ok=True)
            self.callback_func(f"Directory '{new_dir_path}' has been created.")

        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active

        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_num in selected_rows:
                description, text_for_speech = row[0], row[1]
                if description and text_for_speech:  # Check if both cells are not empty
                    self.convert_to_audio(new_dir_path, description, text_for_speech, selected_option, row_num)
        self.callback_func("Finished...")

    def convert_to_audio(self, path, desc, t2s, voice, row):
        # print(out_dir, row, desc, t2s, voice)
        response = client.audio.speech.create(
            model="tts-1",
            voice=voice,
            input=t2s,
            speed=0.85,
        )
        speech_file_path = os.path.join(path, f"{row} {desc}.mp3")
        response.stream_to_file(speech_file_path)
        add_silence_to_mp3(speech_file_path)
        # convert_mp3_to_wav(speech_file_path)


# def get_t2s_from_file(file_path, output_dir, selected_option, callback):
#     # Extract the base filename without extension
#     base_name = os.path.splitext(os.path.basename(file_path))[0]
#
#     # Create a new directory name combining base name and selected option
#     new_dir_name = f"{base_name}_{selected_option}"
#     new_dir_path = os.path.join(output_dir, new_dir_name)
#
#     os.makedirs(new_dir_path, exist_ok=True)
#     callback(f"Directory '{new_dir_path}' has been created.")
#
#     # Check the file extension to determine the action
#     if file_path.endswith(('.xlsx', '.xls')):
#         process_excel_file(file_path, new_dir_path, selected_option, callback)
#     elif file_path.endswith('.csv'):
#         process_csv_file(file_path, new_dir_path, selected_option, callback)
#     elif file_path.endswith('.txt'):
#         print_txt_file_contents(file_path, selected_option)

# def process_excel_file(file_path, output_dir, selected_option, callback):
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active
#
#     for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
#         description, text_for_speech = row[0], row[1]
#         if description and text_for_speech:  # Check if both cells are not empty
#             get_t2s(output_dir, row_num, description, text_for_speech, selected_option)
#     callback("Finished...")

# def process_csv_file(file_path, output_dir, selected_option, callback):
#     with open(file_path, newline='') as csvfile:
#         reader = csv.reader(csvfile)
#         for row_num, row in enumerate(reader, start=1):
#             description, text_for_speech = row[0], row[1]
#             if description and text_for_speech:  # Check if both cells are not empty
#                 get_t2s(output_dir, row_num, description, text_for_speech, selected_option)

def print_txt_file_contents(file_path):
    # Read and print the contents of the TXT file
    with open(file_path, 'r') as file:
        contents = file.read()
        print(contents)

# def get_t2s(out_dir, row, desc, t2s, voice):
#     print(out_dir, row, desc, t2s, voice)
#     response = client.audio.speech.create(
#         model="tts-1",
#         voice=voice,
#         input=t2s,
#         speed=0.85,
#     )
#     speech_file_path = os.path.join(out_dir, f"{row} {desc}.mp3")
#     response.stream_to_file(speech_file_path)
#     convert_mp3_to_wav(speech_file_path)


def convert_mp3_to_wav(input_file):
    # Load the mp3 file
    audio = AudioSegment.from_mp3(input_file)

    # Add 0.5 seconds of silence at the beginning
    silence = AudioSegment.silent(duration=500)  # 500 milliseconds
    audio_with_silence = silence + audio

    # Construct the output file path (same directory as the input file)
    output_file = os.path.splitext(input_file)[0] + ".wav"

    # Convert to wav with desired attributes (8kHz, 16-bit, mono)
    audio_with_silence.set_frame_rate(8000).set_sample_width(2).set_channels(1).export(output_file, format="wav")

    # Delete the original mp3 file
    os.remove(input_file)

    return output_file

from pydub import AudioSegment

def add_silence_to_mp3(file_path):
    # Load the mp3 file
    audio = AudioSegment.from_mp3(file_path)

    # Create 0.5 seconds of silence
    silence = AudioSegment.silent(duration=500)  # Duration is in milliseconds

    # Add silence to the beginning of the audio
    audio_with_silence = silence + audio

    # Overwrite the original file with the new audio
    audio_with_silence.export(file_path, format="mp3")
