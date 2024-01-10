"""
Created on Sat May 13 19:00:28 2023

@author: David Tovar
"""
import os
import fitz
import re
import numpy as np
import tensorflow_hub as hub
import openai
from sklearn.neighbors import NearestNeighbors
import pandas as pd
from tkinter import Tk, simpledialog, filedialog, Entry, Label, Button, StringVar
import sys
import PyPDF2
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import time
from api_key_checker import check_api_key
from openai.error import RateLimitError,OpenAIError


# Set your OpenAI API Key here
check_api_key("OpenAI_API.txt", "OpenAI")
with open("OpenAI_API.txt", "r") as f:
    openai_api_key = f.read().strip()
    
openai.api_key = openai_api_key

def preprocess(text):
    text = text.replace('\n', ' ')
    text = re.sub('\s+', ' ', text)
    return text


def pdf_to_text(path, start_page=1, end_page=None):
    doc = fitz.open(path)
    total_pages = doc.page_count

    if end_page is None:
        end_page = total_pages

    text_list = []

    for i in range(start_page-1, end_page):
        text = doc.load_page(i).get_text("text")
        text = preprocess(text)
        text_list.append(text)

    doc.close()
    return text_list

def extract_apa_citation(path):
    with open(path, 'rb') as f:
        pdf = PyPDF2.PdfReader(f)
        meta = pdf.metadata

    if meta is None:
        return "No APA citation found."

    title = meta.title if meta.title else "Unknown title"
    author = meta.author if meta.author else "Unknown author"
    journal = meta.creator if meta.creator else "Unknown journal"
    year = meta.get('/CreationDate')[2:6] if meta.get('/CreationDate') else "Unknown year"

    citation = f"{author} ({year}). {title}. {journal}."

    return citation
def text_to_chunks(texts, word_length=150, start_page=1):
    text_toks = [t.split(' ') for t in texts]
    page_nums = []
    chunks = []
    
    for idx, words in enumerate(text_toks):
        for i in range(0, len(words), word_length):
            chunk = words[i:i+word_length]
            if (i+word_length) > len(words) and (len(chunk) < word_length) and (
                len(text_toks) != (idx+1)):
                text_toks[idx+1] = chunk + text_toks[idx+1]
                continue
            chunk = ' '.join(chunk).strip()
            chunk = f'[Page no. {idx+start_page}]' + ' ' + '"' + chunk + '"'
            chunks.append(chunk)
    return chunks

class SemanticSearch:
    
    def __init__(self):
        self.use = hub.load('https://tfhub.dev/google/universal-sentence-encoder/4')
        self.fitted = False
    
    
    def fit(self, data, batch=1000, n_neighbors=5):
        self.data = data
        self.embeddings = self.get_text_embedding(data, batch=batch)
        n_neighbors = min(n_neighbors, len(self.embeddings))
        self.nn = NearestNeighbors(n_neighbors=n_neighbors)
        self.nn.fit(self.embeddings)
        self.fitted = True
    
    
    def __call__(self, text, return_data=True):
        inp_emb = self.use([text])
        neighbors = self.nn.kneighbors(inp_emb, return_distance=False)[0]
        
        if return_data:
            return [self.data[i] for i in neighbors]
        else:
            return neighbors
    
    
    def get_text_embedding(self, texts, batch=1000):
        embeddings = []
        for i in range(0, len(texts), batch):
            text_batch = texts[i:(i+batch)]
            emb_batch = self.use(text_batch)
            embeddings.append(emb_batch)
        embeddings = np.vstack(embeddings)
        return embeddings



def load_recommender(path, start_page=1):
    global recommender
    texts = pdf_to_text(path, start_page=start_page)
    chunks = text_to_chunks(texts, start_page=start_page)
    recommender.fit(chunks)
    return 'Corpus Loaded.'


def generate_text(prompt, engine="gpt-3.5-turbo"):
    max_attempts = 500
    sleep_time = 30 # Number of seconds to wait between attempts

    for attempt in range(max_attempts):
        try:
            if engine == "gpt-3.5-turbo":
                completions = openai.ChatCompletion.create(
                    model=engine,
                    messages=[
                        {"role": "system", "content": "You are an expert scientific writer who writes review articles for Nature Neuroscience Reviews."},
                        {"role": "user", "content": prompt}
                    ],
                    max_tokens = 300,
                    temperature=0.7
                )
            else:
                completions = openai.Completion.create(
                    engine=engine,
                    prompt=prompt,
                    max_tokens=700,
                    n=1,
                    stop=None,
                    temperature=0.7,
                )
            message = completions.choices[0].text.strip() if engine != "gpt-3.5-turbo" else completions.choices[0].message['content'].strip()
            return message

        except (RateLimitError, OpenAIError) as e:
            if attempt + 1 == max_attempts:
                raise e  # If this was the last attempt, raise the exception
            else:
                time.sleep(sleep_time)  # Wait for a while before trying again



def generate_answer(question):
    topn_chunks = recommender(question)
    prompt = ""
    prompt += 'search results:\n\n'
    for c in topn_chunks:
        prompt += c + '\n\n'
    
    prompt += f"For the above text, {question}. Compose a comprehensive reply to the question using the search results given. "\
              "Cite each reference using [ Page Number] notation at the end of each sentence (every result has this number at the beginning). "\
              "Citation should be done at the end of each sentence. If the search results mention multiple subjects "\
              "with the same name, create separate answers for each. Only include information found in the results and "\
              "don't add any additional information. Make sure the answer is correct and don't output false content. "\
              "If the text does not relate to the query, simply state 'Text Not Found in PDF'. Ignore outlier "\
              "search results which has nothing to do with the question. Only answer what is asked. The "\
              "answer should be short and concise, less than 200 words. Answer step-by-step."
    print("Calling GPT3 API")
    answer = generate_text(prompt)
    return answer



recommender = SemanticSearch()


def get_pdf_file_list(folder_path):
    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]
    return pdf_files


class QuestionDialog(simpledialog.Dialog):
    def body(self, master):
        self.question = StringVar()
        self.entry = Entry(master, textvariable=self.question)
        self.entry.grid(row=0, column=0)
        return self.entry  # initial focus

    def apply(self):
        self.result = self.question.get()
        

def process_pdfs(manual_input=False):  
    root = Tk()
    root.title("PDF Questions")

    base_folder_path = filedialog.askdirectory(title='Select Folder with PDFs')

    if not base_folder_path:
        print("No folder selected. Exiting.")
        sys.exit(1)

    if manual_input:
        Label(root, text="Enter questions:").grid(row=0, column=0)
        entries = []
        for i in range(3):  
            dialog = QuestionDialog(root, title="Question {}".format(i+1))
            if dialog.result:
                entry = Entry(root)
                entry.grid(row=i+1, column=0)
                entries.append(entry)
                entry.insert(0, dialog.result)

        def on_submit():
            nonlocal questions
            questions = [entry.get() for entry in entries if entry.get().strip()]
            root.destroy()

        submit_button = Button(root, text="Submit", command=on_submit)
        submit_button.grid(row=4, column=0)
        root.mainloop()

        questions = [question.strip() for question in questions if question.strip()]

    else:
        questions = ["What is the summary of the introduction?", 
                     "What methods are used in the study", 
                     "What are the main results found in the study"]

    print("Question Asked.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  

    for folder_path, subfolders, files in os.walk(base_folder_path):
        summary_records = []  
        pdf_files = [f for f in files if f.lower().endswith('.pdf')]

        if not pdf_files:  # If no PDF files found, continue to the next folder
            continue

        for file_name in pdf_files:
            pdf_file_path = os.path.join(folder_path, file_name)
            load_recommender(pdf_file_path)
            citation = f"{file_name[:-4]}"
            record = [citation]

            for question in questions:
                answer = generate_answer(question).strip()
                record.append(answer)

            summary_records.append(record)
            print("Question Complete.")

        columns = ['Citation'] + questions
        df = pd.DataFrame(summary_records, columns=columns)

        subfolder_name = os.path.basename(os.path.normpath(folder_path))

        ws = wb.create_sheet(title=subfolder_name)  

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for ws in wb.worksheets:
            for col in list(ws.columns)[0:]:
                ws.column_dimensions[get_column_letter(col[0].column)].width = 50  

            for row in ws:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

    folder_name = os.path.basename(os.path.normpath(base_folder_path))
    output_file = os.path.join(base_folder_path, f'{folder_name}_question_answers.xlsx')
    
    wb.save(output_file)
    
    print(f"Question answers saved to: {output_file}")
    return output_file


if __name__ == "__main__":
    result = process_pdfs()
    print(f"Resulting file is {result}")
