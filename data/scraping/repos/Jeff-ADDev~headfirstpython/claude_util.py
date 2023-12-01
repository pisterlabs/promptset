from typing import List
from objects.issue import Issue
from colorama import init, Fore, Back, Style
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter, quote_sheetname
from anthropic import Anthropic, HUMAN_PROMPT, AI_PROMPT

def excel_worksheet_ai_create(ws, epics, jira_issue_link, claude_ai_key):
    # Evaluating the description and comments, please answer the following questions:
    # 1. What is the overall review of the epic?

    def claude_examine_epic_questions(description, comments, api_key):
        print("YEP ER")
        anthropic = Anthropic(
        # defaults to os.environ.get("ANTHROPIC_API_KEY")
        api_key=api_key,
        )
        completion = anthropic.completions.create(
            model="claude-2",
            max_tokens_to_sample=50000,
            prompt=f"{HUMAN_PROMPT}" + 
            
            """You have been given a description from a Jira Epic along with any comments in that epic.
            The epic description has formatting data in it, please ignore that formatting data and just read the text.
            Evaluating the description and comments, please answer, is there any questions that are outstanding?
            """ +
            "Description: " + description +
            "Comments: " + comments +
            f"{AI_PROMPT}",
        )
        return(completion.completion)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 40

    table = Table(displayName="TableAIData", ref="A1:E" + str(len(epics) + 1))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.append(["Epic", "Questions", "Oveall Review","Description", "Comments"])

    for epicitem in epics:
        comments = ""
        for comment in epicitem.comments:
            comments += comment + "\n"
        questions = claude_examine_epic_questions(epicitem.description, comments, claude_ai_key)
        ws.append([epicitem.key, questions, " - ", epicitem.description, comments])    

    ws.add_table(table)

    for row in ws[2:ws.max_row]:  # Exclude The Header
        cell = row[0] # zeor based index
        value_use = cell.value
        cell.hyperlink = f"{jira_issue_link}{value_use}"
        cell.value = value_use
        cell.style = "Hyperlink"
    
    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[0] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[2:ws.max_row]:  # skip the header
        cell = row[1] # zeor based index
        cell.alignment = Alignment(wrap_text=True)
        cell.number_format = "text"

    for row in ws[2:ws.max_row]:  # skip the header
        cell = row[2] # zeor based index
        cell.alignment = Alignment(wrap_text=True)
        cell.number_format = "text"
    
    for row in ws[2:ws.max_row]:  # skip the header
        cell = row[3] # zeor based index
        cell.alignment = Alignment(wrap_text=True)
        cell.number_format = "text"

    for row in ws[2:ws.max_row]:  # skip the header
        cell = row[4] # zeor based index
        cell.alignment = Alignment(wrap_text=True)
        cell.number_format = "text"