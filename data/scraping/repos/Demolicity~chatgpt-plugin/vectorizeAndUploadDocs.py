import os
import sys
import uuid
import docx
import PyPDF2
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import numpy as np
from langchain.embeddings.openai import OpenAIEmbeddings
from langchain.text_splitter import CharacterTextSplitter
from langchain.vectorstores import ElasticVectorSearch, Pinecone, Weaviate, FAISS
import os
from PyPDF2 import PdfReader
import pinecone as pc
import numpy as np

def extract_text_from_file(file_path):
    file_ext = os.path.splitext(file_path)[-1].lower()
    text = ""

    if file_ext == ".txt":
        with open(file_path, "r") as file:
            text = file.read()
    elif file_ext == ".pdf":
        with open(file_path, "rb") as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num in range(len(pdf_reader.pages)):
                text += pdf_reader.pages[page_num].extract_text()
    elif file_ext in (".doc", ".docx"):
        doc = docx.Document(file_path)
        text = "\n".join([para.text for para in doc.paragraphs])
    elif file_ext == ".xls" or file_ext == ".xlsx":
        workbook = load_workbook(file_path, read_only=True)
        for sheet in workbook:
            for row in sheet.iter_rows(values_only=True):
                text += " ".join([str(cell) for cell in row if cell is not None])
    elif file_ext == ".csv":
        df = pd.read_csv(file_path)
        text = " ".join(df.to_string(index=False, header=False).split())
    elif file_ext in (".html", ".css", ".js", ".py", ".cpp", ".java"):
        with open(file_path, "r") as file:
            soup = BeautifulSoup(file.read(), "html.parser")
            text = soup.get_text()

    return text

def upsert_data(datastore, texts, embeddings):
    if (datastore == "Pinecone"):
        pineconeAPIKey = os.environ["PINECONE_API_KEY"]

        # Get the Pinecone Environment from the environment
        pineconeEnvironment = os.environ["PINECONE_ENVIRONMENT"]

        # Get the Pinecone Index from the environment
        pineconeIndex = os.environ["PINECONE_INDEX"]

        pc.init(api_key=pineconeAPIKey, environment=pineconeEnvironment)
        index = pc.Index(pineconeIndex)

        pinecone = Pinecone(index, embeddings.embed_query, "text")

        pinecone.add_texts(texts)

def main(datastore, docs_path):
    OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]
    raw_text = ""
    for root, _, files in os.walk(docs_path):
        for file in files:
            file_path = os.path.join(root, file)
            text = extract_text_from_file(file_path)
            raw_text += "\n" + text

    text_splitter = CharacterTextSplitter(
        separator = '\n',
        chunk_size = 2500,
        chunk_overlap = 200,
        length_function = len,
    )

    texts = text_splitter.split_text(raw_text)

    embeddings = OpenAIEmbeddings(openai_api_key=OPENAI_API_KEY)  

    upsert_data(datastore, texts, embeddings)

if __name__ == "__main__":
    datastore = sys.argv[1]
    docs_path = sys.argv[2]
    print("Vectorizing and uploading documents...")
    print("Datastore: ", datastore)
    print("Docs Path: ", docs_path)
    main(datastore, docs_path)
